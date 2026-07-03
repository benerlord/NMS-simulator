"""接口 Excel 导入/导出的内部工具（编解码 + 校验，不直接触碰 DB）。"""
import json as _json
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill


HEADER_COLUMNS = ["name", "required", "expectValue", "example", "description"]
QUERY_COLUMNS = ["name", "type", "required", "example", "description"]
BODY_COLUMNS = ["contentType", "required", "example", "description"]
PARAM_MAPPING_COLUMNS = ["name", "in", "type", "required", "bindTo"]

MAIN_HEADERS = [
    "方法", "路径", "接口名", "启用", "分类", "分组", "数据源", "拓扑",
    "鉴权类型", "鉴权头名",
    "请求头", "Query 参数", "请求体",
    "SQL 语句", "响应模板", "参数映射",
    "静态响应体",
    "故障-延迟毫秒", "故障-错误率", "故障-错误状态码",
]

INSTRUCTION_SHEET_NAME = "_使用说明"
UNCATEGORIZED_SHEET_NAME = "未归类"
SHEET_NAME_MAX_LEN = 31
INVALID_SHEET_CHARS = ":\\/?*[]"

FIELD_SEP = "|"
RECORD_SEP = "\n"


class ExcelValidationError(Exception):
    """行级校验错误——被上层 parse_workbook 捕获并塞进 errors 列表。"""


def _format_field(value: Any) -> str:
    """把 Python 值渲染成单元格里的一段字段字符串。"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    return str(value)


def format_cell_list(records: list[dict], columns: list[str]) -> str:
    """把结构化记录列表序列化成“多行 + 竖线分隔”的单元格文本。

    空列表 -> 空字符串（不是 "|||"）。
    """
    if not records:
        return ""
    lines: list[str] = []
    for rec in records:
        fields = [_format_field(rec.get(col)) for col in columns]
        lines.append(FIELD_SEP.join(fields))
    return RECORD_SEP.join(lines)


_TRUE_TOKENS = {"是", "true", "1", "yes", "y"}
_FALSE_TOKENS = {"否", "false", "0", "no", "n", ""}


def _is_boolean_column(column: str) -> bool:
    """哪些列在解析时按 bool 处理。"""
    return column == "required"


def _parse_field(value: str, column: str) -> Any:
    """把单元格里的字段字符串解回 Python 值。"""
    stripped = value.strip()
    if _is_boolean_column(column):
        lower = stripped.lower()
        if lower in _TRUE_TOKENS:
            return True
        if lower in _FALSE_TOKENS:
            return False
        raise ExcelValidationError(f"必填列的值 '{value}' 无法识别（应为 是/否）")
    return stripped


def parse_cell_list(cell_text: str, columns: list[str], row_hint: str) -> list[dict]:
    """把“多行 + 竖线分隔”的单元格文本解析回记录列表。

    row_hint 用来在报错时给出 sheet/行/列的位置提示，如 "Sheet '网管A' 第 3 行 请求头"。
    """
    if cell_text is None or cell_text == "":
        return []
    records: list[dict] = []
    for line in cell_text.split(RECORD_SEP):
        line = line.strip()
        if not line:
            continue
        fields = line.split(FIELD_SEP)
        if len(fields) > len(columns):
            raise ExcelValidationError(
                f"{row_hint}：字段数 {len(fields)} 超过预期 {len(columns)}（列顺序：{columns}）"
            )
        while len(fields) < len(columns):
            fields.append("")
        record = {}
        for col, raw in zip(columns, fields):
            try:
                record[col] = _parse_field(raw, col)
            except ExcelValidationError as e:
                raise ExcelValidationError(f"{row_hint}：{e}") from e
        records.append(record)
    return records


def sanitize_sheet_name(name: str, used: set[str]) -> str:
    """把 domain 名清洗成 Excel 合法 sheet 名并防重名。

    - 非法字符（`: \\ / ? * [ ]`）替换为下划线
    - 超过 31 字符从右截断
    - 与 used 中已存在的名字冲突时追加 `~2`、`~3`... 后缀，最终长度仍 ≤ 31
    - 每次调用会把返回的名字加入 used

    副作用：修改传入的 used 集合。
    """
    sanitized = name
    for ch in INVALID_SHEET_CHARS:
        sanitized = sanitized.replace(ch, "_")
    if len(sanitized) > SHEET_NAME_MAX_LEN:
        sanitized = sanitized[:SHEET_NAME_MAX_LEN]
    if not sanitized.strip():
        sanitized = "未命名"

    if sanitized not in used:
        used.add(sanitized)
        return sanitized

    suffix = 2
    while True:
        tag = f"~{suffix}"
        base_max = SHEET_NAME_MAX_LEN - len(tag)
        candidate = sanitized[:base_max] + tag
        if candidate not in used:
            used.add(candidate)
            return candidate
        suffix += 1


# 表头样式
_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
_WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

# 每列的期望宽度（字符数）
_COLUMN_WIDTHS = {
    "方法": 10, "路径": 40, "接口名": 18, "启用": 8, "分类": 12, "分组": 12,
    "数据源": 10, "拓扑": 15, "鉴权类型": 12, "鉴权头名": 15,
    "请求头": 40, "Query 参数": 40, "请求体": 30,
    "SQL 语句": 80, "响应模板": 80, "参数映射": 40,
    "静态响应体": 60,
    "故障-延迟毫秒": 14, "故障-错误率": 12, "故障-错误状态码": 16,
}

# 每列表头的 comment 说明（导入方在此对齐语义）
_HEADER_COMMENTS = {
    "方法": "GET/POST/PUT/PATCH/DELETE，空 → 跳过该行",
    "路径": "必填。全局唯一（method+path），跨 Sheet 移动 = 换域",
    "启用": "是/否；空视为「是」",
    "拓扑": "拓扑名（不是 ID）；找不到时留空 + warning",
    "请求头": "每行「名称|必填|期望值|样例|说明」；用换行分隔多条；\n值中禁止出现 |",
    "Query 参数": "每行「名称|类型|必填|样例|说明」；类型 string/int/bool",
    "请求体": "「Content-Type|必填|样例|说明」；不多值",
    "参数映射": "每行「参数名|位置|类型|必填|SQL绑定名」；位置 query/path/body",
    "故障-延迟毫秒": "空 = 不注入延迟",
    "故障-错误率": "0~1 之间，空 = 不注入错误",
    "故障-错误状态码": "空 = 默认 500",
}


def _decode_config(config_json: str) -> dict:
    if not config_json:
        return {}
    try:
        return _json.loads(config_json)
    except _json.JSONDecodeError:
        return {}


def _api_row_to_excel_values(api: dict, topology_name_by_id: dict) -> list:
    """把一行 api_configs 记录转成 Excel 一行的 20 个单元格值（按 MAIN_HEADERS 顺序）。"""
    config = _decode_config(api.get("config") or "")
    request = config.get("request") or {}
    auth = config.get("auth") or {}
    fault = config.get("fault") or {}

    headers = request.get("headers") or []
    query = request.get("query") or []
    body = request.get("body") or None
    param_mappings = config.get("params") or []

    tid = api.get("topology_id")
    topology_name = topology_name_by_id.get(tid, "") if tid else ""

    body_cell = ""
    if isinstance(body, dict):
        body_cell = format_cell_list([body], BODY_COLUMNS)

    return [
        api.get("method") or "",
        api.get("path") or "",
        api.get("name") or "",
        "是" if api.get("enabled") else "否",
        api.get("category") or "",
        api.get("group_name") or "",
        api.get("data_source") or "",
        topology_name,
        auth.get("type") or "none",
        auth.get("headerName") or "",
        format_cell_list(headers, HEADER_COLUMNS),
        format_cell_list(query, QUERY_COLUMNS),
        body_cell,
        api.get("sql_text") or "",
        config.get("responseTemplate") or "",
        format_cell_list(param_mappings, PARAM_MAPPING_COLUMNS),
        config.get("staticBody") or "",
        fault.get("delayMs"),
        fault.get("errorRate"),
        fault.get("errorStatus"),
    ]


def _write_instruction_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = INSTRUCTION_SHEET_NAME
    lines = [
        "接口 Excel 导入/导出使用说明",
        "",
        "1. 每个网管/设备一个 Sheet；一行一个接口。",
        "2. 变长字段（请求头 / Query 参数 / 参数映射）用「换行 + 竖线 |」表达：",
        "   例：请求头单元格 = 'Authorization|是|Bearer x||\\nX-Trace-Id|否|||'",
        "3. 值中禁止出现 |，否则该行导入时报错。",
        "4. 必填字段用「是 / 否」，也接受 true / false。",
        "5. 拓扑列填拓扑名（不是 ID），未找到时导入后为空并给出 warning。",
        "6. 匹配规则：按 (方法, 路径) 全局匹配；命中 → 更新（含换域），未命中 → 新建。",
        "7. 删除接口请到 UI 操作，从 Excel 里删除行不会删接口。",
        "8. Sheet 名 _ 开头（如本 Sheet）导入时被忽略。",
        "9. 域名含非法字符或超 31 字符时 Sheet 名会被自动清洗，A1 单元格的 comment 里记录原始域名。",
    ]
    for i, line in enumerate(lines, start=1):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 100


def _write_data_sheet(ws, api_rows: list, topology_name_by_id: dict, original_domain_name: str) -> None:
    """写入表头 + 数据行，并在 A1 comment 里记录原始域名。"""
    # 表头
    for col_idx, header in enumerate(MAIN_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        if header in _HEADER_COMMENTS:
            cell.comment = Comment(_HEADER_COMMENTS[header], "system")
        # 列宽
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = _COLUMN_WIDTHS.get(header, 15)

    # A1 记录原始域名（用于导入回填，防止 Sheet 名清洗后找不到域）
    if original_domain_name:
        existing_comment_text = ws["A1"].comment.text if ws["A1"].comment else ""
        marker = f"__ORIGINAL_DOMAIN__={original_domain_name}"
        combined = f"{existing_comment_text}\n{marker}" if existing_comment_text else marker
        ws["A1"].comment = Comment(combined, "system")

    # 数据行
    for row_idx, api in enumerate(api_rows, start=2):
        values = _api_row_to_excel_values(api, topology_name_by_id)
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = _WRAP_ALIGNMENT

    # 冻结表头行
    ws.freeze_panes = "A2"


def build_workbook(
    api_rows: list,
    domains: list,
    topologies: list,
) -> Workbook:
    """构建 xlsx。

    api_rows: [{id, name, method, path, enabled, group_name, domain_id, category,
                data_source, topology_id, sql_text, config}]
    domains:  [{id, name}]
    topologies: [{id, name}]
    """
    wb = Workbook()
    _write_instruction_sheet(wb)

    topology_name_by_id = {t["id"]: t["name"] for t in topologies}
    domain_by_id = {d["id"]: d for d in domains}

    # 按 domain_id 分组
    apis_by_domain: dict = {}
    for api in api_rows:
        dom_id = api.get("domain_id")
        apis_by_domain.setdefault(dom_id, []).append(api)

    used_sheet_names: set = {INSTRUCTION_SHEET_NAME}

    # 每个域一个 Sheet
    for dom_id, apis in apis_by_domain.items():
        if dom_id is None:
            sheet_name = UNCATEGORIZED_SHEET_NAME
            original_name = ""
            used_sheet_names.add(sheet_name)
        else:
            dom = domain_by_id.get(dom_id)
            if dom is None:
                sheet_name = sanitize_sheet_name(f"未知域_{dom_id[:8]}", used_sheet_names)
                original_name = ""
            else:
                original_name = dom["name"]
                sheet_name = sanitize_sheet_name(original_name, used_sheet_names)
        ws = wb.create_sheet(title=sheet_name)
        _write_data_sheet(ws, apis, topology_name_by_id, original_name)

    return wb


from dataclasses import dataclass, field
from openpyxl import load_workbook as _load_workbook


ALLOWED_METHODS = {"GET", "POST", "PUT", "PATCH", "DELETE"}
ALLOWED_DATA_SOURCES = {"sql", "static"}


@dataclass
class ParseResult:
    rows: list = field(default_factory=list)
    errors: list = field(default_factory=list)
    warnings: list = field(default_factory=list)
    auto_created_domains: list = field(default_factory=list)


def _extract_original_domain(a1_cell) -> Optional[str]:
    """从 A1 comment 里提取 __ORIGINAL_DOMAIN__= 后面的文本，找不到返回 None。"""
    if a1_cell.comment is None:
        return None
    text = a1_cell.comment.text or ""
    marker = "__ORIGINAL_DOMAIN__="
    for line in text.split("\n"):
        if line.startswith(marker):
            return line[len(marker):].strip()
    return None


def _resolve_topology(
    topo_name: str,
    existing_topologies: list,
    row_hint: str,
    warnings: list,
) -> Optional[str]:
    """按名称查 topology_id；多命中取最早创建；未命中记 warning 返回 None。"""
    if not topo_name:
        return None
    matches = [t for t in existing_topologies if t.get("name") == topo_name]
    if not matches:
        warnings.append(f"{row_hint}：拓扑名 '{topo_name}' 未找到，已留空")
        return None
    if len(matches) > 1:
        matches_sorted = sorted(matches, key=lambda t: t.get("created_at") or "")
        chosen = matches_sorted[0]
        warnings.append(f"{row_hint}：拓扑名 '{topo_name}' 有 {len(matches)} 个匹配，使用 {chosen['id']}")
        return chosen["id"]
    return matches[0]["id"]


def _parse_bool_cell(value: Any, default: bool = True) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if s in _TRUE_TOKENS:
        return True
    if s in _FALSE_TOKENS:
        return False
    return default


def _parse_number_cell(value: Any):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return value
    try:
        s = str(value).strip()
        if not s:
            return None
        return float(s) if "." in s else int(s)
    except (ValueError, TypeError):
        raise ExcelValidationError(f"数字列填了非数字 '{value}'")


def _resolve_domain_for_sheet(
    sheet_title: str,
    original_domain_name: Optional[str],
    existing_domains: list,
    auto_created: list,
):
    """返回 (domain_id, new_domain_name_to_create)。

    - 未归类 Sheet → (None, None)
    - A1 comment 有原始域名，且在 existing_domains 里 → (id, None)
    - A1 comment 有原始域名，但域已被删/改名 → (None, 原始名字)：新建
    - A1 无 comment，用 sheet_title 查 existing_domains
    - 都找不到 → 待创建
    """
    if sheet_title == UNCATEGORIZED_SHEET_NAME:
        return (None, None)

    lookup_name = original_domain_name or sheet_title
    matches = [d for d in existing_domains if d.get("name") == lookup_name]
    if matches:
        matches_sorted = sorted(matches, key=lambda d: d.get("created_at") or "")
        return (matches_sorted[0]["id"], None)

    # 待创建
    if lookup_name not in auto_created:
        auto_created.append(lookup_name)
    return (None, lookup_name)


def _row_to_config(cells: dict, row_hint: str) -> dict:
    """把一行的所有单元格值组装成 api_configs.config 字典。

    cells 是 {表头名: 单元格值} 的字典。抛 ExcelValidationError 表示行级错误。
    """
    config: dict = {}

    headers_text = cells.get("请求头") or ""
    query_text = cells.get("Query 参数") or ""
    body_text = cells.get("请求体") or ""
    params_text = cells.get("参数映射") or ""

    headers = parse_cell_list(headers_text, HEADER_COLUMNS, f"{row_hint} 请求头") if headers_text else []
    query = parse_cell_list(query_text, QUERY_COLUMNS, f"{row_hint} Query 参数") if query_text else []
    body_records = parse_cell_list(body_text, BODY_COLUMNS, f"{row_hint} 请求体") if body_text else []
    params = parse_cell_list(params_text, PARAM_MAPPING_COLUMNS, f"{row_hint} 参数映射") if params_text else []

    request: dict = {}
    if headers:
        request["headers"] = headers
    if query:
        request["query"] = query
    if body_records:
        request["body"] = body_records[0]
    if request:
        config["request"] = request

    auth_type = (cells.get("鉴权类型") or "none")
    auth_type = str(auth_type).strip() or "none"
    auth_header_name = str(cells.get("鉴权头名") or "").strip()
    if auth_type != "none" or auth_header_name:
        auth = {"type": auth_type}
        if auth_header_name:
            auth["headerName"] = auth_header_name
        config["auth"] = auth

    if params:
        config["params"] = params

    resp_tpl = cells.get("响应模板")
    if resp_tpl:
        config["responseTemplate"] = str(resp_tpl)

    static_body = cells.get("静态响应体")
    if static_body:
        config["staticBody"] = str(static_body)

    delay_ms = _parse_number_cell(cells.get("故障-延迟毫秒"))
    error_rate = _parse_number_cell(cells.get("故障-错误率"))
    error_status = _parse_number_cell(cells.get("故障-错误状态码"))
    fault: dict = {}
    if delay_ms is not None:
        fault["delayMs"] = int(delay_ms)
    if error_rate is not None:
        fault["errorRate"] = float(error_rate)
    if error_status is not None:
        fault["errorStatus"] = int(error_status)
    if fault:
        config["fault"] = fault

    return config


def parse_workbook(
    file_like,
    existing_domains: list,
    existing_topologies: list,
) -> ParseResult:
    """解析上传的 xlsx 文件对象。

    file_like: 支持 read() 的对象（如 UploadFile.file 或 io.BytesIO）
    existing_domains: [{id, name, created_at}]
    existing_topologies: [{id, name, created_at}]

    抛 ExcelValidationError → 致命错误（文件损坏 / 无数据 Sheet）。
    行级错误 → 收进 result.errors，那行被跳过。
    """
    try:
        wb = _load_workbook(file_like, read_only=False, data_only=True)
    except Exception as e:
        raise ExcelValidationError(f"Excel 打开失败：{e}") from e

    result = ParseResult()
    data_sheet_count = 0

    for ws in wb.worksheets:
        title = ws.title
        if title.startswith("_"):
            continue
        data_sheet_count += 1

        # 允许列缺失（用 None 填空），按 MAIN_HEADERS 的顺序建索引
        col_index = {h: i + 1 for i, h in enumerate(MAIN_HEADERS)}

        original_domain = _extract_original_domain(ws["A1"])
        domain_id, new_domain_name = _resolve_domain_for_sheet(
            title, original_domain, existing_domains, result.auto_created_domains,
        )

        for row_num in range(2, ws.max_row + 1):
            row_hint = f"Sheet '{title}' 第 {row_num} 行"
            cells = {}
            for header in MAIN_HEADERS:
                cells[header] = ws.cell(row=row_num, column=col_index[header]).value

            method_raw = cells.get("方法")
            path = cells.get("路径")
            if not method_raw or not path:
                # 整行空 → 静默跳过；单缺 method 或 path 也跳过（不算错，用户可能在整理表）
                continue
            method = str(method_raw).strip().upper()
            if method not in ALLOWED_METHODS:
                result.errors.append(f"{row_hint}：method '{method}' 不合法（允许 {sorted(ALLOWED_METHODS)}）")
                continue

            data_source = (str(cells.get("数据源") or "").strip() or "static").lower()
            if data_source not in ALLOWED_DATA_SOURCES:
                result.errors.append(f"{row_hint}：数据源 '{data_source}' 不合法")
                continue

            try:
                config = _row_to_config(cells, row_hint)
            except ExcelValidationError as e:
                result.errors.append(str(e))
                continue

            topology_id = _resolve_topology(
                str(cells.get("拓扑") or "").strip(),
                existing_topologies,
                row_hint,
                result.warnings,
            )

            row = {
                "method": method,
                "path": str(path).strip(),
                "name": str(cells.get("接口名") or "").strip(),
                "enabled": _parse_bool_cell(cells.get("启用"), default=True),
                "category": (str(cells.get("分类") or "").strip() or None),
                "group_name": (str(cells.get("分组") or "").strip() or None),
                "data_source": data_source,
                "topology_id": topology_id,
                "sql_text": (str(cells.get("SQL 语句") or "").strip() or None) if data_source == "sql" else None,
                "config": config,
                "domain_id": domain_id,
                "_new_domain_name": new_domain_name,
                "_sheet": title,
                "_row_num": row_num,
            }
            result.rows.append(row)

    if data_sheet_count == 0:
        raise ExcelValidationError("Excel 中未找到任何数据 Sheet")

    return result
