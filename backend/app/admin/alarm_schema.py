import re
import uuid
from io import BytesIO
from typing import Any

from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook

from app.db.connection import connect, transaction
from app.admin.schemas import (
    AlarmSchemaCreate,
    AlarmSchemaUpdate,
    AlarmSchemaDetail,
    AlarmSchemaItem,
    AlarmSchemaFieldItem,
    AlarmSchemaFieldCreate,
    AlarmSchemaExportRequest,
    AlarmSchemaImportPreviewItem,
    AlarmSchemaImportPreview,
    AlarmSchemaImportResult,
)

router = APIRouter(prefix="/admin/api", tags=["告警模板"])

_IDENT_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")
_FIXED_COLS = {"id", "node_id", "alarm_index", "created_at", "updated_at"}


def _new_id() -> str:
    return f"as_{uuid.uuid4().hex[:12]}"


def _validate_field_keys(fields: list[AlarmSchemaFieldCreate]) -> None:
    seen: set = set()
    for f in fields:
        if not _IDENT_RE.match(f.field_key):
            raise HTTPException(
                status_code=400,
                detail={
                    "code": 40001,
                    "message": f"字段键非法（仅支持字母/数字/下划线，且以字母或下划线开头）: {f.field_key}",
                },
            )
        if f.field_key in _FIXED_COLS:
            raise HTTPException(
                status_code=400,
                detail={
                    "code": 40002,
                    "message": f"字段键与固定列冲突: {f.field_key}（固定列：{', '.join(sorted(_FIXED_COLS))}）",
                },
            )
        if f.field_key in seen:
            raise HTTPException(
                status_code=400,
                detail={"code": 40003, "message": f"字段键重复: {f.field_key}"},
            )
        seen.add(f.field_key)


def _get_fields(conn, schema_id: str) -> list[AlarmSchemaFieldItem]:
    rows = conn.execute(
        "SELECT * FROM alarm_schema_fields WHERE alarm_schema_id = ? ORDER BY sort_order, id",
        (schema_id,),
    ).fetchall()
    return [
        AlarmSchemaFieldItem(
            id=r["id"],
            alarm_schema_id=r["alarm_schema_id"],
            field_key=r["field_key"],
            field_label=r["field_label"],
            field_type=r["field_type"],
            max_length=r["max_length"],
            default_value=r["default_value"],
            options=r["options"],
            required=bool(r["required"]),
            sort_order=r["sort_order"],
            mapping_target=r["mapping_target"],
        )
        for r in rows
    ]


def _row_to_detail(conn, row) -> AlarmSchemaDetail:
    return AlarmSchemaDetail(
        id=row["id"],
        code=row["code"],
        name=row["name"],
        description=row["description"],
        created_at=row["created_at"],
        updated_at=row["updated_at"],
        display_field_key=row["display_field_key"],
        fields=_get_fields(conn, row["id"]),
    )


@router.get("/alarm-schemas")
def list_alarm_schemas() -> dict:
    with connect() as conn:
        rows = conn.execute(
            "SELECT * FROM alarm_schemas ORDER BY created_at DESC"
        ).fetchall()
        items = [
            AlarmSchemaItem(
                id=r["id"], code=r["code"], name=r["name"],
                description=r["description"],
                display_field_key=r["display_field_key"],
                created_at=r["created_at"], updated_at=r["updated_at"],
            ).model_dump(mode="json", by_alias=True)
            for r in rows
        ]
    return {"code": 0, "data": items, "message": "ok"}


@router.get("/alarm-schemas/{schema_id}")
def get_alarm_schema(schema_id: str) -> dict:
    with connect() as conn:
        row = conn.execute(
            "SELECT * FROM alarm_schemas WHERE id = ?", (schema_id,)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail={"code": 40404, "message": "告警模板不存在"})
        d = _row_to_detail(conn, row)
    return {"code": 0, "data": d.model_dump(mode="json", by_alias=True), "message": "ok"}


@router.post("/alarm-schemas")
def create_alarm_schema(data: AlarmSchemaCreate) -> dict:
    _validate_field_keys(data.fields)
    sid = _new_id()
    with transaction() as conn:
        dup = conn.execute(
            "SELECT id FROM alarm_schemas WHERE code = ?", (data.code,)
        ).fetchone()
        if dup:
            raise HTTPException(
                status_code=409,
                detail={"code": 40901, "message": f"code 已存在: {data.code}"},
            )
        conn.execute(
            "INSERT INTO alarm_schemas (id, code, name, description, display_field_key) "
            "VALUES (?, ?, ?, ?, ?)",
            (sid, data.code, data.name, data.description, data.display_field_key),
        )
        for f in data.fields:
            conn.execute(
                "INSERT INTO alarm_schema_fields "
                "(alarm_schema_id, field_key, field_label, field_type, max_length, "
                " default_value, options, required, sort_order, mapping_target) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (sid, f.field_key, f.field_label, f.field_type, f.max_length,
                 f.default_value, f.options, int(f.required), f.sort_order, f.mapping_target),
            )
        row = conn.execute("SELECT * FROM alarm_schemas WHERE id = ?", (sid,)).fetchone()
        d = _row_to_detail(conn, row)
    return {"code": 0, "data": d.model_dump(mode="json", by_alias=True), "message": "ok"}


@router.put("/alarm-schemas/{schema_id}")
def update_alarm_schema(schema_id: str, data: AlarmSchemaUpdate) -> dict:
    if data.fields is not None:
        _validate_field_keys(data.fields)
    with transaction() as conn:
        row = conn.execute("SELECT * FROM alarm_schemas WHERE id = ?", (schema_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail={"code": 40404, "message": "告警模板不存在"})

        sets: list[str] = []
        params: list[Any] = []
        if data.name is not None:
            sets.append("name = ?")
            params.append(data.name)
        if data.description is not None:
            sets.append("description = ?")
            params.append(data.description)
        if 'display_field_key' in data.model_fields_set:
            sets.append("display_field_key = ?")
            params.append(data.display_field_key)
        if sets:
            sets.append("updated_at = datetime('now')")
            params.append(schema_id)
            conn.execute(f"UPDATE alarm_schemas SET {', '.join(sets)} WHERE id = ?", params)

        if data.fields is not None:
            conn.execute("DELETE FROM alarm_schema_fields WHERE alarm_schema_id = ?", (schema_id,))
            for f in data.fields:
                conn.execute(
                    "INSERT INTO alarm_schema_fields "
                    "(alarm_schema_id, field_key, field_label, field_type, max_length, "
                    " default_value, options, required, sort_order, mapping_target) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (schema_id, f.field_key, f.field_label, f.field_type, f.max_length,
                     f.default_value, f.options, int(f.required), f.sort_order, f.mapping_target),
                )

        row = conn.execute("SELECT * FROM alarm_schemas WHERE id = ?", (schema_id,)).fetchone()
        d = _row_to_detail(conn, row)
    return {"code": 0, "data": d.model_dump(mode="json", by_alias=True), "message": "ok"}


@router.delete("/alarm-schemas/{schema_id}")
def delete_alarm_schema(schema_id: str) -> dict:
    with transaction() as conn:
        row = conn.execute("SELECT id FROM alarm_schemas WHERE id = ?", (schema_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail={"code": 40404, "message": "告警模板不存在"})

        refs = conn.execute(
            "SELECT name FROM topologies WHERE alarm_schema_id = ?", (schema_id,)
        ).fetchall()
        if refs:
            raise HTTPException(
                status_code=409,
                detail={
                    "code": 40901,
                    "message": "告警模板被以下拓扑引用，无法删除",
                    "details": {"referencedBy": [r["name"] for r in refs]},
                },
            )
        conn.execute("DELETE FROM alarm_schemas WHERE id = ?", (schema_id,))
    return {"code": 0, "data": None, "message": "ok"}


# ============== Excel 导入导出 ==============

_SHEET_INVALID_CHARS = re.compile(r'[\\\*/\[\]\?:]')


def _safe_sheet_name(code: str) -> str:
    name = _SHEET_INVALID_CHARS.sub('_', code)
    if len(name) > 31:
        name = name[:28] + "..."
    return name


def _build_alarm_schemas_excel(items: list) -> BytesIO:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "模板汇总"
    ws1.append(["Code", "名称", "描述", "展示字段Key",
                "字段数", "创建时间", "更新时间"])
    for item in items:
        ws1.append([
            item.get("code"),
            item.get("name"),
            item.get("description"),
            item.get("displayFieldKey"),
            len(item.get("fields") or []),
            item.get("createdAt"),
            item.get("updatedAt"),
        ])

    for item in items:
        fields = item.get("fields") or []
        sheet_name = _safe_sheet_name(item.get("code", item.get("id", "unknown")))
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["字段标识", "显示名称", "字段类型", "最大长度",
                    "默认值", "选项", "必填", "排序", "映射节点属性"])
        for f in fields:
            ws.append([
                f.get("fieldKey"),
                f.get("fieldLabel"),
                f.get("fieldType"),
                f.get("maxLength"),
                f.get("defaultValue"),
                f.get("options"),
                "是" if f.get("required") else "否",
                f.get("sortOrder"),
                f.get("mappingTarget"),
            ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _build_header_map(ws) -> dict:
    """读取第 1 行构建 {表头名: 列索引}，空返回 {}."""
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not row:
        return {}
    return {str(v).strip(): i for i, v in enumerate(row) if v is not None and str(v).strip()}


def _col(headers: dict, name: str, row: tuple):
    """按表头名取值，找不到列或空值返回 None。"""
    idx = headers.get(name)
    if idx is None or idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    return str(val).strip() or None


def _load_import_workbook(contents: bytes):
    try:
        wb = load_workbook(filename=BytesIO(contents))
    except Exception:
        raise HTTPException(
            status_code=400,
            detail={"code": 40410, "message": "文件无法解析，请确认是有效的 xlsx 文件"},
        )
    if "模板汇总" not in wb.sheetnames:
        raise HTTPException(
            status_code=400,
            detail={"code": 40411, "message": "缺少「模板汇总」Sheet"},
        )
    return wb


@router.post("/alarm-schemas/import/preview")
async def preview_alarm_schemas_import(file: UploadFile = File(...)):
    if not file.filename or not file.filename.endswith('.xlsx'):
        raise HTTPException(
            status_code=400,
            detail={"code": 40412, "message": "仅支持 .xlsx 文件"},
        )

    contents = await file.read()
    wb = _load_import_workbook(contents)
    ws = wb["模板汇总"]

    to_create: list = []
    to_update: list = []
    errors: list = []

    with connect() as conn:
        headers = _build_header_map(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None or v == '' for v in row):
                break
            code = _col(headers, "Code", row)
            name = _col(headers, "名称", row)

            if not code or not name:
                errors.append(f"Code={code or '(空)'} 缺少必填字段（Code/名称），跳过")
                continue

            existing = conn.execute(
                "SELECT id, name FROM alarm_schemas WHERE code = ?", (code,)
            ).fetchone()

            if existing:
                to_update.append({
                    "code": code,
                    "name": name,
                    "old_name": existing["name"],
                })
            else:
                to_create.append({"code": code, "name": name})

    preview = AlarmSchemaImportPreview(
        to_create=[AlarmSchemaImportPreviewItem(**c) for c in to_create],
        to_update=[AlarmSchemaImportPreviewItem(**u) for u in to_update],
        errors=errors,
    )
    return {"code": 0, "data": preview.model_dump(mode="json", by_alias=True), "message": "ok"}


@router.post("/alarm-schemas/export")
def export_alarm_schemas(data: AlarmSchemaExportRequest):
    with connect() as conn:
        if data.ids:
            placeholders = ",".join("?" for _ in data.ids)
            rows = conn.execute(
                f"SELECT * FROM alarm_schemas WHERE id IN ({placeholders}) ORDER BY created_at DESC",
                tuple(data.ids),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM alarm_schemas ORDER BY created_at DESC"
            ).fetchall()
        items = []
        for r in rows:
            detail = _row_to_detail(conn, r)
            items.append(detail.model_dump(mode="json", by_alias=True))

    excel = _build_alarm_schemas_excel(items)
    return StreamingResponse(
        excel,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=alarm-schemas-export.xlsx"},
    )
