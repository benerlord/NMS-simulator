"""画布 Excel 导入导出的内部工具（编解码 + 校验，不直接触碰 DB）。"""
import json as _json
from typing import Any, Optional


INSTRUCTION_SHEET_NAME = "_使用说明"
INDEX_SHEET_NAME = "_总表"
META_SHEET_NAME = "拓扑元信息"
NODE_GROUP_SHEET_NAME = "节点组"
NODE_GROUP_EDGE_STRATEGY_SHEET_NAME = "节点组边策略"
NODE_ALARM_SHEET_NAME = "节点告警"
NODE_GROUP_ALARM_SHEET_NAME = "节点组告警"

SHEET_NAME_MAX_LEN = 31
INVALID_SHEET_CHARS = ":\\/?*[]"

NODE_TYPE_MARKER = "__NODE_TYPE_CODE__"
EDGE_TYPE_MARKER = "__EDGE_TYPE_CODE__"
NODE_GROUP_MARKER = "__NODE_GROUP__"
NODE_GROUP_EDGE_STRATEGY_MARKER = "__NODE_GROUP_EDGE_STRATEGY__"
NODE_ALARM_MARKER = "__NODE_ALARM__"
NODE_GROUP_ALARM_MARKER = "__NODE_GROUP_ALARM__"

FIELD_SEP = "|"
RECORD_SEP = "\n"
PARAM_KV_SEP = ";"

NODE_FIXED_HEADERS = ["名称", "DN", "状态", "画布 X", "画布 Y", "所属组"]
EDGE_FIXED_HEADERS = ["源节点", "目标节点", "状态"]
NODE_GROUP_HEADERS = [
    "组名", "节点类型代码", "节点数量", "命名模板", "已展开",
    "画布 X", "画布 Y", "属性策略",
]
NODE_GROUP_EDGE_STRATEGY_HEADERS = [
    "源组名", "目标", "目标类型", "边类型代码", "模式", "K",
]
NODE_ALARM_FIXED_HEADERS = ["节点类型代码", "节点名称", "告警序号"]
NODE_GROUP_ALARM_FIXED_HEADERS = ["组名", "告警序号"]
META_HEADERS = ["字段", "值"]
INDEX_HEADERS = ["类别", "类型代码", "Sheet 名", "行数", "跳转"]

ALLOWED_STATUS = {"online", "offline", "unknown", "warning", "error"}
ALLOWED_STRATEGY = {"fixed", "random", "increment", "range"}
ALLOWED_MODE = {"modulo", "one_to_n", "all_to_all", "dense"}
ALLOWED_TARGET_KIND = {"组", "节点"}


class ExcelValidationError(Exception):
    """行级/致命校验错误。上层区分——parse_workbook 里被捕获归到 errors 就是行级；
    在打开文件/无数据 Sheet 场景直接抛就是致命。"""


def sanitize_sheet_name(name: str, used: set) -> str:
    """把 domain/nodeType/edgeType 名清洗成 Excel 合法 sheet 名并防重名。副作用：mutating used。"""
    sanitized = name
    for ch in INVALID_SHEET_CHARS:
        sanitized = sanitized.replace(ch, "_")
    if len(sanitized) > SHEET_NAME_MAX_LEN:
        sanitized = sanitized[:SHEET_NAME_MAX_LEN]
    if not sanitized.strip():
        sanitized = "未命名"
    if sanitized not in used:
        used.add(sanitized)
        return sanitized
    suffix = 2
    while True:
        tag = f"~{suffix}"
        base_max = SHEET_NAME_MAX_LEN - len(tag)
        candidate = sanitized[:base_max] + tag
        if candidate not in used:
            used.add(candidate)
            return candidate
        suffix += 1


# --- Attr strategy encoders/decoders ---

def format_attr_strategy_row(strategy: dict) -> str:
    """把一个 attr_strategy dict 序列化成 'field_key|strategy|k=v;k=v' 行。"""
    field_key = strategy["field_key"]
    stype = strategy["strategy"]
    params = []
    if stype == "fixed":
        params.append(f"fixedValue={strategy.get('fixed_value', '')}")
    elif stype == "random":
        pool = strategy.get("pool") or []
        params.append(f"pool={';'.join(str(x) for x in pool)}")
    elif stype == "increment":
        params.append(f"base={strategy.get('base', '')}")
        params.append(f"step={strategy.get('step', '')}")
    elif stype == "range":
        params.append(f"min={strategy.get('min', '')}")
        params.append(f"max={strategy.get('max', '')}")
    param_str = PARAM_KV_SEP.join(params)
    return f"{field_key}{FIELD_SEP}{stype}{FIELD_SEP}{param_str}"


def parse_attr_strategy_row(line: str, row_hint: str) -> dict:
    """把 'field_key|strategy|k=v;k=v' 行解回 dict。抛 ExcelValidationError 表示行级错误。"""
    parts = line.split(FIELD_SEP)
    if len(parts) < 2:
        raise ExcelValidationError(f"{row_hint}：属性策略格式错，缺字段名或策略类型")
    field_key = parts[0].strip()
    stype = parts[1].strip()
    param_str = FIELD_SEP.join(parts[2:]).strip() if len(parts) > 2 else ""

    if stype not in ALLOWED_STRATEGY:
        raise ExcelValidationError(
            f"{row_hint}：属性策略类型 '{stype}' 不合法（允许 {sorted(ALLOWED_STRATEGY)}）"
        )
    kv = {}
    if param_str:
        for pair in param_str.split(PARAM_KV_SEP):
            if "=" not in pair:
                continue
            k, v = pair.split("=", 1)
            kv[k.strip()] = v.strip()

    result: dict = {"field_key": field_key, "strategy": stype}
    if stype == "fixed":
        if "fixedValue" not in kv or not kv["fixedValue"]:
            raise ExcelValidationError(f"{row_hint}：fixed 策略缺 fixedValue")
        result["fixed_value"] = kv["fixedValue"]
    elif stype == "random":
        pool_str = kv.get("pool", "")
        if not pool_str:
            raise ExcelValidationError(f"{row_hint}：random 策略缺 pool")
        result["pool"] = [x for x in pool_str.split(";") if x]
    elif stype == "increment":
        if "base" not in kv or "step" not in kv:
            raise ExcelValidationError(f"{row_hint}：increment 策略缺 base 或 step")
        result["base"] = kv["base"]
        result["step"] = kv["step"]
    elif stype == "range":
        if "min" not in kv or "max" not in kv:
            raise ExcelValidationError(f"{row_hint}：range 策略缺 min 或 max")
        try:
            result["min"] = int(kv["min"])
            result["max"] = int(kv["max"])
        except ValueError:
            raise ExcelValidationError(f"{row_hint}：range 的 min/max 必须是整数")
    return result


# --- Edge strategy encoders/decoders ---

def format_edge_strategy_row(strategy: dict) -> str:
    """把一个 edge_strategy dict 序列化成 '源组名|目标|目标类型|边类型|模式|K' 行。"""
    k_str = "" if strategy.get("ratio_k") is None else str(strategy["ratio_k"])
    return FIELD_SEP.join([
        strategy["source_group_name"],
        strategy["target_name"],
        strategy["target_kind"],
        strategy["edge_type_code"],
        strategy["mode"],
        k_str,
    ])


def parse_edge_strategy_row(line: str, row_hint: str) -> dict:
    parts = line.split(FIELD_SEP)
    if len(parts) < 6:
        raise ExcelValidationError(
            f"{row_hint}：边策略字段数 {len(parts)} 不足 6（源组名|目标|目标类型|边类型|模式|K）"
        )
    if len(parts) > 6:
        raise ExcelValidationError(
            f"{row_hint}：边策略字段数 {len(parts)} 超过 6"
        )
    src, tgt, kind, etype, mode, k_str = [p.strip() for p in parts]
    if kind not in ALLOWED_TARGET_KIND:
        raise ExcelValidationError(
            f"{row_hint}：目标类型 '{kind}' 不合法（应为 组/节点）"
        )
    if mode not in ALLOWED_MODE:
        raise ExcelValidationError(
            f"{row_hint}：边策略模式 '{mode}' 不合法（允许 {sorted(ALLOWED_MODE)}）"
        )
    ratio_k: Optional[int] = None
    if k_str:
        try:
            ratio_k = int(k_str)
        except ValueError:
            raise ExcelValidationError(f"{row_hint}：K 值 '{k_str}' 必须是整数")
    if mode in ("modulo", "one_to_n") and ratio_k is None:
        raise ExcelValidationError(
            f"{row_hint}：{mode} 模式必须提供 K"
        )
    return {
        "source_group_name": src,
        "target_name": tgt,
        "target_kind": kind,
        "edge_type_code": etype,
        "mode": mode,
        "ratio_k": ratio_k,
    }


# --- Workbook builder ---

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.hyperlink import Hyperlink

_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
_WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
_HYPERLINK_FONT = Font(color="0563C1", underline="single")


def _set_a1_marker(ws, marker: str, value: str = "1") -> None:
    """在 A1 单元格 comment 里追加 marker=value（不覆盖已有 comment 内容）。"""
    existing = ws["A1"].comment.text if ws["A1"].comment else ""
    line = f"{marker}={value}"
    combined = f"{existing}\n{line}" if existing else line
    ws["A1"].comment = Comment(combined, "system")


def _apply_header_style(ws, num_cols: int) -> None:
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL


def _write_instruction_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = INSTRUCTION_SHEET_NAME
    lines = [
        "画布 Excel 导入/导出使用说明",
        "",
        "1. 一个 xlsx = 一个拓扑。导入时始终新建拓扑（名字冲突加\"(导入 N)\"后缀）。",
        "2. 拓扑元信息 Sheet：拓扑名/描述/网管/告警模板绑定（key-value 表）。",
        "3. 每种节点类型一 Sheet（如\"路由器\"）；每种边类型一 Sheet（如\"连接\"）。",
        "4. 边的\"源节点\"/\"目标节点\"列填节点名称；同 (类型, 名称) 组合必须唯一。",
        "5. 节点组 Sheet 存组定义；节点组边策略 Sheet 存组间连线策略（一策略一行）。",
        "6. 拓扑绑了告警模板才生成\"节点告警\"+\"节点组告警\"两 Sheet。",
        "7. Sheet 名以 _ 开头（如本 Sheet 与 _总表）导入时被忽略。",
        "8. 变长字段（属性策略）用\"换行 + 竖线 |\"，跟接口 Excel 一致：值中禁止出现 |。",
        "9. 已 materialize 的组，其物理节点在 nodeType Sheet 里通过\"所属组\"列关联组名。",
    ]
    for i, line in enumerate(lines, start=1):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 100


def _write_meta_sheet(wb: Workbook, topo: dict) -> None:
    ws = wb.create_sheet(title=META_SHEET_NAME)
    for c, h in enumerate(META_HEADERS, start=1):
        ws.cell(row=1, column=c, value=h)
    _apply_header_style(ws, len(META_HEADERS))
    rows = [
        ("拓扑名称", topo.get("name") or ""),
        ("描述", topo.get("description") or ""),
        ("版本", topo.get("version") or 1),
        ("所属网管/设备", topo.get("domain_name") or ""),
        ("告警模板", topo.get("alarm_schema_code") or ""),
    ]
    for i, (k, v) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40


def _write_node_sheet(wb: Workbook, node_type: dict, nodes: list, used: set) -> str:
    """写一个 nodeType 的 Sheet，返回实际用的 sheet name。"""
    sheet_name = sanitize_sheet_name(node_type["name"], used)
    ws = wb.create_sheet(title=sheet_name)
    field_keys = [f["field_key"] for f in node_type.get("fields", [])]
    headers = NODE_FIXED_HEADERS + field_keys
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _apply_header_style(ws, len(headers))
    _set_a1_marker(ws, NODE_TYPE_MARKER, node_type["code"])
    for i, n in enumerate(nodes, start=2):
        ws.cell(row=i, column=1, value=n.get("name") or "")
        ws.cell(row=i, column=2, value=n.get("dn") or "")
        ws.cell(row=i, column=3, value=n.get("status") or "online")
        ws.cell(row=i, column=4, value=n.get("canvas_x"))
        ws.cell(row=i, column=5, value=n.get("canvas_y"))
        ws.cell(row=i, column=6, value=n.get("group_name") or "")
        for c, fk in enumerate(field_keys, start=7):
            ws.cell(row=i, column=c, value=n.get("attrs", {}).get(fk))
    ws.freeze_panes = "A2"
    return sheet_name


def _write_edge_sheet(wb: Workbook, edge_type: dict, edges: list, node_name_by_id: dict, used: set) -> str:
    sheet_name = sanitize_sheet_name(edge_type["name"], used)
    ws = wb.create_sheet(title=sheet_name)
    field_keys = [f["field_key"] for f in edge_type.get("fields", [])]
    headers = EDGE_FIXED_HEADERS + field_keys
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _apply_header_style(ws, len(headers))
    _set_a1_marker(ws, EDGE_TYPE_MARKER, edge_type["code"])
    for i, e in enumerate(edges, start=2):
        ws.cell(row=i, column=1, value=node_name_by_id.get(e["source_id"], ""))
        ws.cell(row=i, column=2, value=node_name_by_id.get(e["target_id"], ""))
        ws.cell(row=i, column=3, value=e.get("status") or "online")
        for c, fk in enumerate(field_keys, start=4):
            ws.cell(row=i, column=c, value=e.get("attrs", {}).get(fk))
    ws.freeze_panes = "A2"
    return sheet_name


def _write_node_group_sheet(wb: Workbook, groups: list, node_type_code_by_id: dict) -> None:
    ws = wb.create_sheet(title=NODE_GROUP_SHEET_NAME)
    for c, h in enumerate(NODE_GROUP_HEADERS, start=1):
        ws.cell(row=1, column=c, value=h)
    _apply_header_style(ws, len(NODE_GROUP_HEADERS))
    _set_a1_marker(ws, NODE_GROUP_MARKER)
    for i, g in enumerate(groups, start=2):
        ws.cell(row=i, column=1, value=g.get("group_name") or "")
        ws.cell(row=i, column=2, value=node_type_code_by_id.get(g.get("node_type_id"), ""))
        ws.cell(row=i, column=3, value=g.get("node_count"))
        ws.cell(row=i, column=4, value=g.get("name_template") or "")
        ws.cell(row=i, column=5, value="是" if g.get("materialized_at") else "否")
        ws.cell(row=i, column=6, value=g.get("canvas_x"))
        ws.cell(row=i, column=7, value=g.get("canvas_y"))
        attrs = g.get("attr_strategies") or []
        lines = [format_attr_strategy_row(s) for s in attrs]
        cell = ws.cell(row=i, column=8, value=RECORD_SEP.join(lines) if lines else "")
        cell.alignment = _WRAP_ALIGNMENT
    ws.freeze_panes = "A2"


def _write_node_group_edge_strategy_sheet(wb: Workbook, strategies: list) -> None:
    ws = wb.create_sheet(title=NODE_GROUP_EDGE_STRATEGY_SHEET_NAME)
    for c, h in enumerate(NODE_GROUP_EDGE_STRATEGY_HEADERS, start=1):
        ws.cell(row=1, column=c, value=h)
    _apply_header_style(ws, len(NODE_GROUP_EDGE_STRATEGY_HEADERS))
    _set_a1_marker(ws, NODE_GROUP_EDGE_STRATEGY_MARKER)
    for i, s in enumerate(strategies, start=2):
        ws.cell(row=i, column=1, value=s.get("source_group_name") or "")
        ws.cell(row=i, column=2, value=s.get("target_name") or "")
        ws.cell(row=i, column=3, value=s.get("target_kind") or "组")
        ws.cell(row=i, column=4, value=s.get("edge_type_code") or "")
        ws.cell(row=i, column=5, value=s.get("mode") or "")
        k = s.get("ratio_k")
        ws.cell(row=i, column=6, value=k if k is not None else "")
    ws.freeze_panes = "A2"


def _write_node_alarm_sheet(wb: Workbook, alarms: list, alarm_field_keys: list,
                             node_id_to_type_code: dict, node_name_by_id: dict) -> None:
    """alarms 结构: [{node_id, alarm_index, attrs}]。node_id_to_type_code / node_name_by_id 反查用。"""
    ws = wb.create_sheet(title=NODE_ALARM_SHEET_NAME)
    headers = NODE_ALARM_FIXED_HEADERS + alarm_field_keys
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _apply_header_style(ws, len(headers))
    _set_a1_marker(ws, NODE_ALARM_MARKER)
    for i, a in enumerate(alarms, start=2):
        nid = a.get("node_id")
        ws.cell(row=i, column=1, value=node_id_to_type_code.get(nid, ""))
        ws.cell(row=i, column=2, value=node_name_by_id.get(nid, ""))
        ws.cell(row=i, column=3, value=a.get("alarm_index"))
        for c, fk in enumerate(alarm_field_keys, start=4):
            ws.cell(row=i, column=c, value=a.get("attrs", {}).get(fk))
    ws.freeze_panes = "A2"


def _write_node_group_alarm_sheet(wb: Workbook, alarms: list, alarm_field_keys: list,
                                    group_name_by_id: dict) -> None:
    ws = wb.create_sheet(title=NODE_GROUP_ALARM_SHEET_NAME)
    headers = NODE_GROUP_ALARM_FIXED_HEADERS + alarm_field_keys
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _apply_header_style(ws, len(headers))
    _set_a1_marker(ws, NODE_GROUP_ALARM_MARKER)
    for i, a in enumerate(alarms, start=2):
        ws.cell(row=i, column=1, value=group_name_by_id.get(a.get("node_group_id"), ""))
        ws.cell(row=i, column=2, value=a.get("alarm_index"))
        for c, fk in enumerate(alarm_field_keys, start=3):
            ws.cell(row=i, column=c, value=a.get("attrs", {}).get(fk))
    ws.freeze_panes = "A2"


def _write_index_sheet(wb: Workbook, index_rows: list) -> None:
    """index_rows: list[dict{category, type_code, sheet_name, row_count}]"""
    ws = wb.create_sheet(title=INDEX_SHEET_NAME, index=1)
    for c, h in enumerate(INDEX_HEADERS, start=1):
        ws.cell(row=1, column=c, value=h)
    _apply_header_style(ws, len(INDEX_HEADERS))
    for i, row in enumerate(index_rows, start=2):
        ws.cell(row=i, column=1, value=row["category"])
        ws.cell(row=i, column=2, value=row.get("type_code") or "")
        cell = ws.cell(row=i, column=3, value=row["sheet_name"])
        cell.hyperlink = Hyperlink(
            ref=cell.coordinate,
            location=f"'{row['sheet_name']}'!A1",
            display=row["sheet_name"],
        )
        cell.font = _HYPERLINK_FONT
        ws.cell(row=i, column=4, value=row.get("row_count", 0))
        ws.cell(row=i, column=5, value="点击左侧跳转")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 20


def build_workbook(
    topology: dict,
    node_types: list,
    edge_types: list,
    nodes_by_type_code: dict,
    edges_by_type_code: dict,
    node_groups: list,
    node_group_edge_strategies: list,
    alarm_schema_fields: list,
    node_alarms: list,
    node_group_alarms: list,
) -> Workbook:
    """构建拓扑 Excel workbook。参见 spec Section: 数据入参形状。"""
    wb = Workbook()
    _write_instruction_sheet(wb)
    _write_meta_sheet(wb, topology)

    # 反查映射
    node_name_by_id: dict = {}
    node_id_to_type_code: dict = {}
    for nt in node_types:
        for n in nodes_by_type_code.get(nt["code"], []):
            node_name_by_id[n["id"]] = n["name"]
            node_id_to_type_code[n["id"]] = nt["code"]
    node_type_code_by_id = {nt["id"]: nt["code"] for nt in node_types}
    group_name_by_id = {g["id"]: g["group_name"] for g in node_groups}

    used: set = {INSTRUCTION_SHEET_NAME, INDEX_SHEET_NAME, META_SHEET_NAME,
                 NODE_GROUP_SHEET_NAME, NODE_GROUP_EDGE_STRATEGY_SHEET_NAME,
                 NODE_ALARM_SHEET_NAME, NODE_GROUP_ALARM_SHEET_NAME}

    index_rows: list = []

    for nt in node_types:
        nodes = nodes_by_type_code.get(nt["code"], [])
        name = _write_node_sheet(wb, nt, nodes, used)
        index_rows.append({"category": "节点", "type_code": nt["code"],
                           "sheet_name": name, "row_count": len(nodes)})

    for et in edge_types:
        edges = edges_by_type_code.get(et["code"], [])
        name = _write_edge_sheet(wb, et, edges, node_name_by_id, used)
        index_rows.append({"category": "边", "type_code": et["code"],
                           "sheet_name": name, "row_count": len(edges)})

    _write_node_group_sheet(wb, node_groups, node_type_code_by_id)
    index_rows.append({"category": "节点组", "type_code": "",
                       "sheet_name": NODE_GROUP_SHEET_NAME, "row_count": len(node_groups)})

    _write_node_group_edge_strategy_sheet(wb, node_group_edge_strategies)
    index_rows.append({"category": "节点组边策略", "type_code": "",
                       "sheet_name": NODE_GROUP_EDGE_STRATEGY_SHEET_NAME,
                       "row_count": len(node_group_edge_strategies)})

    if topology.get("alarm_schema_code") and alarm_schema_fields:
        alarm_field_keys = [f["field_key"] for f in alarm_schema_fields if not f.get("mapping_target")]
        _write_node_alarm_sheet(wb, node_alarms, alarm_field_keys,
                                 node_id_to_type_code, node_name_by_id)
        index_rows.append({"category": "节点告警", "type_code": "",
                           "sheet_name": NODE_ALARM_SHEET_NAME, "row_count": len(node_alarms)})

        _write_node_group_alarm_sheet(wb, node_group_alarms, alarm_field_keys, group_name_by_id)
        index_rows.append({"category": "节点组告警", "type_code": "",
                           "sheet_name": NODE_GROUP_ALARM_SHEET_NAME,
                           "row_count": len(node_group_alarms)})

    _write_index_sheet(wb, index_rows)

    return wb


# --- Workbook parser ---

from dataclasses import dataclass, field
from openpyxl import load_workbook as _load_workbook


@dataclass
class ParseResult:
    meta: dict = field(default_factory=dict)
    nodes_by_type_code: dict = field(default_factory=dict)
    edges_by_type_code: dict = field(default_factory=dict)
    node_groups: list = field(default_factory=list)
    node_group_edge_strategies: list = field(default_factory=list)
    node_alarms: list = field(default_factory=list)
    node_group_alarms: list = field(default_factory=list)
    errors: list = field(default_factory=list)
    warnings: list = field(default_factory=list)


def _extract_marker(a1_cell, marker: str) -> Optional[str]:
    """从 A1 comment 里读 marker=value 的 value；不存在返回 None。"""
    if a1_cell.comment is None:
        return None
    text = a1_cell.comment.text or ""
    prefix = f"{marker}="
    for line in text.split("\n"):
        if line.startswith(prefix):
            return line[len(prefix):].strip()
    return None


def _read_meta_sheet(ws) -> dict:
    """两列 key-value → dict。"""
    kv = {}
    for r in range(2, ws.max_row + 1):
        k = ws.cell(row=r, column=1).value
        v = ws.cell(row=r, column=2).value
        if k:
            kv[str(k).strip()] = v
    name_raw = kv.get("拓扑名称")
    name = name_raw.strip() if isinstance(name_raw, str) else name_raw
    return {
        "name": name,
        "description": kv.get("描述"),
        "version": kv.get("版本") or 1,
        "domain_name": (kv.get("所属网管/设备") or None),
        "alarm_schema_code": (kv.get("告警模板") or None),
    }


def _read_node_sheet(ws, code: str) -> list:
    field_start_col = len(NODE_FIXED_HEADERS) + 1
    field_keys = []
    c = field_start_col
    while ws.cell(row=1, column=c).value:
        field_keys.append(str(ws.cell(row=1, column=c).value).strip())
        c += 1

    nodes = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if not name:
            continue
        n = {
            "name": str(name).strip(),
            "dn": ws.cell(row=r, column=2).value,
            "status": ws.cell(row=r, column=3).value or "online",
            "canvas_x": ws.cell(row=r, column=4).value,
            "canvas_y": ws.cell(row=r, column=5).value,
            "group_name": ws.cell(row=r, column=6).value,
            "attrs": {},
        }
        for i, fk in enumerate(field_keys):
            v = ws.cell(row=r, column=field_start_col + i).value
            if v is not None and v != "":
                n["attrs"][fk] = str(v) if not isinstance(v, (int, float, bool)) else v
        nodes.append(n)
    return nodes


def _read_edge_sheet(ws) -> list:
    field_start_col = len(EDGE_FIXED_HEADERS) + 1
    field_keys = []
    c = field_start_col
    while ws.cell(row=1, column=c).value:
        field_keys.append(str(ws.cell(row=1, column=c).value).strip())
        c += 1

    edges = []
    for r in range(2, ws.max_row + 1):
        src = ws.cell(row=r, column=1).value
        tgt = ws.cell(row=r, column=2).value
        if not src or not tgt:
            continue
        e = {
            "source_name": str(src).strip(),
            "target_name": str(tgt).strip(),
            "status": ws.cell(row=r, column=3).value or "online",
            "attrs": {},
        }
        for i, fk in enumerate(field_keys):
            v = ws.cell(row=r, column=field_start_col + i).value
            if v is not None and v != "":
                e["attrs"][fk] = str(v) if not isinstance(v, (int, float, bool)) else v
        edges.append(e)
    return edges


def _read_node_group_sheet(ws, errors: list) -> list:
    groups = []
    for r in range(2, ws.max_row + 1):
        gname = ws.cell(row=r, column=1).value
        if not gname:
            continue
        row_hint = f"Sheet '{ws.title}' 第 {r} 行"
        attrs_cell = ws.cell(row=r, column=8).value or ""
        attr_strategies = []
        if attrs_cell:
            for line in str(attrs_cell).split(RECORD_SEP):
                line = line.strip()
                if not line:
                    continue
                try:
                    attr_strategies.append(parse_attr_strategy_row(line, row_hint))
                except ExcelValidationError as e:
                    errors.append(str(e))
        groups.append({
            "group_name": str(gname).strip(),
            "node_type_code": ws.cell(row=r, column=2).value,
            "node_count": ws.cell(row=r, column=3).value,
            "name_template": ws.cell(row=r, column=4).value,
            "canvas_x": ws.cell(row=r, column=6).value,
            "canvas_y": ws.cell(row=r, column=7).value,
            "attr_strategies": attr_strategies,
        })
    return groups


def _read_node_group_edge_strategy_sheet(ws, errors: list) -> list:
    strategies = []
    for r in range(2, ws.max_row + 1):
        src = ws.cell(row=r, column=1).value
        if not src:
            continue
        row_hint = f"Sheet '{ws.title}' 第 {r} 行"
        line = FIELD_SEP.join([
            str(ws.cell(row=r, column=c).value or "") for c in range(1, 7)
        ])
        try:
            strategies.append(parse_edge_strategy_row(line, row_hint))
        except ExcelValidationError as e:
            errors.append(str(e))
    return strategies


def _read_alarm_sheet(ws, fixed_col_count: int, alarm_field_keys_start: int) -> list:
    field_keys = []
    c = alarm_field_keys_start
    while ws.cell(row=1, column=c).value:
        field_keys.append(str(ws.cell(row=1, column=c).value).strip())
        c += 1

    rows = []
    for r in range(2, ws.max_row + 1):
        first_cell = ws.cell(row=r, column=1).value
        if not first_cell:
            continue
        row_data = {"attrs": {}}
        for i in range(1, fixed_col_count + 1):
            row_data[f"_col_{i}"] = ws.cell(row=r, column=i).value
        for i, fk in enumerate(field_keys):
            v = ws.cell(row=r, column=alarm_field_keys_start + i).value
            if v is not None and v != "":
                row_data["attrs"][fk] = str(v) if not isinstance(v, (int, float, bool)) else v
        rows.append(row_data)
    return rows


def parse_workbook(file_like) -> ParseResult:
    """解析上传的 xlsx。抛 ExcelValidationError 表示致命错。行级错误进 result.errors。"""
    try:
        wb = _load_workbook(file_like, read_only=False, data_only=True)
    except Exception as e:
        raise ExcelValidationError(f"Excel 打开失败：{e}") from e

    result = ParseResult()

    if META_SHEET_NAME not in wb.sheetnames:
        raise ExcelValidationError(f"缺少 '{META_SHEET_NAME}' Sheet")
    result.meta = _read_meta_sheet(wb[META_SHEET_NAME])

    data_sheet_count = 0

    for ws in wb.worksheets:
        title = ws.title
        if title.startswith("_") or title == META_SHEET_NAME:
            continue

        node_type_code = _extract_marker(ws["A1"], NODE_TYPE_MARKER)
        edge_type_code = _extract_marker(ws["A1"], EDGE_TYPE_MARKER)

        if node_type_code:
            result.nodes_by_type_code[node_type_code] = _read_node_sheet(ws, node_type_code)
            data_sheet_count += 1
            continue

        if edge_type_code:
            result.edges_by_type_code[edge_type_code] = _read_edge_sheet(ws)
            data_sheet_count += 1
            continue

        if _extract_marker(ws["A1"], NODE_GROUP_MARKER):
            result.node_groups = _read_node_group_sheet(ws, result.errors)
            data_sheet_count += 1
            continue

        if _extract_marker(ws["A1"], NODE_GROUP_EDGE_STRATEGY_MARKER):
            result.node_group_edge_strategies = _read_node_group_edge_strategy_sheet(ws, result.errors)
            data_sheet_count += 1
            continue

        if _extract_marker(ws["A1"], NODE_ALARM_MARKER):
            rows = _read_alarm_sheet(ws, len(NODE_ALARM_FIXED_HEADERS), len(NODE_ALARM_FIXED_HEADERS) + 1)
            for r in rows:
                result.node_alarms.append({
                    "node_type_code": r.get("_col_1"),
                    "node_name": r.get("_col_2"),
                    "alarm_index": r.get("_col_3"),
                    "attrs": r["attrs"],
                })
            data_sheet_count += 1
            continue

        if _extract_marker(ws["A1"], NODE_GROUP_ALARM_MARKER):
            rows = _read_alarm_sheet(ws, len(NODE_GROUP_ALARM_FIXED_HEADERS), len(NODE_GROUP_ALARM_FIXED_HEADERS) + 1)
            for r in rows:
                result.node_group_alarms.append({
                    "group_name": r.get("_col_1"),
                    "alarm_index": r.get("_col_2"),
                    "attrs": r["attrs"],
                })
            data_sheet_count += 1
            continue

        # 未识别的 Sheet 静默跳过

    if data_sheet_count == 0:
        raise ExcelValidationError("Excel 中未找到任何数据 Sheet")

    return result
