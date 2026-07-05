import re
import uuid
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook

from app.db.connection import connect, transaction
from app.admin.schemas.node_type import (
    NodeTypeCreate,
    NodeTypeUpdate,
    NodeTypeDetail,
    NodeTypeItem,
    NodeTypeDomainsUpdate,
    NodeTypeBatchDomainsUpdate,
    NodeTypeFieldItem,
    NodeTypeFieldInput,
    NodeTypeBatchDelete,
    EdgeTypeBatchDelete,
    TypeExportRequest,
    TypeImportResult,
    TypeImportPreview,
    TypeImportPreviewItem,
    EdgeTypeCreate,
    EdgeTypeUpdate,
    EdgeTypeDetail,
    EdgeTypeItem,
    EdgeTypeFieldItem,
    EdgeTypeFieldInput,
    FieldDeleteImpactRequest,
    FieldDeleteImpactItem,
    FieldDeleteImpactResponse,
    EdgeTypeImportPreviewItem,
    EdgeTypeImportPreview,
    EdgeTypeImportResult,
)

router = APIRouter(prefix="/admin/api", tags=["类型"])


def _new_id() -> str:
    return f"ntype_{uuid.uuid4().hex[:12]}"


def _new_edge_id() -> str:
    return f"etype_{uuid.uuid4().hex[:12]}"


def _row_to_node_type_item(row) -> NodeTypeItem:
    return NodeTypeItem(
        id=row["id"],
        code=row["code"],
        name=row["name"],
        category=row["category"],
        description=row["description"],
        created_at=row["created_at"],
        updated_at=row["updated_at"],
    )


def _row_to_edge_type_item(row) -> EdgeTypeItem:
    return EdgeTypeItem(
        id=row["id"],
        code=row["code"],
        name=row["name"],
        semantic=row["semantic"],
        directed=bool(row["directed"]),
        exclusive_target=bool(row["exclusive_target"]),
        allow_source_type_codes=row["allow_source_type_codes"],
        allow_target_type_codes=row["allow_target_type_codes"],
        line_style=row["line_style"],
        color=row["color"],
        description=row["description"],
        created_at=row["created_at"],
        updated_at=row["updated_at"],
    )


def _get_node_type_fields(conn, node_type_id: str) -> list[NodeTypeFieldItem]:
    rows = conn.execute(
        "SELECT * FROM node_type_fields WHERE node_type_id = ? ORDER BY sort_order",
        (node_type_id,),
    ).fetchall()
    return [
        NodeTypeFieldItem(
            id=r["id"],
            node_type_id=r["node_type_id"],
            field_key=r["field_key"],
            field_label=r["field_label"],
            field_type=r["field_type"],
            max_length=r["max_length"],
            default_value=r["default_value"],
            options=r["options"],
            required=bool(r["required"]),
            sort_order=r["sort_order"],
        )
        for r in rows
    ]


def _sync_node_type_fields(conn, node_type_id: str, incoming: list) -> None:
    """整批同步 node_type_fields，按 field_key 做 diff。

    incoming: list[NodeTypeFieldInput]
    事务内顺序：delete → update → insert
    删除字段时一并清理同类型节点的孤儿 node_attrs 行。
    """
    # 校验 incoming 内 field_key 不重复
    seen_keys = set()
    for f in incoming:
        if f.field_key in seen_keys:
            raise HTTPException(
                status_code=400,
                detail={"code": 40206, "message": f"字段 Key 重复: {f.field_key}"},
            )
        seen_keys.add(f.field_key)

    existing_keys = {
        r["field_key"]
        for r in conn.execute(
            "SELECT field_key FROM node_type_fields WHERE node_type_id = ?",
            (node_type_id,),
        ).fetchall()
    }

    incoming_keys = {f.field_key for f in incoming}
    to_delete = existing_keys - incoming_keys

    # DELETE：先删字段定义，再清理孤儿 node_attrs（仅同类型节点）
    for k in to_delete:
        conn.execute(
            "DELETE FROM node_type_fields WHERE node_type_id = ? AND field_key = ?",
            (node_type_id, k),
        )
        conn.execute(
            """DELETE FROM node_attrs
               WHERE field_key = ?
                 AND node_id IN (SELECT id FROM nodes WHERE node_type_id = ?)""",
            (k, node_type_id),
        )

    # UPDATE：按 sort_order = 数组下标重写
    for idx, f in enumerate(incoming):
        if f.field_key not in existing_keys:
            continue
        conn.execute(
            """UPDATE node_type_fields
               SET field_label = ?, field_type = ?, max_length = ?,
                   default_value = ?, options = ?, required = ?, sort_order = ?
               WHERE node_type_id = ? AND field_key = ?""",
            (f.field_label, f.field_type, f.max_length, f.default_value,
             f.options, int(f.required), idx,
             node_type_id, f.field_key),
        )

    # INSERT：sort_order = 数组下标
    for idx, f in enumerate(incoming):
        if f.field_key in existing_keys:
            continue
        conn.execute(
            """INSERT INTO node_type_fields
               (node_type_id, field_key, field_label, field_type, max_length,
                default_value, options, required, sort_order)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (node_type_id, f.field_key, f.field_label, f.field_type,
             f.max_length, f.default_value, f.options, int(f.required), idx),
        )


def _sync_edge_type_fields(conn, edge_type_id: str, incoming: list) -> None:
    """整批同步 edge_type_fields，按 field_key 做 diff。

    incoming: list[EdgeTypeFieldInput]
    事务内顺序：delete → update → insert
    删除字段时一并清理同类型边的孤儿 edge_attrs 行。
    """
    # 校验 incoming 内 field_key 不重复
    seen_keys = set()
    for f in incoming:
        if f.field_key in seen_keys:
            raise HTTPException(
                status_code=400,
                detail={"code": 40306, "message": f"字段 Key 重复: {f.field_key}"},
            )
        seen_keys.add(f.field_key)

    existing_keys = {
        r["field_key"]
        for r in conn.execute(
            "SELECT field_key FROM edge_type_fields WHERE edge_type_id = ?",
            (edge_type_id,),
        ).fetchall()
    }

    incoming_keys = {f.field_key for f in incoming}
    to_delete = existing_keys - incoming_keys

    # DELETE：先删字段定义，再清理孤儿 edge_attrs（仅同类型边）
    for k in to_delete:
        conn.execute(
            "DELETE FROM edge_type_fields WHERE edge_type_id = ? AND field_key = ?",
            (edge_type_id, k),
        )
        conn.execute(
            """DELETE FROM edge_attrs
               WHERE field_key = ?
                 AND edge_id IN (SELECT id FROM edges WHERE edge_type_id = ?)""",
            (k, edge_type_id),
        )

    # UPDATE：按 sort_order = 数组下标重写
    for idx, f in enumerate(incoming):
        if f.field_key not in existing_keys:
            continue
        conn.execute(
            """UPDATE edge_type_fields
               SET field_label = ?, field_type = ?, max_length = ?,
                   default_value = ?, options = ?, required = ?, sort_order = ?
               WHERE edge_type_id = ? AND field_key = ?""",
            (f.field_label, f.field_type, f.max_length, f.default_value,
             f.options, int(f.required), idx,
             edge_type_id, f.field_key),
        )

    # INSERT：sort_order = 数组下标
    for idx, f in enumerate(incoming):
        if f.field_key in existing_keys:
            continue
        conn.execute(
            """INSERT INTO edge_type_fields
               (edge_type_id, field_key, field_label, field_type, max_length,
                default_value, options, required, sort_order)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (edge_type_id, f.field_key, f.field_label, f.field_type,
             f.max_length, f.default_value, f.options, int(f.required), idx),
        )


def _get_edge_type_fields(conn, edge_type_id: str) -> list[EdgeTypeFieldItem]:
    rows = conn.execute(
        "SELECT * FROM edge_type_fields WHERE edge_type_id = ? ORDER BY sort_order",
        (edge_type_id,),
    ).fetchall()
    return [
        EdgeTypeFieldItem(
            id=r["id"],
            edge_type_id=r["edge_type_id"],
            field_key=r["field_key"],
            field_label=r["field_label"],
            field_type=r["field_type"],
            max_length=r["max_length"],
            default_value=r["default_value"],
            options=r["options"],
            required=bool(r["required"]),
            sort_order=r["sort_order"],
        )
        for r in rows
    ]


# ============== node_types ==============

@router.get("/node-types")
def list_node_types(domain_id: str = Query(None, description="按域过滤，NULL=返回全部")) -> dict:
    with connect() as conn:
        if domain_id:
            has_bindings = conn.execute(
                "SELECT COUNT(*) AS cnt FROM domain_node_types WHERE domain_id = ?",
                (domain_id,),
            ).fetchone()["cnt"] > 0
            if has_bindings:
                rows = conn.execute("""
                    SELECT nt.* FROM node_types nt
                    INNER JOIN domain_node_types dnt ON dnt.node_type_id = nt.id
                    WHERE dnt.domain_id = ?
                    ORDER BY nt.category, nt.name
                """, (domain_id,)).fetchall()
            else:
                rows = conn.execute(
                    "SELECT * FROM node_types ORDER BY category, name"
                ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM node_types ORDER BY category, name"
            ).fetchall()
        items = []
        for r in rows:
            dom_rows = conn.execute("""
                SELECT d.id, d.name FROM domains d
                INNER JOIN domain_node_types dnt ON dnt.domain_id = d.id
                WHERE dnt.node_type_id = ?
            """, (r["id"],)).fetchall()
            item = NodeTypeDetail(
                id=r["id"],
                code=r["code"],
                name=r["name"],
                category=r["category"],
                description=r["description"],
                domain_ids=[dr["id"] for dr in dom_rows],
                domain_names=[dr["name"] for dr in dom_rows],
                created_at=r["created_at"],
                updated_at=r["updated_at"],
                fields=_get_node_type_fields(conn, r["id"]),
            )
            items.append(item.model_dump(mode="json", by_alias=True))
        return {"code": 0, "data": {"items": items, "total": len(items)}, "message": "ok"}


@router.get("/node-types/{type_id}")
def get_node_type(type_id: str) -> dict:
    with connect() as conn:
        row = conn.execute(
            "SELECT * FROM node_types WHERE id = ?", (type_id,)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail={"code": 40201, "message": "类型不存在"})
        dom_rows = conn.execute("""
            SELECT d.id, d.name FROM domains d
            INNER JOIN domain_node_types dnt ON dnt.domain_id = d.id
            WHERE dnt.node_type_id = ?
        """, (type_id,)).fetchall()
        item = NodeTypeDetail(
            id=row["id"],
            code=row["code"],
            name=row["name"],
            category=row["category"],
            description=row["description"],
            domain_ids=[dr["id"] for dr in dom_rows],
            domain_names=[dr["name"] for dr in dom_rows],
            created_at=row["created_at"],
            updated_at=row["updated_at"],
            fields=_get_node_type_fields(conn, row["id"]),
        )
        return {"code": 0, "data": item.model_dump(mode="json", by_alias=True), "message": "ok"}


@router.post("/node-types")
def create_node_type(data: NodeTypeCreate) -> dict:
    type_id = _new_id()
    with transaction() as conn:
        existing = conn.execute(
            "SELECT id FROM node_types WHERE code = ?", (data.code,)
        ).fetchone()
        if existing:
            raise HTTPException(status_code=409, detail={"code": 40202, "message": "类型代码已存在"})
        conn.execute(
            """INSERT INTO node_types (id, code, name, category, description)
               VALUES (?, ?, ?, ?, ?)""",
            (type_id, data.code, data.name, data.category, data.description),
        )
        if data.fields is not None:
            _sync_node_type_fields(conn, type_id, data.fields)
        if data.domain_ids is not None:
            conn.execute("DELETE FROM domain_node_types WHERE node_type_id = ?", (type_id,))
            for dom_id in data.domain_ids:
                conn.execute(
                    "INSERT OR IGNORE INTO domain_node_types (domain_id, node_type_id) VALUES (?, ?)",
                    (dom_id, type_id),
                )
    return {"code": 0, "data": {"id": type_id}, "message": "ok"}


# ============== node_type <-> domain 关联 ==============

@router.put("/node-types/domains")
def batch_update_node_type_domains(data: NodeTypeBatchDomainsUpdate) -> dict:
    with transaction() as conn:
        for type_id in data.node_type_ids:
            existing = conn.execute(
                "SELECT id FROM node_types WHERE id = ?", (type_id,)
            ).fetchone()
            if not existing:
                continue
            conn.execute("DELETE FROM domain_node_types WHERE node_type_id = ?", (type_id,))
            for dom_id in data.domain_ids:
                conn.execute(
                    "INSERT OR IGNORE INTO domain_node_types (domain_id, node_type_id) VALUES (?, ?)",
                    (dom_id, type_id),
                )
    return {
        "code": 0,
        "data": {"nodeTypeIds": data.node_type_ids, "domainIds": data.domain_ids},
        "message": "ok",
    }


@router.put("/node-types/{type_id}/domains")
def update_node_type_domains(type_id: str, data: NodeTypeDomainsUpdate) -> dict:
    with transaction() as conn:
        existing = conn.execute(
            "SELECT id FROM node_types WHERE id = ?", (type_id,)
        ).fetchone()
        if not existing:
            raise HTTPException(
                status_code=404,
                detail={"code": 40201, "message": "类型不存在"},
            )
        conn.execute("DELETE FROM domain_node_types WHERE node_type_id = ?", (type_id,))
        for dom_id in data.domain_ids:
            conn.execute(
                "INSERT OR IGNORE INTO domain_node_types (domain_id, node_type_id) VALUES (?, ?)",
                (dom_id, type_id),
            )
    return {"code": 0, "data": {"id": type_id}, "message": "ok"}


@router.put("/node-types/{type_id}")
def update_node_type(type_id: str, data: NodeTypeUpdate) -> dict:
    raw = data.model_dump(exclude_unset=True)
    fields_payload = data.fields if "fields" in raw else None
    domain_ids_payload = data.domain_ids if "domain_ids" in raw else None
    raw.pop("fields", None)
    raw.pop("domain_ids", None)

    # 防御：空 body {} 拒绝
    if not raw and fields_payload is None and domain_ids_payload is None:
        raise HTTPException(status_code=400, detail={"code": 40203, "message": "无更新字段"})

    with transaction() as conn:
        existing = conn.execute(
            "SELECT id FROM node_types WHERE id = ?", (type_id,)
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail={"code": 40201, "message": "类型不存在"})

        if raw:
            set_clause = ", ".join(f"{k} = ?" for k in raw.keys())
            conn.execute(
                f"UPDATE node_types SET {set_clause}, updated_at = datetime('now') WHERE id = ?",
                (*raw.values(), type_id),
            )

        if fields_payload is not None:
            _sync_node_type_fields(conn, type_id, fields_payload)

        if domain_ids_payload is not None:
            conn.execute("DELETE FROM domain_node_types WHERE node_type_id = ?", (type_id,))
            for dom_id in domain_ids_payload:
                conn.execute(
                    "INSERT OR IGNORE INTO domain_node_types (domain_id, node_type_id) VALUES (?, ?)",
                    (dom_id, type_id),
                )

    return {"code": 0, "data": {"id": type_id}, "message": "ok"}


@router.post("/node-types/batch-delete")
def batch_delete_node_types(data: NodeTypeBatchDelete) -> dict:
    deleted_count = 0
    skipped: list[dict] = []
    with transaction() as conn:
        for type_id in data.ids:
            existing = conn.execute(
                "SELECT id FROM node_types WHERE id = ?", (type_id,)
            ).fetchone()
            if not existing:
                skipped.append({"id": type_id, "reason": "类型不存在"})
                continue
            in_use = conn.execute(
                "SELECT id FROM nodes WHERE node_type_id = ? LIMIT 1", (type_id,)
            ).fetchone()
            if in_use:
                skipped.append({"id": type_id, "reason": "该类型已被节点使用，无法删除"})
                continue
            in_use = conn.execute(
                "SELECT id FROM node_groups WHERE node_type_id = ? LIMIT 1", (type_id,)
            ).fetchone()
            if in_use:
                skipped.append({"id": type_id, "reason": "该类型已被节点组使用，无法删除"})
                continue
            conn.execute("DELETE FROM node_type_fields WHERE node_type_id = ?", (type_id,))
            conn.execute("DELETE FROM node_types WHERE id = ?", (type_id,))
            deleted_count += 1
    return {"code": 0, "data": {"deletedCount": deleted_count, "skipped": skipped}, "message": "ok"}


_SHEET_INVALID_CHARS = re.compile(r'[\\\*/\[\]\?:]')


def _safe_sheet_name(code: str) -> str:
    name = _SHEET_INVALID_CHARS.sub('_', code)
    if len(name) > 31:
        name = name[:28] + "..."
    return name


def _build_header_map(ws) -> dict[str, int]:
    """Read row 1 and map header name -> column index. Returns empty dict if no header."""
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not row:
        return {}
    return {str(v).strip(): i for i, v in enumerate(row) if v is not None and str(v).strip()}


def _col(headers: dict[str, int], name: str, row: tuple) -> Optional[str]:
    """Get a cell value by header name, returning None if the column is missing."""
    idx = headers.get(name)
    if idx is None or idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    return str(val).strip() or None


def _build_node_types_excel(items: list[dict]) -> BytesIO:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "类型汇总"
    ws1.append(["编码", "名称", "分类", "所属网管/设备",
                "描述", "创建时间", "更新时间"])
    for item in items:
        dom_names = item.get("domainNames") or []
        ws1.append([
            item.get("code"), item.get("name"), item.get("category"),
            "|".join(dom_names) if dom_names else None,
            item.get("description"), item.get("createdAt"), item.get("updatedAt"),
        ])

    for item in items:
        fields = item.get("fields") or []
        sheet_name = _safe_sheet_name(item.get("code", item.get("id", "unknown")))
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["字段标识", "显示名称", "字段类型", "最大长度",
                    "默认值", "选项", "必填", "排序"])
        for f in fields:
            ws.append([
                f.get("fieldKey"), f.get("fieldLabel"),
                f.get("fieldType"), f.get("maxLength"), f.get("defaultValue"),
                f.get("options"), "是" if f.get("required") else "否",
                f.get("sortOrder"),
            ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _build_edge_types_excel(items: list) -> BytesIO:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "边类型汇总"
    ws1.append(["Code", "名称", "语义", "有向", "唯一目标",
                "允许源类型", "允许目标类型", "线条样式", "颜色",
                "描述", "字段数", "创建时间", "更新时间"])
    for item in items:
        ws1.append([
            item.get("code"),
            item.get("name"),
            item.get("semantic"),
            "是" if item.get("directed") else "否",
            "是" if item.get("exclusiveTarget") else "否",
            item.get("allowSourceTypeCodes"),
            item.get("allowTargetTypeCodes"),
            item.get("lineStyle"),
            item.get("color"),
            item.get("description"),
            len(item.get("fields") or []),
            item.get("createdAt"),
            item.get("updatedAt"),
        ])

    for item in items:
        fields = item.get("fields") or []
        sheet_name = _safe_sheet_name(item.get("code", item.get("id", "unknown")))
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["字段标识", "显示名称", "字段类型", "最大长度",
                    "默认值", "选项", "必填", "排序"])
        for f in fields:
            ws.append([
                f.get("fieldKey"),
                f.get("fieldLabel"),
                f.get("fieldType"),
                f.get("maxLength"),
                f.get("defaultValue"),
                f.get("options"),
                "是" if f.get("required") else "否",
                f.get("sortOrder"),
            ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.post("/node-types/export")
def export_node_types(data: TypeExportRequest):
    with connect() as conn:
        if data.ids:
            placeholders = ",".join("?" for _ in data.ids)
            rows = conn.execute(
                f"SELECT * FROM node_types WHERE id IN ({placeholders}) ORDER BY category, name",
                tuple(data.ids),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM node_types ORDER BY category, name"
            ).fetchall()
        items = []
        for r in rows:
            dom_rows = conn.execute("""
                SELECT d.id, d.name FROM domains d
                INNER JOIN domain_node_types dnt ON dnt.domain_id = d.id
                WHERE dnt.node_type_id = ?
            """, (r["id"],)).fetchall()
            item = NodeTypeDetail(
                id=r["id"],
                code=r["code"],
                name=r["name"],
                category=r["category"],
                description=r["description"],
                domain_ids=[dr["id"] for dr in dom_rows],
                domain_names=[dr["name"] for dr in dom_rows],
                created_at=r["created_at"],
                updated_at=r["updated_at"],
                fields=_get_node_type_fields(conn, r["id"]),
            )
            items.append(item.model_dump(mode="json", by_alias=True))

    excel = _build_node_types_excel(items)
    return StreamingResponse(
        excel,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=node-types-export.xlsx"},
    )


def _load_import_workbook(contents: bytes, expected_sheet: str = "类型汇总") -> Workbook:
    try:
        wb = load_workbook(filename=BytesIO(contents))
    except Exception:
        raise HTTPException(
            status_code=400,
            detail={"code": 40211, "message": "文件无法解析，请确认是有效的 xlsx 文件"},
        )
    if expected_sheet not in wb.sheetnames:
        raise HTTPException(
            status_code=400,
            detail={"code": 40212, "message": f"缺少「{expected_sheet}」Sheet"},
        )
    return wb


@router.post("/node-types/import/preview")
async def preview_node_types_import(file: UploadFile = File(...)):
    if not file.filename or not file.filename.endswith('.xlsx'):
        raise HTTPException(
            status_code=400,
            detail={"code": 40210, "message": "仅支持 .xlsx 文件"},
        )

    contents = await file.read()
    wb = _load_import_workbook(contents)
    ws = wb["类型汇总"]

    to_create: list[dict] = []
    to_update: list[dict] = []
    errors: list[str] = []

    with connect() as conn:
        headers = _build_header_map(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None or v == '' for v in row):
                break
            code = _col(headers, "编码", row)
            name = _col(headers, "名称", row)
            category = _col(headers, "分类", row)

            if not code or not name or not category:
                errors.append(f"编码={code or '(空)'} 缺少必填字段（编码/名称/分类），跳过")
                continue

            existing = conn.execute(
                "SELECT id, name FROM node_types WHERE code = ?", (code,)
            ).fetchone()

            if existing:
                to_update.append({
                    "code": code,
                    "name": name,
                    "old_name": existing["name"],
                })
            else:
                to_create.append({"code": code, "name": name})

    return {
        "code": 0,
        "data": {
            "toCreate": to_create,
            "toUpdate": to_update,
            "errors": errors,
        },
        "message": "ok",
    }


@router.post("/node-types/import")
async def import_node_types(file: UploadFile = File(...)):
    if not file.filename or not file.filename.endswith('.xlsx'):
        raise HTTPException(
            status_code=400,
            detail={"code": 40210, "message": "仅支持 .xlsx 文件"},
        )

    contents = await file.read()
    wb = _load_import_workbook(contents)
    ws = wb["类型汇总"]
    result = TypeImportResult()

    with transaction() as conn:
        headers = _build_header_map(ws)
        # type_id -> [domain_id, ...]，key 存在与否决定"是否处理关联"
        pending_links: dict[str, list] = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None or v == '' for v in row):
                break
            code = _col(headers, "编码", row)
            name = _col(headers, "名称", row)
            category = _col(headers, "分类", row)

            if not code or not name or not category:
                result.errors.append(f"编码={code or '(空)'} 缺少必填字段（编码/名称/分类），跳过")
                continue

            description = _col(headers, "描述", row)

            existing = conn.execute(
                "SELECT id FROM node_types WHERE code = ?", (code,)
            ).fetchone()

            if existing:
                type_id = existing["id"]
                conn.execute(
                    """UPDATE node_types SET name=?, category=?, description=?,
                       updated_at=datetime('now')
                       WHERE id=?""",
                    (name, category, description, type_id),
                )
                result.updated += 1
            else:
                type_id = _new_id()
                conn.execute(
                    """INSERT INTO node_types
                       (id, code, name, category, description)
                       VALUES (?,?,?,?,?)""",
                    (type_id, code, name, category, description),
                )
                result.created += 1

            # 解析"所属网管/设备"列 → 收集待关联 domain_id
            if "所属网管/设备" in headers:
                cell = _col(headers, "所属网管/设备", row)
                if cell:
                    names = [n.strip() for n in cell.split("|") if n.strip()]
                    dom_ids: list = []
                    for dname in names:
                        drow = conn.execute(
                            "SELECT id FROM domains WHERE name = ?", (dname,)
                        ).fetchone()
                        if drow:
                            dom_ids.append(drow["id"])
                        else:
                            result.errors.append(
                                f"[{code}] 网管/设备 '{dname}' 不存在，关联跳过"
                            )
                    pending_links[type_id] = dom_ids
                else:
                    pending_links[type_id] = []

            sheet_name = _safe_sheet_name(code)
            if sheet_name in wb.sheetnames:
                conn.execute(
                    "DELETE FROM node_type_fields WHERE node_type_id = ?",
                    (type_id,),
                )
                fheaders = _build_header_map(wb[sheet_name])
                seen_fields: set[str] = set()
                for frow in wb[sheet_name].iter_rows(min_row=2, values_only=True):
                    if all(v is None or v == '' for v in frow):
                        break
                    fkey = _col(fheaders, "字段标识", frow)
                    flabel = _col(fheaders, "显示名称", frow)
                    ftype_raw = _col(fheaders, "字段类型", frow)
                    if not fkey or not flabel or not ftype_raw:
                        continue
                    if fkey in seen_fields:
                        result.errors.append(
                            f"[{code}] 字段标识 {fkey} 重复，跳过"
                        )
                        continue
                    seen_fields.add(fkey)
                    ftype = ftype_raw.strip().lower()
                    if ftype not in ('text', 'number', 'select', 'boolean', 'array'):
                        result.errors.append(
                            f"[{code}] 字段 {fkey} 类型 '{ftype_raw}' 无效，仅支持 text/number/select/boolean/array，跳过"
                        )
                        continue
                    maxlen_raw = _col(fheaders, "最大长度", frow)
                    if ftype == 'text':
                        # text 类型 max_length 兜底 255
                        try:
                            maxlen = int(maxlen_raw) if maxlen_raw else 255
                            if maxlen < 1:
                                maxlen = 255
                        except (ValueError, TypeError):
                            maxlen = 255
                    else:
                        try:
                            maxlen = int(maxlen_raw) if maxlen_raw else None
                        except (ValueError, TypeError):
                            maxlen = None
                    defval = _col(fheaders, "默认值", frow)
                    opts = _col(fheaders, "选项", frow)
                    req_raw = _col(fheaders, "必填", frow) or ""
                    req = 1 if req_raw == "是" else 0
                    sort_raw = _col(fheaders, "排序", frow)
                    sort = int(sort_raw) if sort_raw and sort_raw.isdigit() else 0

                    # array 类型：default_value 必须是合法 JSON array
                    if ftype == 'array' and defval:
                        import json as _json
                        try:
                            _parsed = _json.loads(defval)
                            if not isinstance(_parsed, list):
                                result.errors.append(
                                    f"[{code}] 字段 {fkey} 的默认值必须是 JSON array，跳过"
                                )
                                continue
                        except _json.JSONDecodeError:
                            result.errors.append(
                                f"[{code}] 字段 {fkey} 的默认值不是合法 JSON，跳过"
                            )
                            continue

                    conn.execute(
                        """INSERT INTO node_type_fields
                           (node_type_id, field_key, field_label, field_type,
                            max_length, default_value, options, required, sort_order)
                           VALUES (?,?,?,?,?,?,?,?,?)""",
                        (type_id, fkey, flabel, ftype,
                         maxlen, defval, opts, req, sort),
                    )
                    result.total_fields += 1

        # 一次性写入 domain_node_types
        for type_id, dom_ids in pending_links.items():
            conn.execute("DELETE FROM domain_node_types WHERE node_type_id = ?", (type_id,))
            for did in dom_ids:
                conn.execute(
                    "INSERT OR IGNORE INTO domain_node_types (domain_id, node_type_id) VALUES (?, ?)",
                    (did, type_id),
                )

    return {
        "code": 0,
        "data": result.model_dump(mode="json", by_alias=True),
        "message": "ok",
    }


@router.delete("/node-types/{type_id}")
def delete_node_type(type_id: str) -> dict:
    with transaction() as conn:
        existing = conn.execute(
            "SELECT id FROM node_types WHERE id = ?", (type_id,)
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail={"code": 40201, "message": "类型不存在"})
        in_use = conn.execute(
            "SELECT id FROM nodes WHERE node_type_id = ? LIMIT 1", (type_id,)
        ).fetchone()
        if in_use:
            raise HTTPException(status_code=409, detail={"code": 40201, "message": "该类型已被节点使用，无法删除"})
        in_use = conn.execute(
            "SELECT id FROM node_groups WHERE node_type_id = ? LIMIT 1", (type_id,)
        ).fetchone()
        if in_use:
            raise HTTPException(status_code=409, detail={"code": 40201, "message": "该类型已被节点组使用，无法删除"})
        conn.execute("DELETE FROM node_types WHERE id = ?", (type_id,))
    return {"code": 0, "data": {"id": type_id}, "message": "ok"}


@router.post("/node-types/{type_id}/fields/delete-impact")
def get_node_type_field_delete_impact(
    type_id: str, payload: FieldDeleteImpactRequest
) -> dict:
    """统计若删除 fieldKeys 中的字段，会影响多少同类型节点的 node_attrs。"""
    with connect() as conn:
        existing = conn.execute(
            "SELECT id FROM node_types WHERE id = ?", (type_id,)
        ).fetchone()
        if not existing:
            raise HTTPException(
                status_code=404, detail={"code": 40201, "message": "类型不存在"}
            )
        items = []
        for k in payload.field_keys:
            count = conn.execute(
                """SELECT COUNT(*) AS cnt FROM node_attrs a
                   JOIN nodes n ON a.node_id = n.id
                   WHERE n.node_type_id = ? AND a.field_key = ?""",
                (type_id, k),
            ).fetchone()["cnt"]
            items.append(
                FieldDeleteImpactItem(field_key=k, affected_node_count=count)
            )
        resp = FieldDeleteImpactResponse(items=items)
    return {"code": 0, "data": resp.model_dump(mode="json", by_alias=True), "message": "ok"}


# ============== edge_types ==============

@router.get("/edge-types")
def list_edge_types() -> dict:
    with connect() as conn:
        rows = conn.execute(
            "SELECT * FROM edge_types ORDER BY name"
        ).fetchall()
        items = []
        for r in rows:
            item = EdgeTypeDetail(
                id=r["id"],
                code=r["code"],
                name=r["name"],
                semantic=r["semantic"],
                directed=bool(r["directed"]),
                exclusive_target=bool(r["exclusive_target"]),
                allow_source_type_codes=r["allow_source_type_codes"],
                allow_target_type_codes=r["allow_target_type_codes"],
                line_style=r["line_style"],
                color=r["color"],
                description=r["description"],
                created_at=r["created_at"],
                updated_at=r["updated_at"],
                fields=_get_edge_type_fields(conn, r["id"]),
            )
            items.append(item.model_dump(mode="json", by_alias=True))
        return {"code": 0, "data": {"items": items, "total": len(items)}, "message": "ok"}


@router.get("/edge-types/{type_id}")
def get_edge_type(type_id: str) -> dict:
    with connect() as conn:
        row = conn.execute(
            "SELECT * FROM edge_types WHERE id = ?", (type_id,)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail={"code": 40301, "message": "类型不存在"})
        item = EdgeTypeDetail(
            id=row["id"],
            code=row["code"],
            name=row["name"],
            semantic=row["semantic"],
            directed=bool(row["directed"]),
            exclusive_target=bool(row["exclusive_target"]),
            allow_source_type_codes=row["allow_source_type_codes"],
            allow_target_type_codes=row["allow_target_type_codes"],
            line_style=row["line_style"],
            color=row["color"],
            description=row["description"],
            created_at=row["created_at"],
            updated_at=row["updated_at"],
            fields=_get_edge_type_fields(conn, row["id"]),
        )
        return {"code": 0, "data": item.model_dump(mode="json", by_alias=True), "message": "ok"}


@router.post("/edge-types")
def create_edge_type(data: EdgeTypeCreate) -> dict:
    type_id = _new_edge_id()
    with transaction() as conn:
        existing = conn.execute(
            "SELECT id FROM edge_types WHERE code = ?", (data.code,)
        ).fetchone()
        if existing:
            raise HTTPException(status_code=409, detail={"code": 40302, "message": "类型代码已存在"})
        conn.execute(
            """INSERT INTO edge_types
               (id, code, name, semantic, directed, exclusive_target,
                allow_source_type_codes, allow_target_type_codes, line_style, color, description)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (type_id, data.code, data.name, data.semantic, int(data.directed),
             int(data.exclusive_target), data.allow_source_type_codes,
             data.allow_target_type_codes, data.line_style, data.color, data.description),
        )
        if data.fields is not None:
            _sync_edge_type_fields(conn, type_id, data.fields)
    return {"code": 0, "data": {"id": type_id}, "message": "ok"}


@router.put("/edge-types/{type_id}")
def update_edge_type(type_id: str, data: EdgeTypeUpdate) -> dict:
    raw = data.model_dump(exclude_unset=True)
    fields_payload = data.fields if "fields" in raw else None  # 用户是否传入了 fields
    raw.pop("fields", None)  # 不放进 UPDATE SQL

    # 防御：空 body {} 拒绝
    if not raw and fields_payload is None:
        raise HTTPException(status_code=400, detail={"code": 40303, "message": "无更新字段"})

    with transaction() as conn:
        existing = conn.execute(
            "SELECT id FROM edge_types WHERE id = ?", (type_id,)
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail={"code": 40301, "message": "类型不存在"})

        if raw:
            set_clause = ", ".join(f"{k} = ?" for k in raw.keys())
            conn.execute(
                f"UPDATE edge_types SET {set_clause}, updated_at = datetime('now') WHERE id = ?",
                (*raw.values(), type_id),
            )

        if fields_payload is not None:
            _sync_edge_type_fields(conn, type_id, fields_payload)

    return {"code": 0, "data": {"id": type_id}, "message": "ok"}


@router.post("/edge-types/batch-delete")
def batch_delete_edge_types(data: EdgeTypeBatchDelete) -> dict:
    deleted_count = 0
    skipped: list[dict] = []
    with transaction() as conn:
        for type_id in data.ids:
            existing = conn.execute(
                "SELECT id FROM edge_types WHERE id = ?", (type_id,)
            ).fetchone()
            if not existing:
                skipped.append({"id": type_id, "reason": "类型不存在"})
                continue
            in_use = conn.execute(
                "SELECT id FROM edges WHERE edge_type_id = ? LIMIT 1", (type_id,)
            ).fetchone()
            if in_use:
                skipped.append({"id": type_id, "reason": "该类型已被边使用，无法删除"})
                continue
            conn.execute("DELETE FROM edge_type_fields WHERE edge_type_id = ?", (type_id,))
            conn.execute("DELETE FROM edge_types WHERE id = ?", (type_id,))
            deleted_count += 1
    return {"code": 0, "data": {"deletedCount": deleted_count, "skipped": skipped}, "message": "ok"}


@router.post("/edge-types/export")
def export_edge_types(data: TypeExportRequest):
    with connect() as conn:
        if data.ids:
            placeholders = ",".join("?" for _ in data.ids)
            rows = conn.execute(
                f"SELECT * FROM edge_types WHERE id IN ({placeholders}) ORDER BY name",
                tuple(data.ids),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM edge_types ORDER BY name"
            ).fetchall()
        items = []
        for r in rows:
            item = EdgeTypeDetail(
                id=r["id"],
                code=r["code"],
                name=r["name"],
                semantic=r["semantic"],
                directed=bool(r["directed"]),
                exclusive_target=bool(r["exclusive_target"]),
                allow_source_type_codes=r["allow_source_type_codes"],
                allow_target_type_codes=r["allow_target_type_codes"],
                line_style=r["line_style"],
                color=r["color"],
                description=r["description"],
                created_at=r["created_at"],
                updated_at=r["updated_at"],
                fields=_get_edge_type_fields(conn, r["id"]),
            )
            items.append(item.model_dump(mode="json", by_alias=True))

    excel = _build_edge_types_excel(items)
    return StreamingResponse(
        excel,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=edge-types-export.xlsx"},
    )


@router.post("/edge-types/import/preview")
async def preview_edge_types_import(file: UploadFile = File(...)):
    if not file.filename or not file.filename.endswith('.xlsx'):
        raise HTTPException(
            status_code=400,
            detail={"code": 40310, "message": "仅支持 .xlsx 文件"},
        )

    contents = await file.read()
    wb = _load_import_workbook(contents, "边类型汇总")
    ws = wb["边类型汇总"]

    to_create: list[dict] = []
    to_update: list[dict] = []
    errors: list[str] = []

    with connect() as conn:
        headers = _build_header_map(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None or v == '' for v in row):
                break
            code = _col(headers, "Code", row)
            name = _col(headers, "名称", row)

            if not code or not name:
                errors.append(f"Code={code or '(空)'} 缺少必填字段（Code/名称），跳过")
                continue

            existing = conn.execute(
                "SELECT id, name FROM edge_types WHERE code = ?", (code,)
            ).fetchone()

            if existing:
                to_update.append({
                    "code": code,
                    "name": name,
                    "old_name": existing["name"],
                })
            else:
                to_create.append({"code": code, "name": name})

    preview = EdgeTypeImportPreview(
        to_create=[EdgeTypeImportPreviewItem(**c) for c in to_create],
        to_update=[EdgeTypeImportPreviewItem(**u) for u in to_update],
        errors=errors,
    )
    return {"code": 0, "data": preview.model_dump(mode="json", by_alias=True), "message": "ok"}


@router.delete("/edge-types/{type_id}")
def delete_edge_type(type_id: str) -> dict:
    with transaction() as conn:
        existing = conn.execute(
            "SELECT id FROM edge_types WHERE id = ?", (type_id,)
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail={"code": 40301, "message": "类型不存在"})
        in_use = conn.execute(
            "SELECT id FROM edges WHERE edge_type_id = ? LIMIT 1", (type_id,)
        ).fetchone()
        if in_use:
            raise HTTPException(status_code=409, detail={"code": 40301, "message": "该类型已被边使用，无法删除"})
        conn.execute("DELETE FROM edge_types WHERE id = ?", (type_id,))
    return {"code": 0, "data": {"id": type_id}, "message": "ok"}


@router.post("/edge-types/{type_id}/fields/delete-impact")
def get_edge_type_field_delete_impact(
    type_id: str, payload: FieldDeleteImpactRequest
) -> dict:
    """统计若删除 fieldKeys 中的字段，会影响多少同类型边的 edge_attrs。"""
    with connect() as conn:
        existing = conn.execute(
            "SELECT id FROM edge_types WHERE id = ?", (type_id,)
        ).fetchone()
        if not existing:
            raise HTTPException(
                status_code=404, detail={"code": 40301, "message": "类型不存在"}
            )
        items = []
        for k in payload.field_keys:
            count = conn.execute(
                """SELECT COUNT(*) AS cnt FROM edge_attrs a
                   JOIN edges e ON a.edge_id = e.id
                   WHERE e.edge_type_id = ? AND a.field_key = ?""",
                (type_id, k),
            ).fetchone()["cnt"]
            items.append(
                FieldDeleteImpactItem(field_key=k, affected_node_count=count)
            )
        resp = FieldDeleteImpactResponse(items=items)
    return {"code": 0, "data": resp.model_dump(mode="json", by_alias=True), "message": "ok"}
