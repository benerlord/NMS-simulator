import pytest

from app.admin._topology_excel import (
    ExcelValidationError,
    sanitize_sheet_name,
    format_attr_strategy_row,
    parse_attr_strategy_row,
    format_edge_strategy_row,
    parse_edge_strategy_row,
    INVALID_SHEET_CHARS,
    SHEET_NAME_MAX_LEN,
)


def test_sanitize_sheet_name_valid_passthrough():
    used = set()
    assert sanitize_sheet_name("路由器", used) == "路由器"
    assert "路由器" in used


def test_sanitize_sheet_name_invalid_chars_replaced():
    used = set()
    assert sanitize_sheet_name("A/B", used) == "A_B"
    assert sanitize_sheet_name("C:D", used) == "C_D"
    assert sanitize_sheet_name("E[F]G", used) == "E_F_G"


def test_sanitize_sheet_name_truncates_to_31():
    used = set()
    long_name = "X" * 40
    assert len(sanitize_sheet_name(long_name, used)) == 31


def test_sanitize_sheet_name_dedupe_suffix():
    used = set()
    a = sanitize_sheet_name("路由器", used)
    b = sanitize_sheet_name("路由器", used)
    assert a == "路由器"
    assert b == "路由器~2"


def test_sanitize_sheet_name_empty_falls_back():
    used = set()
    assert sanitize_sheet_name("", used) == "未命名"


# --- Attr strategy ---

def test_format_attr_strategy_fixed():
    line = format_attr_strategy_row({
        "field_key": "role", "strategy": "fixed", "fixed_value": "core-router",
    })
    assert line == "role|fixed|fixedValue=core-router"


def test_format_attr_strategy_range():
    line = format_attr_strategy_row({
        "field_key": "vlan_id", "strategy": "range", "min": 100, "max": 200,
    })
    assert line == "vlan_id|range|min=100;max=200"


def test_format_attr_strategy_random():
    line = format_attr_strategy_row({
        "field_key": "region", "strategy": "random",
        "pool": ["北京", "上海", "广州"],
    })
    assert line == "region|random|pool=北京;上海;广州"


def test_format_attr_strategy_increment():
    line = format_attr_strategy_row({
        "field_key": "mgmt_ip", "strategy": "increment",
        "base": "10.0.0.1", "step": "1",
    })
    assert line == "mgmt_ip|increment|base=10.0.0.1;step=1"


def test_parse_attr_strategy_round_trip():
    strategies = [
        {"field_key": "vlan_id", "strategy": "range", "min": 100, "max": 200},
        {"field_key": "role", "strategy": "fixed", "fixed_value": "core"},
        {"field_key": "region", "strategy": "random", "pool": ["北京", "上海"]},
        {"field_key": "ip", "strategy": "increment", "base": "10.0.0.1", "step": "1"},
    ]
    for s in strategies:
        line = format_attr_strategy_row(s)
        parsed = parse_attr_strategy_row(line, row_hint="")
        assert parsed["field_key"] == s["field_key"]
        assert parsed["strategy"] == s["strategy"]


def test_parse_attr_strategy_invalid_strategy_raises():
    with pytest.raises(ExcelValidationError):
        parse_attr_strategy_row("role|BOGUS|fixedValue=x", row_hint="第 3 行")


def test_parse_attr_strategy_missing_params_raises():
    with pytest.raises(ExcelValidationError):
        parse_attr_strategy_row("vlan_id|range|min=100", row_hint="第 3 行")


# --- Edge strategy ---

def test_format_edge_strategy_row_all_to_all():
    line = format_edge_strategy_row({
        "source_group_name": "核心路由器组",
        "target_name": "接入交换机组",
        "target_kind": "组",
        "edge_type_code": "link",
        "mode": "all_to_all",
    })
    assert line == "核心路由器组|接入交换机组|组|link|all_to_all|"


def test_format_edge_strategy_row_modulo():
    line = format_edge_strategy_row({
        "source_group_name": "组A",
        "target_name": "组B",
        "target_kind": "组",
        "edge_type_code": "link",
        "mode": "modulo",
        "ratio_k": 4,
    })
    assert line == "组A|组B|组|link|modulo|4"


def test_format_edge_strategy_row_hybrid():
    line = format_edge_strategy_row({
        "source_group_name": "服务器组",
        "target_name": "gateway-01",
        "target_kind": "节点",
        "edge_type_code": "link",
        "mode": "all_to_all",
    })
    assert line == "服务器组|gateway-01|节点|link|all_to_all|"


def test_parse_edge_strategy_round_trip():
    d = {
        "source_group_name": "A",
        "target_name": "B",
        "target_kind": "组",
        "edge_type_code": "link",
        "mode": "modulo",
        "ratio_k": 4,
    }
    line = format_edge_strategy_row(d)
    parsed = parse_edge_strategy_row(line, row_hint="")
    assert parsed == d


def test_parse_edge_strategy_missing_K_for_modulo_raises():
    with pytest.raises(ExcelValidationError):
        parse_edge_strategy_row("A|B|组|link|modulo|", row_hint="第 5 行")


def test_parse_edge_strategy_invalid_target_kind_raises():
    with pytest.raises(ExcelValidationError):
        parse_edge_strategy_row("A|B|BOGUS|link|all_to_all|", row_hint="第 5 行")


def test_parse_edge_strategy_invalid_mode_raises():
    with pytest.raises(ExcelValidationError):
        parse_edge_strategy_row("A|B|组|link|FAKE|", row_hint="第 5 行")


# --- build_workbook ---

import io
from openpyxl import load_workbook

from app.admin._topology_excel import (
    build_workbook,
    INSTRUCTION_SHEET_NAME,
    INDEX_SHEET_NAME,
    META_SHEET_NAME,
    NODE_GROUP_SHEET_NAME,
    NODE_GROUP_EDGE_STRATEGY_SHEET_NAME,
    NODE_GROUP_EDGE_STRATEGY_HEADERS,
    NODE_ALARM_SHEET_NAME,
    NODE_GROUP_ALARM_SHEET_NAME,
    NODE_FIXED_HEADERS,
    NODE_TYPE_MARKER,
    EDGE_FIXED_HEADERS,
    EDGE_TYPE_MARKER,
    NODE_GROUP_EDGE_STRATEGY_MARKER,
)


def _load_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return load_workbook(buf)


def _basic_context():
    return {
        "topology": {"name": "T", "description": "", "version": 1,
                     "domain_name": None, "alarm_schema_code": None},
        "node_types": [],
        "edge_types": [],
        "nodes_by_type_code": {},
        "edges_by_type_code": {},
        "node_groups": [],
        "node_group_edge_strategies": [],
        "alarm_schema_fields": [],
        "node_alarms": [],
        "node_group_alarms": [],
    }


def test_build_workbook_minimal_has_meta_and_index_and_instruction():
    wb = build_workbook(**_basic_context())
    reloaded = _load_bytes(wb)
    assert INSTRUCTION_SHEET_NAME in reloaded.sheetnames
    assert INDEX_SHEET_NAME in reloaded.sheetnames
    assert META_SHEET_NAME in reloaded.sheetnames
    assert NODE_GROUP_SHEET_NAME in reloaded.sheetnames
    assert NODE_GROUP_EDGE_STRATEGY_SHEET_NAME in reloaded.sheetnames
    assert NODE_ALARM_SHEET_NAME not in reloaded.sheetnames
    assert NODE_GROUP_ALARM_SHEET_NAME not in reloaded.sheetnames


def test_build_workbook_meta_sheet_key_value():
    ctx = _basic_context()
    ctx["topology"] = {
        "name": "机房", "description": "描述", "version": 3,
        "domain_name": "网管A", "alarm_schema_code": "s1",
    }
    wb = build_workbook(**ctx)
    reloaded = _load_bytes(wb)
    ws = reloaded[META_SHEET_NAME]
    kv = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
          for r in range(2, ws.max_row + 1)}
    assert kv["拓扑名称"] == "机房"
    assert kv["描述"] == "描述"
    assert kv["版本"] == 3
    assert kv["所属网管/设备"] == "网管A"
    assert kv["告警模板"] == "s1"


def test_build_workbook_per_nodetype_sheet_with_a1_marker():
    ctx = _basic_context()
    ctx["node_types"] = [{"id": "nt_r", "code": "router", "name": "路由器",
                          "fields": [{"field_key": "vlan_id"}]}]
    ctx["nodes_by_type_code"] = {"router": [
        {"id": "n1", "name": "R1", "dn": None, "status": "online",
         "canvas_x": 10.0, "canvas_y": 20.0, "group_name": None,
         "attrs": {"vlan_id": "100"}},
    ]}
    wb = build_workbook(**ctx)
    reloaded = _load_bytes(wb)
    assert "路由器" in reloaded.sheetnames
    ws = reloaded["路由器"]
    assert [ws.cell(row=1, column=i).value for i in range(1, 7)] == NODE_FIXED_HEADERS
    assert ws.cell(row=1, column=7).value == "vlan_id"
    assert ws.cell(row=2, column=1).value == "R1"
    assert ws["A1"].comment is not None
    assert f"{NODE_TYPE_MARKER}=router" in ws["A1"].comment.text


def test_build_workbook_index_lists_data_sheets():
    ctx = _basic_context()
    ctx["node_types"] = [{"id": "nt_r", "code": "router", "name": "路由器",
                          "fields": []}]
    ctx["nodes_by_type_code"] = {"router": []}
    ctx["edge_types"] = [{"id": "et_l", "code": "link", "name": "连接",
                          "fields": []}]
    ctx["edges_by_type_code"] = {"link": []}
    wb = build_workbook(**ctx)
    reloaded = _load_bytes(wb)
    ws = reloaded[INDEX_SHEET_NAME]
    categories = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert "节点" in categories
    assert "边" in categories
    assert "节点组" in categories
    assert "节点组边策略" in categories
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "节点":
            sheet_name_cell = ws.cell(row=r, column=3)
            assert sheet_name_cell.hyperlink is not None
            assert "路由器" in (sheet_name_cell.hyperlink.location or "")


def test_build_workbook_edge_uses_source_target_names():
    ctx = _basic_context()
    ctx["node_types"] = [{"id": "nt_r", "code": "router", "name": "路由器", "fields": []}]
    ctx["nodes_by_type_code"] = {"router": [
        {"id": "n1", "name": "R1", "dn": None, "status": "online",
         "canvas_x": None, "canvas_y": None, "group_name": None, "attrs": {}},
        {"id": "n2", "name": "R2", "dn": None, "status": "online",
         "canvas_x": None, "canvas_y": None, "group_name": None, "attrs": {}},
    ]}
    ctx["edge_types"] = [{"id": "et_l", "code": "link", "name": "连接",
                          "fields": [{"field_key": "bandwidth"}]}]
    ctx["edges_by_type_code"] = {"link": [
        {"id": "e1", "source_id": "n1", "target_id": "n2", "status": "online",
         "attrs": {"bandwidth": "10G"}},
    ]}
    wb = build_workbook(**ctx)
    reloaded = _load_bytes(wb)
    ws = reloaded["连接"]
    assert ws.cell(row=2, column=1).value == "R1"
    assert ws.cell(row=2, column=2).value == "R2"
    assert ws.cell(row=2, column=4).value == "10G"


def test_build_workbook_alarm_sheets_when_schema_bound():
    ctx = _basic_context()
    ctx["topology"]["alarm_schema_code"] = "s1"
    ctx["alarm_schema_fields"] = [
        {"field_key": "severity", "mapping_target": None},
        {"field_key": "node_dn", "mapping_target": "dn"},
    ]
    ctx["node_alarms"] = []
    ctx["node_group_alarms"] = []
    wb = build_workbook(**ctx)
    reloaded = _load_bytes(wb)
    assert NODE_ALARM_SHEET_NAME in reloaded.sheetnames
    assert NODE_GROUP_ALARM_SHEET_NAME in reloaded.sheetnames
    ws = reloaded[NODE_ALARM_SHEET_NAME]
    assert ws.cell(row=1, column=4).value == "severity"
    assert ws.cell(row=1, column=5).value is None


# --- parse_workbook ---

from app.admin._topology_excel import parse_workbook, ParseResult


def _build_test_wb_bytes(builder):
    from openpyxl import Workbook as _Wb
    wb = _Wb()
    default = wb.active
    wb.remove(default)
    builder(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_workbook_reads_meta_sheet():
    def build(wb):
        ws = wb.create_sheet(META_SHEET_NAME)
        rows = [("字段", "值"), ("拓扑名称", "机房"), ("描述", "d"),
                ("版本", 2), ("所属网管/设备", "网管A"), ("告警模板", "s1")]
        for r, (k, v) in enumerate(rows, start=1):
            ws.cell(row=r, column=1, value=k)
            ws.cell(row=r, column=2, value=v)
        ws2 = wb.create_sheet("路由器")
        for c, h in enumerate(NODE_FIXED_HEADERS, start=1):
            ws2.cell(row=1, column=c, value=h)
        from openpyxl.comments import Comment
        ws2["A1"].comment = Comment(f"{NODE_TYPE_MARKER}=router", "system")
    data = _build_test_wb_bytes(build)

    result = parse_workbook(io.BytesIO(data))
    assert isinstance(result, ParseResult)
    assert result.meta["name"] == "机房"
    assert result.meta["version"] == 2
    assert result.meta["domain_name"] == "网管A"
    assert result.meta["alarm_schema_code"] == "s1"


def test_parse_workbook_meta_sheet_missing_fatal():
    def build(wb):
        ws = wb.create_sheet("路由器")
        for c, h in enumerate(NODE_FIXED_HEADERS, start=1):
            ws.cell(row=1, column=c, value=h)
    data = _build_test_wb_bytes(build)

    with pytest.raises(ExcelValidationError) as exc:
        parse_workbook(io.BytesIO(data))
    assert "拓扑元信息" in str(exc.value)


def test_parse_workbook_reads_node_sheet_by_A1_marker():
    def build(wb):
        wsm = wb.create_sheet(META_SHEET_NAME)
        wsm.cell(row=1, column=1, value="字段")
        wsm.cell(row=1, column=2, value="值")
        wsm.cell(row=2, column=1, value="拓扑名称")
        wsm.cell(row=2, column=2, value="T")

        ws = wb.create_sheet("路由器_A")  # 名字被"改过"
        headers = NODE_FIXED_HEADERS + ["vlan_id"]
        for c, h in enumerate(headers, start=1):
            ws.cell(row=1, column=c, value=h)
        from openpyxl.comments import Comment
        ws["A1"].comment = Comment(f"{NODE_TYPE_MARKER}=router", "system")
        ws.cell(row=2, column=1, value="R1")
        ws.cell(row=2, column=7, value="100")
    data = _build_test_wb_bytes(build)

    result = parse_workbook(io.BytesIO(data))
    assert "router" in result.nodes_by_type_code
    nodes = result.nodes_by_type_code["router"]
    assert nodes[0]["name"] == "R1"
    assert nodes[0]["attrs"]["vlan_id"] == "100"


def test_parse_workbook_edge_sheet_edges_captured():
    def build(wb):
        wsm = wb.create_sheet(META_SHEET_NAME)
        wsm.cell(row=1, column=1, value="字段")
        wsm.cell(row=1, column=2, value="值")
        wsm.cell(row=2, column=1, value="拓扑名称")
        wsm.cell(row=2, column=2, value="T")

        from openpyxl.comments import Comment
        ws_r = wb.create_sheet("路由器")
        for c, h in enumerate(NODE_FIXED_HEADERS, start=1):
            ws_r.cell(row=1, column=c, value=h)
        ws_r["A1"].comment = Comment(f"{NODE_TYPE_MARKER}=router", "system")
        ws_r.cell(row=2, column=1, value="R1")
        ws_r.cell(row=3, column=1, value="R2")

        ws_e = wb.create_sheet("连接")
        headers = EDGE_FIXED_HEADERS + ["bandwidth"]
        for c, h in enumerate(headers, start=1):
            ws_e.cell(row=1, column=c, value=h)
        ws_e["A1"].comment = Comment(f"{EDGE_TYPE_MARKER}=link", "system")
        ws_e.cell(row=2, column=1, value="R1")
        ws_e.cell(row=2, column=2, value="R2")
        ws_e.cell(row=2, column=4, value="10G")
    data = _build_test_wb_bytes(build)

    result = parse_workbook(io.BytesIO(data))
    assert "link" in result.edges_by_type_code
    e = result.edges_by_type_code["link"][0]
    assert e["source_name"] == "R1"
    assert e["target_name"] == "R2"
    assert e["attrs"]["bandwidth"] == "10G"


def test_parse_workbook_ignores_underscore_sheets():
    def build(wb):
        wsm = wb.create_sheet(META_SHEET_NAME)
        wsm.cell(row=1, column=1, value="字段")
        wsm.cell(row=1, column=2, value="值")
        wsm.cell(row=2, column=1, value="拓扑名称")
        wsm.cell(row=2, column=2, value="T")

        wb.create_sheet("_使用说明")
        wb.create_sheet("_总表")

        from openpyxl.comments import Comment
        ws = wb.create_sheet("路由器")
        for c, h in enumerate(NODE_FIXED_HEADERS, start=1):
            ws.cell(row=1, column=c, value=h)
        ws["A1"].comment = Comment(f"{NODE_TYPE_MARKER}=router", "system")
    data = _build_test_wb_bytes(build)

    result = parse_workbook(io.BytesIO(data))
    assert "router" in result.nodes_by_type_code


def test_parse_workbook_group_edge_strategy_row_read():
    def build(wb):
        wsm = wb.create_sheet(META_SHEET_NAME)
        wsm.cell(row=1, column=1, value="字段")
        wsm.cell(row=1, column=2, value="值")
        wsm.cell(row=2, column=1, value="拓扑名称")
        wsm.cell(row=2, column=2, value="T")

        from openpyxl.comments import Comment
        ws = wb.create_sheet(NODE_GROUP_EDGE_STRATEGY_SHEET_NAME)
        for c, h in enumerate(NODE_GROUP_EDGE_STRATEGY_HEADERS, start=1):
            ws.cell(row=1, column=c, value=h)
        ws["A1"].comment = Comment(f"{NODE_GROUP_EDGE_STRATEGY_MARKER}=1", "system")
        ws.cell(row=2, column=1, value="组A")
        ws.cell(row=2, column=2, value="组B")
        ws.cell(row=2, column=3, value="组")
        ws.cell(row=2, column=4, value="link")
        ws.cell(row=2, column=5, value="modulo")
        ws.cell(row=2, column=6, value=4)

        ws2 = wb.create_sheet("路由器")
        for c, h in enumerate(NODE_FIXED_HEADERS, start=1):
            ws2.cell(row=1, column=c, value=h)
        ws2["A1"].comment = Comment(f"{NODE_TYPE_MARKER}=router", "system")
    data = _build_test_wb_bytes(build)

    result = parse_workbook(io.BytesIO(data))
    assert len(result.node_group_edge_strategies) == 1
    s = result.node_group_edge_strategies[0]
    assert s["source_group_name"] == "组A"
    assert s["mode"] == "modulo"
    assert s["ratio_k"] == 4
