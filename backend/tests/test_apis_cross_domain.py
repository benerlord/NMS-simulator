"""CRUD + 导入的跨域同名接口行为测试。"""


def test_create_same_method_path_different_domains_ok(client):
    d1 = client.post("/admin/api/domains", json={"name": "网管A"}).json()["data"]["id"]
    d2 = client.post("/admin/api/domains", json={"name": "网管B"}).json()["data"]["id"]

    r1 = client.post("/admin/api/apis", json={
        "method": "PUT", "path": "/rest/token", "name": "A-token",
        "domainId": d1, "dataSource": "static", "config": {},
    })
    assert r1.status_code == 200, r1.text

    r2 = client.post("/admin/api/apis", json={
        "method": "PUT", "path": "/rest/token", "name": "B-token",
        "domainId": d2, "dataSource": "static", "config": {},
    })
    assert r2.status_code == 200, r2.text
    assert r1.json()["data"]["id"] != r2.json()["data"]["id"]


def test_create_same_method_path_same_domain_blocked(client):
    d = client.post("/admin/api/domains", json={"name": "网管A"}).json()["data"]["id"]
    r1 = client.post("/admin/api/apis", json={
        "method": "GET", "path": "/dup", "name": "first",
        "domainId": d, "dataSource": "static", "config": {},
    })
    assert r1.status_code == 200

    r2 = client.post("/admin/api/apis", json={
        "method": "GET", "path": "/dup", "name": "second",
        "domainId": d, "dataSource": "static", "config": {},
    })
    assert r2.status_code == 409
    assert r2.json()["detail"]["code"] == 40301


def test_create_null_domain_duplicates_allowed(client):
    """未归类接口不做重复预检——跟 SQLite NULL 语义一致。"""
    r1 = client.post("/admin/api/apis", json={
        "method": "GET", "path": "/orphan", "name": "o1",
        "dataSource": "static", "config": {},
    })
    assert r1.status_code == 200
    r2 = client.post("/admin/api/apis", json={
        "method": "GET", "path": "/orphan", "name": "o2",
        "dataSource": "static", "config": {},
    })
    assert r2.status_code == 200
    assert r1.json()["data"]["id"] != r2.json()["data"]["id"]


def test_update_moving_path_within_same_domain_conflict_blocked(client):
    d = client.post("/admin/api/domains", json={"name": "网管A"}).json()["data"]["id"]
    a1 = client.post("/admin/api/apis", json={
        "method": "GET", "path": "/a", "name": "a",
        "domainId": d, "dataSource": "static", "config": {},
    }).json()["data"]["id"]
    client.post("/admin/api/apis", json={
        "method": "GET", "path": "/b", "name": "b",
        "domainId": d, "dataSource": "static", "config": {},
    })

    r = client.put(f"/admin/api/apis/{a1}", json={"path": "/b"})
    assert r.status_code == 409
    assert r.json()["detail"]["code"] == 40301


def test_update_moving_path_across_domain_ok(client):
    """A 域下 /x → 更新 path 为 /y，B 域下已有 /y 不应阻挡。"""
    dA = client.post("/admin/api/domains", json={"name": "网管A"}).json()["data"]["id"]
    dB = client.post("/admin/api/domains", json={"name": "网管B"}).json()["data"]["id"]
    aA = client.post("/admin/api/apis", json={
        "method": "GET", "path": "/x", "name": "xA",
        "domainId": dA, "dataSource": "static", "config": {},
    }).json()["data"]["id"]
    client.post("/admin/api/apis", json={
        "method": "GET", "path": "/y", "name": "yB",
        "domainId": dB, "dataSource": "static", "config": {},
    })

    r = client.put(f"/admin/api/apis/{aA}", json={"path": "/y"})
    assert r.status_code == 200, r.text


def test_duplicate_api_generates_unique_path_within_domain(client):
    """复制接口时生成 _copy 后缀，冲突判断按域，不看其它域。"""
    dA = client.post("/admin/api/domains", json={"name": "网管A"}).json()["data"]["id"]
    dB = client.post("/admin/api/domains", json={"name": "网管B"}).json()["data"]["id"]
    aA = client.post("/admin/api/apis", json={
        "method": "GET", "path": "/foo", "name": "fooA",
        "domainId": dA, "dataSource": "static", "config": {},
    }).json()["data"]["id"]
    # B 域下预置一个 /foo_copy，确保复制时 A 域的判定不会误参考它
    client.post("/admin/api/apis", json={
        "method": "GET", "path": "/foo_copy", "name": "fooB_copy",
        "domainId": dB, "dataSource": "static", "config": {},
    })

    r = client.post(f"/admin/api/apis/{aA}/duplicate")
    assert r.status_code == 200
    new_id = r.json()["data"]["id"]
    detail = client.get(f"/admin/api/apis/{new_id}").json()["data"]
    # 因为 A 域下不存在 /foo_copy，直接可用（不应因为 B 域的 /foo_copy 而变成 /foo_copy2）
    assert detail["path"] == "/foo_copy"
    assert detail["domainId"] == dA


import io
from openpyxl import Workbook
from app.admin._api_excel import MAIN_HEADERS, UNCATEGORIZED_SHEET_NAME


def _build_xlsx(builder) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    builder(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload_xlsx(client, xlsx_bytes: bytes):
    return client.post(
        "/admin/api/apis/import",
        files={"file": ("t.xlsx", xlsx_bytes,
                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )


def test_import_cross_sheet_move_leaves_source_untouched(client):
    """Sheet A 有 /token；Excel 里 Sheet B 也有 /token → 导入后 A 保留、B 新建。"""
    dA = client.post("/admin/api/domains", json={"name": "网管A"}).json()["data"]["id"]
    dB = client.post("/admin/api/domains", json={"name": "网管B"}).json()["data"]["id"]
    a_id = client.post("/admin/api/apis", json={
        "method": "PUT", "path": "/rest/token", "name": "A-token",
        "domainId": dA, "dataSource": "static", "config": {},
    }).json()["data"]["id"]

    def build(wb):
        ws = wb.create_sheet("网管B")
        for i, h in enumerate(MAIN_HEADERS, start=1):
            ws.cell(row=1, column=i, value=h)
        ws.cell(row=2, column=1, value="PUT")
        ws.cell(row=2, column=2, value="/rest/token")
        ws.cell(row=2, column=3, value="B-token-from-excel")
        ws.cell(row=2, column=7, value="static")
    data = _build_xlsx(build)

    r = _upload_xlsx(client, data)
    assert r.status_code == 200
    result = r.json()["data"]
    assert result["created"] == 1
    assert result["updated"] == 0

    # 源 A 域接口不动
    a_detail = client.get(f"/admin/api/apis/{a_id}").json()["data"]
    assert a_detail["name"] == "A-token"
    assert a_detail["domainId"] == dA

    # 目标 B 域新建了一份
    b_apis = client.get(f"/admin/api/apis?domainId={dB}").json()["data"]["items"]
    assert any(a["path"] == "/rest/token" and a["name"] == "B-token-from-excel" for a in b_apis)


def test_import_same_domain_same_path_updates_that_row(client):
    """Excel Sheet=网管A 且行 (method, path) 命中 A 域下已有接口 → UPDATE 该行。"""
    dA = client.post("/admin/api/domains", json={"name": "网管A"}).json()["data"]["id"]
    a_id = client.post("/admin/api/apis", json={
        "method": "PUT", "path": "/rest/token", "name": "旧名",
        "domainId": dA, "dataSource": "static", "config": {},
    }).json()["data"]["id"]

    def build(wb):
        ws = wb.create_sheet("网管A")
        for i, h in enumerate(MAIN_HEADERS, start=1):
            ws.cell(row=1, column=i, value=h)
        ws.cell(row=2, column=1, value="PUT")
        ws.cell(row=2, column=2, value="/rest/token")
        ws.cell(row=2, column=3, value="新名")
        ws.cell(row=2, column=7, value="static")
    data = _build_xlsx(build)

    r = _upload_xlsx(client, data)
    assert r.json()["data"]["updated"] == 1
    detail = client.get(f"/admin/api/apis/{a_id}").json()["data"]
    assert detail["name"] == "新名"


def test_import_cross_domain_does_not_affect_other_domain(client):
    """A/B 两域各有 /token；Excel Sheet=A 只改 A 的 name → B 不受影响。"""
    dA = client.post("/admin/api/domains", json={"name": "网管A"}).json()["data"]["id"]
    dB = client.post("/admin/api/domains", json={"name": "网管B"}).json()["data"]["id"]
    aA = client.post("/admin/api/apis", json={
        "method": "PUT", "path": "/rest/token", "name": "A-orig",
        "domainId": dA, "dataSource": "static", "config": {},
    }).json()["data"]["id"]
    aB = client.post("/admin/api/apis", json={
        "method": "PUT", "path": "/rest/token", "name": "B-orig",
        "domainId": dB, "dataSource": "static", "config": {},
    }).json()["data"]["id"]

    def build(wb):
        ws = wb.create_sheet("网管A")
        for i, h in enumerate(MAIN_HEADERS, start=1):
            ws.cell(row=1, column=i, value=h)
        ws.cell(row=2, column=1, value="PUT")
        ws.cell(row=2, column=2, value="/rest/token")
        ws.cell(row=2, column=3, value="A-new")
        ws.cell(row=2, column=7, value="static")
    data = _build_xlsx(build)

    r = _upload_xlsx(client, data)
    assert r.json()["data"]["updated"] == 1

    assert client.get(f"/admin/api/apis/{aA}").json()["data"]["name"] == "A-new"
    assert client.get(f"/admin/api/apis/{aB}").json()["data"]["name"] == "B-orig"


def test_admin_8080_no_longer_mounts_mock_routes(client):
    """admin 端口停用 mock 服务后，任何 mock 路径打 admin 都应 404。"""
    d = client.post("/admin/api/domains", json={"name": "网管A"}).json()["data"]["id"]
    client.post("/admin/api/apis", json={
        "method": "GET", "path": "/api/some-mock", "name": "m",
        "domainId": d, "dataSource": "static",
        "config": {"staticBody": '{"ok":true}'},
    })
    # 直接打 mock 路径应 404（admin 上不再挂载）
    r = client.get("/api/some-mock")
    assert r.status_code == 404


def test_admin_still_serves_admin_api(client):
    """删掉 mock_registry 后 admin API 本身仍然可用。"""
    r = client.get("/admin/api/health")
    assert r.status_code == 200
