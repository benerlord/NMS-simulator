"""画布 Excel 导出/导入端到端测试。"""
import io

from openpyxl import load_workbook

from app.admin._topology_excel import (
    META_SHEET_NAME,
    NODE_GROUP_SHEET_NAME,
    NODE_GROUP_EDGE_STRATEGY_SHEET_NAME,
    NODE_ALARM_SHEET_NAME,
    NODE_FIXED_HEADERS,
    EDGE_FIXED_HEADERS,
    NODE_TYPE_MARKER,
)


def _make_topology_with_nodes(client, with_schema=True):
    r = client.post("/admin/api/topologies", json={"name": "T"})
    tid = r.json()["data"]["id"]
    r = client.post("/admin/api/node-types", json={
        "code": "xlsx_router", "name": "路由器", "category": "switch",
        "fields": [{"fieldKey": "vlan_id", "fieldLabel": "VLAN",
                    "fieldType": "text", "maxLength": 20}],
    })
    ntid = r.json()["data"]["id"]
    r = client.post("/admin/api/edge-types", json={
        "code": "xlsx_link", "name": "连接",
        "fields": [{"fieldKey": "bandwidth", "fieldLabel": "带宽",
                    "fieldType": "text", "maxLength": 20}],
    })
    etid = r.json()["data"]["id"]
    r = client.post(f"/admin/api/topologies/{tid}/nodes", json={
        "nodeTypeId": ntid, "name": "R1",
    })
    nid1 = r.json()["data"]["id"]
    client.put(f"/admin/api/nodes/{nid1}/attrs",
               json=[{"fieldKey": "vlan_id", "value": "100"}])
    r = client.post(f"/admin/api/topologies/{tid}/nodes", json={
        "nodeTypeId": ntid, "name": "R2",
    })
    nid2 = r.json()["data"]["id"]
    client.put(f"/admin/api/nodes/{nid2}/attrs",
               json=[{"fieldKey": "vlan_id", "value": "200"}])
    return tid, ntid, etid, nid1, nid2


def test_export_excel_returns_xlsx(client):
    tid, _, _, _, _ = _make_topology_with_nodes(client)
    r = client.get(f"/admin/api/topologies/{tid}/export-excel")
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    wb = load_workbook(io.BytesIO(r.content))
    assert META_SHEET_NAME in wb.sheetnames
    assert "路由器" in wb.sheetnames
    ws = wb["路由器"]
    assert ws.cell(row=2, column=1).value == "R1"
    assert ws.cell(row=2, column=7).value == "100"


def test_export_excel_edges_use_node_names(client):
    tid, ntid, etid, nid1, nid2 = _make_topology_with_nodes(client)
    r = client.post(f"/admin/api/topologies/{tid}/edges", json={
        "edgeTypeId": etid, "sourceId": nid1, "targetId": nid2,
    })
    eid = r.json()["data"]["id"]
    client.put(f"/admin/api/edges/{eid}/attrs",
               json=[{"fieldKey": "bandwidth", "value": "10G"}])
    r = client.get(f"/admin/api/topologies/{tid}/export-excel")
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["连接"]
    assert ws.cell(row=2, column=1).value == "R1"
    assert ws.cell(row=2, column=2).value == "R2"
    assert ws.cell(row=2, column=4).value == "10G"


def test_import_excel_creates_new_topology(client):
    src_tid, _, _, _, _ = _make_topology_with_nodes(client)
    r = client.get(f"/admin/api/topologies/{src_tid}/export-excel")
    xlsx = r.content

    r = client.post(
        "/admin/api/topologies/import-excel",
        files={"file": ("t.xlsx", xlsx,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text
    result = r.json()["data"]
    assert result["topologyId"] != src_tid
    assert "(导入" in result["topologyName"]
    assert result["counts"]["nodes"] == 2


def test_import_excel_rejects_non_xlsx(client):
    r = client.post(
        "/admin/api/topologies/import-excel",
        files={"file": ("t.json", b'{"a":1}', "application/json")},
    )
    assert r.status_code == 400
    assert r.json()["detail"]["code"] == 40410


def test_import_excel_unknown_node_type_code_rejected(client):
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    wb = Workbook()
    wb.remove(wb.active)
    ws_meta = wb.create_sheet(META_SHEET_NAME)
    ws_meta.cell(row=1, column=1, value="字段")
    ws_meta.cell(row=1, column=2, value="值")
    ws_meta.cell(row=2, column=1, value="拓扑名称")
    ws_meta.cell(row=2, column=2, value="T2")
    ws_g = wb.create_sheet("Ghost")
    for c, h in enumerate(NODE_FIXED_HEADERS, start=1):
        ws_g.cell(row=1, column=c, value=h)
    ws_g["A1"].comment = Comment(f"{NODE_TYPE_MARKER}=ghost_type_xyz", "system")
    ws_g.cell(row=2, column=1, value="foo")
    buf = io.BytesIO()
    wb.save(buf)

    r = client.post(
        "/admin/api/topologies/import-excel",
        files={"file": ("t.xlsx", buf.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 400
    assert r.json()["detail"]["code"] == 40431


def test_import_excel_duplicate_node_name_rejected(client):
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    client.post("/admin/api/node-types", json={
        "code": "sw2", "name": "SW2", "category": "switch",
    })
    wb = Workbook()
    wb.remove(wb.active)
    ws_meta = wb.create_sheet(META_SHEET_NAME)
    ws_meta.cell(row=1, column=1, value="字段")
    ws_meta.cell(row=1, column=2, value="值")
    ws_meta.cell(row=2, column=1, value="拓扑名称")
    ws_meta.cell(row=2, column=2, value="T3")
    ws = wb.create_sheet("SW2")
    for c, h in enumerate(NODE_FIXED_HEADERS, start=1):
        ws.cell(row=1, column=c, value=h)
    ws["A1"].comment = Comment(f"{NODE_TYPE_MARKER}=sw2", "system")
    ws.cell(row=2, column=1, value="dup")
    ws.cell(row=3, column=1, value="dup")
    buf = io.BytesIO()
    wb.save(buf)

    r = client.post(
        "/admin/api/topologies/import-excel",
        files={"file": ("t.xlsx", buf.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 400
    assert r.json()["detail"]["code"] == 40432


def test_import_excel_missing_meta_sheet_rejected(client):
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("路由器")
    buf = io.BytesIO()
    wb.save(buf)

    r = client.post(
        "/admin/api/topologies/import-excel",
        files={"file": ("t.xlsx", buf.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 400
    assert r.json()["detail"]["code"] == 40411
