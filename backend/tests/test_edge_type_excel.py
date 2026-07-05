"""边类型 Excel 导入导出 e2e 测试。"""
import io

import openpyxl


def _create_edge_type(
    client, code: str, name: str, semantic: str = "connect",
    directed: bool = True, exclusive_target: bool = False,
    allow_source: str = None, allow_target: str = None,
    line_style: str = None, color: str = None, description: str = None,
    fields: list = None,
) -> str:
    payload = {
        "code": code, "name": name, "semantic": semantic,
        "directed": directed, "exclusiveTarget": exclusive_target,
        "allowSourceTypeCodes": allow_source,
        "allowTargetTypeCodes": allow_target,
        "lineStyle": line_style, "color": color,
        "description": description,
        "fields": fields or [],
    }
    r = client.post("/admin/api/edge-types", json=payload)
    assert r.status_code == 200, r.text
    return r.json()["data"]["id"]


# ============== 导出测试 ==============

def test_export_all_returns_xlsx_with_summary_sheet(client):
    """导出全部 → xlsx 含'边类型汇总' Sheet + 每 code 独立字段 Sheet。"""
    _create_edge_type(client, "et_a", "边A")
    _create_edge_type(client, "et_b", "边B")

    r = client.post("/admin/api/edge-types/export", json={})
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert "边类型汇总" in wb.sheetnames
    assert "et_a" in wb.sheetnames
    assert "et_b" in wb.sheetnames


def test_export_summary_contains_semantic_and_directed_columns(client):
    """汇总 Sheet 表头正确（13 列）。"""
    _create_edge_type(client, "et_hdr", "边表头")

    r = client.post("/admin/api/edge-types/export", json={})
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb["边类型汇总"]
    headers = [c.value for c in ws[1] if c.value]

    for h in ["Code", "名称", "语义", "有向", "唯一目标",
              "允许源类型", "允许目标类型", "线条样式", "颜色",
              "描述", "字段数", "创建时间", "更新时间"]:
        assert h in headers, f"缺少表头 {h}"


def test_export_directed_and_exclusive_serialized_as_yes_no(client):
    """有向 / 唯一目标 布尔 → '是' / '否'。"""
    _create_edge_type(client, "et_yes", "有向的", directed=True, exclusive_target=True)
    _create_edge_type(client, "et_no", "无向的", directed=False, exclusive_target=False)

    r = client.post("/admin/api/edge-types/export", json={})
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb["边类型汇总"]
    headers = {c.value: idx for idx, c in enumerate(ws[1]) if c.value}
    dir_col = headers["有向"] + 1
    exc_col = headers["唯一目标"] + 1
    code_col = headers["Code"] + 1

    values = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        code = row[code_col - 1].value
        if code in ("et_yes", "et_no"):
            values[code] = (row[dir_col - 1].value, row[exc_col - 1].value)

    assert values["et_yes"] == ("是", "是")
    assert values["et_no"] == ("否", "否")


def test_export_allow_codes_kept_as_comma_separated_string(client):
    """允许源/目标类型 逗号分隔字符串原样输出。"""
    _create_edge_type(
        client, "et_al", "有白名单",
        allow_source="switch,router", allow_target="server",
    )

    r = client.post("/admin/api/edge-types/export", json={})
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb["边类型汇总"]
    headers = {c.value: idx for idx, c in enumerate(ws[1]) if c.value}
    src_col = headers["允许源类型"] + 1
    tgt_col = headers["允许目标类型"] + 1
    code_col = headers["Code"] + 1

    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[code_col - 1].value == "et_al":
            assert row[src_col - 1].value == "switch,router"
            assert row[tgt_col - 1].value == "server"
            return
    raise AssertionError("et_al 行未找到")


def test_export_ids_only_returns_selected(client):
    """按 ids 过滤导出。"""
    tid_a = _create_edge_type(client, "et_only_a", "只A")
    _create_edge_type(client, "et_only_b", "只B")

    r = client.post("/admin/api/edge-types/export", json={"ids": [tid_a]})
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert "et_only_a" in wb.sheetnames
    assert "et_only_b" not in wb.sheetnames


# ============== 导入 preview 测试 ==============

def _build_edge_import_xlsx(rows: list[dict]) -> io.BytesIO:
    """构造一份最小导入 xlsx。rows: [{code, name, semantic, directed, exclusive_target,
    allow_source, allow_target, line_style, color, description, fields: [...]}]"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "边类型汇总"
    ws.append(["Code", "名称", "语义", "有向", "唯一目标",
               "允许源类型", "允许目标类型", "线条样式", "颜色", "描述"])
    for r in rows:
        ws.append([
            r.get("code"), r.get("name"),
            r.get("semantic", "connect"),
            "是" if r.get("directed", True) else "否",
            "是" if r.get("exclusive_target", False) else "否",
            r.get("allow_source"), r.get("allow_target"),
            r.get("line_style"), r.get("color"),
            r.get("description"),
        ])

    for r in rows:
        code = r.get("code")
        fields = r.get("fields")
        if fields is None:
            continue
        fs = wb.create_sheet(title=code)
        fs.append(["字段标识", "显示名称", "字段类型", "最大长度",
                    "默认值", "选项", "必填", "排序"])
        for f in fields:
            fs.append([
                f.get("fieldKey"), f.get("fieldLabel"), f.get("fieldType"),
                f.get("maxLength"), f.get("defaultValue"), f.get("options"),
                "是" if f.get("required") else "否",
                f.get("sortOrder", 0),
            ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_import_preview_categorizes_create_and_update(client):
    """已存在 code → toUpdate；不存在 → toCreate。"""
    _create_edge_type(client, "et_exists", "已存在")
    buf = _build_edge_import_xlsx([
        {"code": "et_exists", "name": "已存在改名"},
        {"code": "et_new_1", "name": "新1"},
        {"code": "et_new_2", "name": "新2"},
    ])

    r = client.post(
        "/admin/api/edge-types/import/preview",
        files={"file": ("test.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text
    data = r.json()["data"]

    to_create_codes = [item["code"] for item in data["toCreate"]]
    to_update_codes = [item["code"] for item in data["toUpdate"]]
    assert sorted(to_create_codes) == ["et_new_1", "et_new_2"]
    assert to_update_codes == ["et_exists"]


def test_import_preview_records_old_name_on_update(client):
    """覆盖项 oldName 与新 name 不同时正确记录。"""
    _create_edge_type(client, "et_rename", "旧名字")
    buf = _build_edge_import_xlsx([
        {"code": "et_rename", "name": "新名字"},
    ])

    r = client.post(
        "/admin/api/edge-types/import/preview",
        files={"file": ("test.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200
    item = r.json()["data"]["toUpdate"][0]
    assert item["code"] == "et_rename"
    assert item["name"] == "新名字"
    assert item["oldName"] == "旧名字"


def test_import_preview_missing_summary_sheet_returns_400(client):
    """缺'边类型汇总' Sheet → 400。"""
    wb = openpyxl.Workbook()
    wb.active.title = "OtherSheet"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    r = client.post(
        "/admin/api/edge-types/import/preview",
        files={"file": ("bad.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 400
    assert "边类型汇总" in r.json()["detail"]["message"]


# ============== 正式导入测试 ==============

def test_import_creates_new_edge_type_with_fields(client):
    """新边类型 + 字段一起导入。"""
    buf = _build_edge_import_xlsx([
        {"code": "et_imp_new", "name": "导入新", "semantic": "connect",
         "directed": True, "line_style": "solid", "color": "#1890ff",
         "fields": [
             {"fieldKey": "bandwidth", "fieldLabel": "带宽",
              "fieldType": "text", "maxLength": 30, "sortOrder": 0},
             {"fieldKey": "priority", "fieldLabel": "优先级",
              "fieldType": "number", "sortOrder": 1},
         ]},
    ])

    r = client.post(
        "/admin/api/edge-types/import",
        files={"file": ("test.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text
    data = r.json()["data"]
    assert data["created"] == 1
    assert data["updated"] == 0
    assert data["totalFields"] == 2

    lst = client.get("/admin/api/edge-types").json()["data"]["items"]
    match = next(it for it in lst if it["code"] == "et_imp_new")
    detail = client.get(f"/admin/api/edge-types/{match['id']}").json()["data"]
    assert [f["fieldKey"] for f in detail["fields"]] == ["bandwidth", "priority"]
    assert detail["lineStyle"] == "solid"
    assert detail["color"] == "#1890ff"


def test_import_overwrites_existing_edge_type(client):
    """已存在 code → 主表覆盖，字段全部替换。"""
    _create_edge_type(
        client, "et_imp_ov", "旧名字",
        fields=[
            {"fieldKey": "old_field", "fieldLabel": "旧字段",
             "fieldType": "text", "maxLength": 10, "sortOrder": 0}
        ],
    )
    buf = _build_edge_import_xlsx([
        {"code": "et_imp_ov", "name": "新名字",
         "fields": [
             {"fieldKey": "new_field", "fieldLabel": "新字段",
              "fieldType": "text", "maxLength": 30, "sortOrder": 0}
         ]},
    ])

    r = client.post(
        "/admin/api/edge-types/import",
        files={"file": ("test.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200
    data = r.json()["data"]
    assert data["updated"] == 1
    assert data["created"] == 0

    lst = client.get("/admin/api/edge-types").json()["data"]["items"]
    match = next(it for it in lst if it["code"] == "et_imp_ov")
    assert match["name"] == "新名字"
    detail = client.get(f"/admin/api/edge-types/{match['id']}").json()["data"]
    field_keys = [f["fieldKey"] for f in detail["fields"]]
    assert field_keys == ["new_field"]
    assert "old_field" not in field_keys


def test_import_directed_column_yes_maps_to_true(client):
    """'有向'='是' → True；'否' 或其他 → False。"""
    buf = _build_edge_import_xlsx([
        {"code": "et_dir_yes", "name": "有向是", "directed": True},
        {"code": "et_dir_no", "name": "有向否", "directed": False},
    ])

    r = client.post(
        "/admin/api/edge-types/import",
        files={"file": ("test.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200
    lst = client.get("/admin/api/edge-types").json()["data"]["items"]
    yes = next(it for it in lst if it["code"] == "et_dir_yes")
    no = next(it for it in lst if it["code"] == "et_dir_no")
    assert yes["directed"] is True
    assert no["directed"] is False


def test_import_invalid_semantic_records_error_and_skips_row(client):
    """'语义' 非 connect/contain → errors 记录，跳过整行。"""
    buf = _build_edge_import_xlsx([
        {"code": "et_bad_sem", "name": "非法语义", "semantic": "invalid_sem"},
        {"code": "et_ok_sem", "name": "合法", "semantic": "connect"},
    ])

    r = client.post(
        "/admin/api/edge-types/import",
        files={"file": ("test.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200
    data = r.json()["data"]
    assert data["created"] == 1  # 只有 et_ok_sem 被创建
    assert any("invalid_sem" in e for e in data["errors"])

    lst = client.get("/admin/api/edge-types").json()["data"]["items"]
    codes = [it["code"] for it in lst]
    assert "et_ok_sem" in codes
    assert "et_bad_sem" not in codes


def test_import_text_field_missing_maxlen_defaults_to_255(client):
    """text 字段最大长度为空 → 落库 255。"""
    buf = _build_edge_import_xlsx([
        {"code": "et_imp_mx", "name": "空长度",
         "fields": [
             {"fieldKey": "note", "fieldLabel": "备注",
              "fieldType": "text", "maxLength": None, "sortOrder": 0}
         ]},
    ])

    r = client.post(
        "/admin/api/edge-types/import",
        files={"file": ("test.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200
    lst = client.get("/admin/api/edge-types").json()["data"]["items"]
    match = next(it for it in lst if it["code"] == "et_imp_mx")
    detail = client.get(f"/admin/api/edge-types/{match['id']}").json()["data"]
    note = next(f for f in detail["fields"] if f["fieldKey"] == "note")
    assert note["maxLength"] == 255


def test_import_allow_codes_unknown_node_code_records_warning(client):
    """allow_source/target 引用不存在的 node_type code → warning，字符串仍保存。"""
    # 种一个已知的 node_type
    r = client.post("/admin/api/node-types", json={
        "code": "known_node", "name": "已知节点", "category": "physical",
    })
    assert r.status_code == 200, r.text

    buf = _build_edge_import_xlsx([
        {"code": "et_allow", "name": "白名单",
         "allow_source": "known_node,ghost_node",
         "allow_target": "another_ghost"},
    ])

    r = client.post(
        "/admin/api/edge-types/import",
        files={"file": ("test.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200
    data = r.json()["data"]
    assert data["created"] == 1
    assert any("ghost_node" in e for e in data["errors"])
    assert any("another_ghost" in e for e in data["errors"])

    lst = client.get("/admin/api/edge-types").json()["data"]["items"]
    match = next(it for it in lst if it["code"] == "et_allow")
    # 字符串仍保存
    assert match["allowSourceTypeCodes"] == "known_node,ghost_node"
    assert match["allowTargetTypeCodes"] == "another_ghost"


def test_import_partial_failure_isolated_per_row(client):
    """某边类型字段解析失败不影响其他边类型。"""
    buf = _build_edge_import_xlsx([
        {"code": "et_imp_ok", "name": "正常",
         "fields": [
             {"fieldKey": "level", "fieldLabel": "级别",
              "fieldType": "text", "maxLength": 20, "sortOrder": 0},
         ]},
        {"code": "et_imp_bad", "name": "含错误字段",
         "fields": [
             {"fieldKey": "bad", "fieldLabel": "坏",
              "fieldType": "not_a_type", "sortOrder": 0},
         ]},
        {"code": "et_imp_ok2", "name": "又正常",
         "fields": [
             {"fieldKey": "count", "fieldLabel": "计数",
              "fieldType": "number", "sortOrder": 0},
         ]},
    ])

    r = client.post(
        "/admin/api/edge-types/import",
        files={"file": ("test.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200
    data = r.json()["data"]
    assert data["created"] == 3  # 3 个边类型都创建
    assert data["totalFields"] == 2  # 只有 2 个字段成功
    assert len(data["errors"]) >= 1
