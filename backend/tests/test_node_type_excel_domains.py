"""节点类型 Excel 导入导出的所属网管列。"""
import io

import openpyxl


def _seed_domain(client, name: str) -> str:
    r = client.post("/admin/api/domains", json={"name": name})
    assert r.status_code == 200, r.text
    return r.json()["data"]["id"]


def test_export_has_domain_column_no_legacy_columns(client):
    """导出 xlsx 类型汇总表头含"所属网管/设备"，不含死字段列。"""
    dom_a = _seed_domain(client, "网管A")
    r = client.post("/admin/api/node-types", json={
        "code": "sw_e1", "name": "交换机", "category": "physical",
        "domainIds": [dom_a],
    })
    type_id = r.json()["data"]["id"]

    r = client.post("/admin/api/node-types/export", json={"ids": [type_id]})
    assert r.status_code == 200, r.text
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb["类型汇总"]
    headers = [c.value for c in ws[1] if c.value]

    assert "所属网管/设备" in headers
    for legacy in ("图标", "颜色", "形状", "渲染模式", "DN模板"):
        assert legacy not in headers, f"导出不应保留死字段列 {legacy}"


def test_export_domain_column_uses_pipe_separated_names(client):
    """导出的网管列值格式 '网管A|网管B'。"""
    dom_a = _seed_domain(client, "网管A")
    dom_b = _seed_domain(client, "网管B")
    r = client.post("/admin/api/node-types", json={
        "code": "sw_e2", "name": "交换机", "category": "physical",
        "domainIds": [dom_a, dom_b],
    })
    type_id = r.json()["data"]["id"]

    r = client.post("/admin/api/node-types/export", json={"ids": [type_id]})
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb["类型汇总"]
    headers = {c.value: idx for idx, c in enumerate(ws[1]) if c.value}
    dom_idx = headers["所属网管/设备"]
    val = ws.cell(row=2, column=dom_idx + 1).value

    assert set(val.split("|")) == {"网管A", "网管B"}


def _build_workbook_with_domains(code: str, domain_cell: str, field_maxlen=None):
    """构造一份包含"所属网管/设备"列的最小导入 xlsx。"""
    import openpyxl as _op

    wb = _op.Workbook()
    ws = wb.active
    ws.title = "类型汇总"
    ws.append(["编码", "名称", "分类", "所属网管/设备", "描述"])
    ws.append([code, "测试", "physical", domain_cell, None])

    fs = wb.create_sheet(title=code)
    fs.append(["字段标识", "显示名称", "字段类型", "最大长度",
                "默认值", "选项", "必填", "排序"])
    fs.append(["ip", "IP", "text", field_maxlen, None, None, "否", 0])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_import_reads_domain_column_and_links(client):
    """导入 xlsx 含"所属网管/设备"列 → 类型创建时自动关联对应网管。"""
    dom_a = _seed_domain(client, "网管A")
    buf = _build_workbook_with_domains("sw_imp_link", "网管A", field_maxlen=50)

    r = client.post("/admin/api/node-types/import",
                    files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200, r.text
    assert r.json()["data"]["created"] == 1

    items = client.get("/admin/api/node-types").json()["data"]["items"]
    match = next(it for it in items if it["code"] == "sw_imp_link")
    assert match["domainIds"] == [dom_a]


def test_import_unknown_domain_records_error_and_skips_link(client):
    """导入 xlsx 里网管名不存在 → 类型创建成功，errors 记录，关联跳过。"""
    _seed_domain(client, "网管A")
    buf = _build_workbook_with_domains("sw_imp_bad", "网管A|幽灵网管", field_maxlen=50)

    r = client.post("/admin/api/node-types/import",
                    files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200, r.text
    result = r.json()["data"]
    assert result["created"] == 1
    assert any("幽灵网管" in e for e in result["errors"])

    items = client.get("/admin/api/node-types").json()["data"]["items"]
    match = next(it for it in items if it["code"] == "sw_imp_bad")
    # 只关联到"网管A"，跳过不存在的
    assert len(match["domainIds"]) == 1


def test_import_text_field_missing_maxlen_defaults_to_255(client):
    """导入 xlsx text 字段"最大长度"列为空 → 落库 max_length=255。"""
    buf = _build_workbook_with_domains("sw_imp_mx", "", field_maxlen=None)

    r = client.post("/admin/api/node-types/import",
                    files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200, r.text
    assert r.json()["data"]["created"] == 1

    items = client.get("/admin/api/node-types").json()["data"]["items"]
    match = next(it for it in items if it["code"] == "sw_imp_mx")
    ip_field = next(f for f in match["fields"] if f["fieldKey"] == "ip")
    assert ip_field["maxLength"] == 255


def test_import_legacy_workbook_without_domain_column_still_works(client):
    """老 xlsx（含"图标/颜色"列，无"所属网管/设备"列）仍能正常导入。"""
    import openpyxl as _op

    wb = _op.Workbook()
    ws = wb.active
    ws.title = "类型汇总"
    ws.append(["编码", "名称", "分类", "图标", "颜色", "形状", "渲染模式", "DN模板", "描述"])
    ws.append(["sw_legacy", "老交换机", "physical",
                "🔀", "#123456", "rect", "flat", "sw={ip}", "老格式"])
    fs = wb.create_sheet(title="sw_legacy")
    fs.append(["字段标识", "显示名称", "字段类型", "最大长度", "默认值", "选项", "必填", "排序"])
    fs.append(["ip", "IP", "text", 50, None, None, "否", 0])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    r = client.post("/admin/api/node-types/import",
                    files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200, r.text
    assert r.json()["data"]["created"] == 1

    items = client.get("/admin/api/node-types").json()["data"]["items"]
    match = next(it for it in items if it["code"] == "sw_legacy")
    # 死字段列被忽略，导入仍成功
    assert match["name"] == "老交换机"
    assert match["domainIds"] == []
