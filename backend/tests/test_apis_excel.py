import io
import json

import pytest
from openpyxl import load_workbook

from app.admin._api_excel import (
    ExcelValidationError,
    build_workbook,
    format_cell_list,
    parse_cell_list,
    parse_workbook,
    sanitize_sheet_name,
    HEADER_COLUMNS,
    INSTRUCTION_SHEET_NAME,
    MAIN_HEADERS,
    ParseResult,
    UNCATEGORIZED_SHEET_NAME,
)


def test_format_cell_list_empty_returns_empty_string():
    assert format_cell_list([], HEADER_COLUMNS) == ""


def test_format_cell_list_single_record():
    records = [{"name": "Authorization", "required": True, "expectValue": "Bearer x", "example": "", "description": ""}]
    assert format_cell_list(records, HEADER_COLUMNS) == "Authorization|是|Bearer x||"


def test_format_cell_list_multiple_records_uses_newline():
    records = [
        {"name": "A", "required": True, "expectValue": "", "example": "", "description": ""},
        {"name": "B", "required": False, "expectValue": "", "example": "", "description": ""},
    ]
    assert format_cell_list(records, HEADER_COLUMNS) == "A|是|||\nB|否|||"


def test_format_cell_list_missing_key_treated_as_empty():
    records = [{"name": "A"}]
    assert format_cell_list(records, HEADER_COLUMNS) == "A||||"


def test_parse_cell_list_empty_returns_empty_list():
    assert parse_cell_list("", HEADER_COLUMNS, row_hint="第 1 行") == []


def test_parse_cell_list_single_record():
    result = parse_cell_list("Authorization|是|Bearer x||", HEADER_COLUMNS, row_hint="第 1 行")
    assert result == [{"name": "Authorization", "required": True, "expectValue": "Bearer x", "example": "", "description": ""}]


def test_parse_cell_list_multiple_records():
    text = "A|是|||\nB|否|||"
    result = parse_cell_list(text, HEADER_COLUMNS, row_hint="第 1 行")
    assert len(result) == 2
    assert result[0]["name"] == "A" and result[0]["required"] is True
    assert result[1]["name"] == "B" and result[1]["required"] is False


def test_parse_cell_list_skips_blank_lines():
    text = "A|是|||\n\nB|否|||"
    result = parse_cell_list(text, HEADER_COLUMNS, row_hint="第 1 行")
    assert len(result) == 2


def test_parse_cell_list_fewer_fields_treated_as_empty():
    result = parse_cell_list("A|是", HEADER_COLUMNS, row_hint="第 1 行")
    assert result == [{"name": "A", "required": True, "expectValue": "", "example": "", "description": ""}]


def test_parse_cell_list_more_fields_raises():
    with pytest.raises(ExcelValidationError) as exc:
        parse_cell_list("A|是|X|Y|Z|EXTRA", HEADER_COLUMNS, row_hint="第 3 行 请求头")
    assert "第 3 行 请求头" in str(exc.value)
    assert "字段数" in str(exc.value)


def test_parse_cell_list_accepts_true_false_synonyms():
    result = parse_cell_list("A|true|||", HEADER_COLUMNS, row_hint="")
    assert result[0]["required"] is True
    result = parse_cell_list("A|false|||", HEADER_COLUMNS, row_hint="")
    assert result[0]["required"] is False


def test_parse_cell_list_round_trip():
    records = [
        {"name": "X", "required": True, "expectValue": "v1", "example": "e1", "description": "d1"},
        {"name": "Y", "required": False, "expectValue": "", "example": "", "description": ""},
    ]
    text = format_cell_list(records, HEADER_COLUMNS)
    parsed = parse_cell_list(text, HEADER_COLUMNS, row_hint="")
    assert parsed == records


def test_parse_cell_list_value_with_pipe_raises():
    # 第 3 个字段（expectValue）含 |，会被 split 成多余字段 → 字段数超限报错
    with pytest.raises(ExcelValidationError) as exc:
        parse_cell_list("A|是|foo|bar||", HEADER_COLUMNS, row_hint="第 5 行 请求头")
    assert "第 5 行" in str(exc.value)


def test_sanitize_sheet_name_passthrough_valid():
    used: set[str] = set()
    assert sanitize_sheet_name("网管A", used) == "网管A"
    assert "网管A" in used


def test_sanitize_sheet_name_replaces_invalid_chars():
    used: set[str] = set()
    assert sanitize_sheet_name("A/B", used) == "A_B"
    assert sanitize_sheet_name(r"C\D", used) == "C_D"
    assert sanitize_sheet_name("E?F*G", used) == "E_F_G"
    assert sanitize_sheet_name("H[I]J", used) == "H_I_J"
    assert sanitize_sheet_name("K:L", used) == "K_L"


def test_sanitize_sheet_name_truncates_to_31_chars():
    used: set[str] = set()
    long_name = "网" * 40
    result = sanitize_sheet_name(long_name, used)
    assert len(result) == 31


def test_sanitize_sheet_name_appends_dedupe_suffix():
    used: set[str] = set()
    a = sanitize_sheet_name("A/B", used)
    b = sanitize_sheet_name("A_B", used)
    c = sanitize_sheet_name("A_B", used)
    assert a == "A_B"
    assert b == "A_B~2"
    assert c == "A_B~3"


def test_sanitize_sheet_name_dedupe_after_truncation():
    used: set[str] = set()
    long_a = "X" * 40
    long_b = "X" * 40
    assert sanitize_sheet_name(long_a, used) == "X" * 31
    # 第二次被截断成同名 → 追加后缀，且总长仍 ≤ 31
    result = sanitize_sheet_name(long_b, used)
    assert result.endswith("~2")
    assert len(result) <= 31


def test_sanitize_sheet_name_empty_or_whitespace_falls_back():
    used: set[str] = set()
    assert sanitize_sheet_name("", used) == "未命名"
    # 第二次调用同样触发 fallback，然后走 dedupe
    assert sanitize_sheet_name("   ", used) == "未命名~2"
    # 全非法字符也走 fallback（清洗后只剩 _ 也算"非空白"，仍然是 _；这个测保底空白路径）
    used2: set[str] = set()
    assert sanitize_sheet_name(" \t\n", used2) == "未命名"


def _make_api_row(**overrides):
    """构造一行 api_configs 数据（用 dict 模拟 sqlite3.Row）。"""
    base = {
        "id": "api_test1",
        "name": "测试接口",
        "method": "GET",
        "path": "/api/test",
        "enabled": 1,
        "group_name": None,
        "domain_id": None,
        "category": None,
        "data_source": "static",
        "topology_id": None,
        "sql_text": None,
        "config": "{}",
    }
    base.update(overrides)
    return base


def _load_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return load_workbook(buf)


def test_build_workbook_empty_produces_instruction_sheet_only():
    wb = build_workbook(api_rows=[], domains=[], topologies=[])
    reloaded = _load_bytes(wb)
    assert INSTRUCTION_SHEET_NAME in reloaded.sheetnames
    # 首个 Sheet 必须是使用说明
    assert reloaded.sheetnames[0] == INSTRUCTION_SHEET_NAME


def test_build_workbook_uncategorized_apis_go_to_special_sheet():
    apis = [_make_api_row(domain_id=None)]
    wb = build_workbook(api_rows=apis, domains=[], topologies=[])
    reloaded = _load_bytes(wb)
    assert UNCATEGORIZED_SHEET_NAME in reloaded.sheetnames
    ws = reloaded[UNCATEGORIZED_SHEET_NAME]
    # 第 1 行是表头
    assert [c.value for c in ws[1]] == MAIN_HEADERS
    # 第 2 行是数据
    assert ws.cell(row=2, column=1).value == "GET"
    assert ws.cell(row=2, column=2).value == "/api/test"


def test_build_workbook_per_domain_sheet_with_original_name_in_comment():
    apis = [_make_api_row(domain_id="dom_1")]
    domains = [{"id": "dom_1", "name": "网管/A"}]
    wb = build_workbook(api_rows=apis, domains=domains, topologies=[])
    reloaded = _load_bytes(wb)
    # 域名 "网管/A" 被清洗成 "网管_A"
    assert "网管_A" in reloaded.sheetnames
    ws = reloaded["网管_A"]
    # A1 comment 里存原始域名
    assert ws["A1"].comment is not None
    assert "网管/A" in ws["A1"].comment.text


def test_build_workbook_uses_topology_name_not_id():
    apis = [_make_api_row(topology_id="topo_x", data_source="sql", sql_text="SELECT 1")]
    topos = [{"id": "topo_x", "name": "MyTopo"}]
    wb = build_workbook(api_rows=apis, domains=[], topologies=topos)
    reloaded = _load_bytes(wb)
    ws = reloaded[UNCATEGORIZED_SHEET_NAME]
    # 第 8 列 = "拓扑"
    topo_col = MAIN_HEADERS.index("拓扑") + 1
    assert ws.cell(row=2, column=topo_col).value == "MyTopo"


def test_build_workbook_serializes_headers_query_params():
    config = {
        "request": {
            "headers": [{"name": "Auth", "required": True, "expectValue": "Bearer"}],
            "query": [{"name": "pageNo", "type": "int", "required": True}],
        },
        "params": [{"name": "id", "in": "query", "type": "string", "required": True, "bindTo": "id_"}],
    }
    apis = [_make_api_row(config=json.dumps(config))]
    wb = build_workbook(api_rows=apis, domains=[], topologies=[])
    reloaded = _load_bytes(wb)
    ws = reloaded[UNCATEGORIZED_SHEET_NAME]
    headers_col = MAIN_HEADERS.index("请求头") + 1
    query_col = MAIN_HEADERS.index("Query 参数") + 1
    param_col = MAIN_HEADERS.index("参数映射") + 1
    assert ws.cell(row=2, column=headers_col).value == "Auth|是|Bearer||"
    assert ws.cell(row=2, column=query_col).value == "pageNo|int|是||"
    assert ws.cell(row=2, column=param_col).value == "id|query|string|是|id_"


def test_build_workbook_fault_columns():
    config = {"fault": {"delayMs": 100, "errorRate": 0.5, "errorStatus": 503}}
    apis = [_make_api_row(config=json.dumps(config))]
    wb = build_workbook(api_rows=apis, domains=[], topologies=[])
    reloaded = _load_bytes(wb)
    ws = reloaded[UNCATEGORIZED_SHEET_NAME]
    delay_col = MAIN_HEADERS.index("故障-延迟毫秒") + 1
    rate_col = MAIN_HEADERS.index("故障-错误率") + 1
    status_col = MAIN_HEADERS.index("故障-错误状态码") + 1
    assert ws.cell(row=2, column=delay_col).value == 100
    assert ws.cell(row=2, column=rate_col).value == 0.5
    assert ws.cell(row=2, column=status_col).value == 503


def _build_test_workbook_bytes(builder):
    """helper: 传一个 (wb) -> None 的回调，写入完毕后返回 bytes。"""
    from openpyxl import Workbook as _Wb
    wb = _Wb()
    # 删掉默认 Sheet
    default = wb.active
    wb.remove(default)
    builder(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_workbook_ignores_underscore_sheets():
    def build(wb):
        ws = wb.create_sheet("_使用说明")
        ws["A1"] = "instructions"
        ws2 = wb.create_sheet("网管A")
        for col_idx, h in enumerate(MAIN_HEADERS, start=1):
            ws2.cell(row=1, column=col_idx, value=h)
        ws2.cell(row=2, column=1, value="GET")
        ws2.cell(row=2, column=2, value="/api/x")
        ws2.cell(row=2, column=3, value="X")
        ws2.cell(row=2, column=7, value="static")
    data = _build_test_workbook_bytes(build)

    result = parse_workbook(io.BytesIO(data), existing_domains=[], existing_topologies=[])
    assert isinstance(result, ParseResult)
    assert len(result.rows) == 1
    assert result.rows[0]["method"] == "GET"
    assert result.rows[0]["path"] == "/api/x"


def test_parse_workbook_empty_raises_fatal():
    def build(wb):
        wb.create_sheet("_使用说明")  # 只有说明 Sheet
    data = _build_test_workbook_bytes(build)

    with pytest.raises(ExcelValidationError) as exc:
        parse_workbook(io.BytesIO(data), existing_domains=[], existing_topologies=[])
    assert "未找到" in str(exc.value)


def test_parse_workbook_skip_row_without_method_or_path():
    def build(wb):
        ws = wb.create_sheet("网管A")
        for col_idx, h in enumerate(MAIN_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        # 第 2 行只有 path 没 method
        ws.cell(row=2, column=2, value="/api/y")
        # 第 3 行完整
        ws.cell(row=3, column=1, value="POST")
        ws.cell(row=3, column=2, value="/api/z")
        ws.cell(row=3, column=3, value="Z")
        ws.cell(row=3, column=7, value="static")
    data = _build_test_workbook_bytes(build)

    result = parse_workbook(io.BytesIO(data), existing_domains=[], existing_topologies=[])
    assert len(result.rows) == 1
    assert result.rows[0]["path"] == "/api/z"


def test_parse_workbook_uses_a1_comment_for_original_domain():
    def build(wb):
        ws = wb.create_sheet("网管_A")
        for col_idx, h in enumerate(MAIN_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        from openpyxl.comments import Comment
        ws["A1"].comment = Comment("__ORIGINAL_DOMAIN__=网管/A", "system")
        ws.cell(row=2, column=1, value="GET")
        ws.cell(row=2, column=2, value="/api/x")
        ws.cell(row=2, column=3, value="X")
        ws.cell(row=2, column=7, value="static")
    data = _build_test_workbook_bytes(build)

    result = parse_workbook(
        io.BytesIO(data),
        existing_domains=[{"id": "dom_orig", "name": "网管/A"}],
        existing_topologies=[],
    )
    # 应能识别原始域名并回填 domain_id
    assert result.rows[0]["domain_id"] == "dom_orig"


def test_parse_workbook_topology_name_resolution():
    def build(wb):
        ws = wb.create_sheet("网管A")
        for col_idx, h in enumerate(MAIN_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        ws.cell(row=2, column=1, value="GET")
        ws.cell(row=2, column=2, value="/a")
        ws.cell(row=2, column=3, value="A")
        ws.cell(row=2, column=7, value="sql")
        # 拓扑列 = 第 8 列
        ws.cell(row=2, column=8, value="TopoX")

        ws.cell(row=3, column=1, value="GET")
        ws.cell(row=3, column=2, value="/b")
        ws.cell(row=3, column=3, value="B")
        ws.cell(row=3, column=7, value="sql")
        ws.cell(row=3, column=8, value="不存在的拓扑")
    data = _build_test_workbook_bytes(build)

    result = parse_workbook(
        io.BytesIO(data),
        existing_domains=[],
        existing_topologies=[{"id": "topo_x", "name": "TopoX", "created_at": "2020-01-01"}],
    )
    assert result.rows[0]["topology_id"] == "topo_x"
    assert result.rows[1]["topology_id"] is None
    assert any("不存在的拓扑" in w for w in result.warnings)


def test_parse_workbook_invalid_method_row_error():
    def build(wb):
        ws = wb.create_sheet("网管A")
        for col_idx, h in enumerate(MAIN_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        ws.cell(row=2, column=1, value="FOO")
        ws.cell(row=2, column=2, value="/api/x")
        ws.cell(row=2, column=3, value="X")
        ws.cell(row=2, column=7, value="static")
    data = _build_test_workbook_bytes(build)

    result = parse_workbook(io.BytesIO(data), existing_domains=[], existing_topologies=[])
    assert result.rows == []
    assert any("FOO" in e for e in result.errors)


def test_parse_workbook_deserializes_headers_and_params():
    def build(wb):
        ws = wb.create_sheet("网管A")
        for col_idx, h in enumerate(MAIN_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        ws.cell(row=2, column=1, value="GET")
        ws.cell(row=2, column=2, value="/api/x")
        ws.cell(row=2, column=3, value="X")
        ws.cell(row=2, column=7, value="sql")
        # 请求头列 = 11
        ws.cell(row=2, column=11, value="Auth|是|Bearer||")
        # Query 参数列 = 12
        ws.cell(row=2, column=12, value="pageNo|int|是||")
        # 参数映射列 = 16
        ws.cell(row=2, column=16, value="id|query|string|是|id_")
        # SQL 列 = 14
        ws.cell(row=2, column=14, value="SELECT 1")
        # 响应模板列 = 15
        ws.cell(row=2, column=15, value="{}")
    data = _build_test_workbook_bytes(build)

    result = parse_workbook(io.BytesIO(data), existing_domains=[], existing_topologies=[])
    row = result.rows[0]
    assert row["config"]["request"]["headers"] == [
        {"name": "Auth", "required": True, "expectValue": "Bearer", "example": "", "description": ""}
    ]
    assert row["config"]["request"]["query"] == [
        {"name": "pageNo", "type": "int", "required": True, "example": "", "description": ""}
    ]
    assert row["config"]["params"] == [
        {"name": "id", "in": "query", "type": "string", "required": True, "bindTo": "id_"}
    ]
    assert row["sql_text"] == "SELECT 1"


def test_parse_workbook_fault_columns_optional():
    def build(wb):
        ws = wb.create_sheet("网管A")
        for col_idx, h in enumerate(MAIN_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        ws.cell(row=2, column=1, value="GET")
        ws.cell(row=2, column=2, value="/a")
        ws.cell(row=2, column=3, value="A")
        ws.cell(row=2, column=7, value="static")
        # 故障列 18/19/20
        ws.cell(row=2, column=18, value=100)
        ws.cell(row=2, column=19, value=0.5)
        ws.cell(row=2, column=20, value=503)

        # 第二行不写故障 → config.fault 不应存在
        ws.cell(row=3, column=1, value="GET")
        ws.cell(row=3, column=2, value="/b")
        ws.cell(row=3, column=3, value="B")
        ws.cell(row=3, column=7, value="static")
    data = _build_test_workbook_bytes(build)

    result = parse_workbook(io.BytesIO(data), existing_domains=[], existing_topologies=[])
    assert result.rows[0]["config"]["fault"] == {"delayMs": 100, "errorRate": 0.5, "errorStatus": 503}
    assert "fault" not in result.rows[1]["config"]


def test_parse_workbook_auto_creates_domain_from_new_sheet():
    def build(wb):
        ws = wb.create_sheet("新网管")
        for col_idx, h in enumerate(MAIN_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        ws.cell(row=2, column=1, value="GET")
        ws.cell(row=2, column=2, value="/a")
        ws.cell(row=2, column=3, value="A")
        ws.cell(row=2, column=7, value="static")
    data = _build_test_workbook_bytes(build)

    result = parse_workbook(io.BytesIO(data), existing_domains=[], existing_topologies=[])
    assert result.rows[0]["domain_id"] is None  # 待创建
    assert result.rows[0]["_new_domain_name"] == "新网管"
    assert "新网管" in result.auto_created_domains


def test_parse_workbook_uncategorized_sheet_has_no_domain():
    def build(wb):
        ws = wb.create_sheet(UNCATEGORIZED_SHEET_NAME)
        for col_idx, h in enumerate(MAIN_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        ws.cell(row=2, column=1, value="GET")
        ws.cell(row=2, column=2, value="/a")
        ws.cell(row=2, column=3, value="A")
        ws.cell(row=2, column=7, value="static")
    data = _build_test_workbook_bytes(build)

    result = parse_workbook(io.BytesIO(data), existing_domains=[], existing_topologies=[])
    assert result.rows[0]["domain_id"] is None
    assert result.rows[0].get("_new_domain_name") is None


def test_parse_workbook_default_sheet_without_headers_is_not_data():
    """默认空 Workbook 的 'Sheet' 应被识别为非数据 Sheet 而非静默 auto-create '域 Sheet'。"""
    def build(wb):
        wb.create_sheet("Sheet")  # 空 Sheet，无表头
    data = _build_test_workbook_bytes(build)

    with pytest.raises(ExcelValidationError) as exc:
        parse_workbook(io.BytesIO(data), existing_domains=[], existing_topologies=[])
    assert "未找到" in str(exc.value)


def test_parse_workbook_sheet_with_wrong_first_header_is_skipped():
    """首列不是 '方法' 的 Sheet 也被跳过（用户误在 Excel 里新建了工作表）。"""
    def build(wb):
        ws1 = wb.create_sheet("我的笔记")
        ws1.cell(row=1, column=1, value="随手记")
        ws2 = wb.create_sheet("网管A")
        for col_idx, h in enumerate(MAIN_HEADERS, start=1):
            ws2.cell(row=1, column=col_idx, value=h)
        ws2.cell(row=2, column=1, value="GET")
        ws2.cell(row=2, column=2, value="/api/x")
        ws2.cell(row=2, column=3, value="X")
        ws2.cell(row=2, column=7, value="static")
    data = _build_test_workbook_bytes(build)

    result = parse_workbook(io.BytesIO(data), existing_domains=[], existing_topologies=[])
    assert len(result.rows) == 1
    assert result.rows[0]["path"] == "/api/x"
    # "我的笔记" 不应作为域被 auto-create
    assert "我的笔记" not in result.auto_created_domains


# ============== 端到端测试：POST /apis/export ==============


def test_export_endpoint_returns_xlsx_binary(client):
    # 造 1 个域 + 1 个静态接口
    r = client.post("/admin/api/domains", json={"name": "网管X"})
    assert r.status_code == 200, r.text
    dom_id = r.json()["data"]["id"]

    r = client.post("/admin/api/apis", json={
        "method": "GET",
        "path": "/api/e2e-export",
        "name": "E2E导出",
        "enabled": True,
        "domainId": dom_id,
        "dataSource": "static",
        "config": {"staticBody": '{"ok":true}'},
    })
    assert r.status_code == 200, r.text

    r = client.post("/admin/api/apis/export", json={})
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    wb = load_workbook(io.BytesIO(r.content))
    assert "网管X" in wb.sheetnames
    ws = wb["网管X"]
    assert ws.cell(row=2, column=1).value == "GET"
    assert ws.cell(row=2, column=2).value == "/api/e2e-export"


def test_export_endpoint_filter_by_ids(client):
    r = client.post("/admin/api/apis", json={
        "method": "GET", "path": "/api/one", "name": "one",
        "dataSource": "static", "config": {},
    })
    api1_id = r.json()["data"]["id"]
    r = client.post("/admin/api/apis", json={
        "method": "GET", "path": "/api/two", "name": "two",
        "dataSource": "static", "config": {},
    })

    r = client.post("/admin/api/apis/export", json={"ids": [api1_id]})
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    # 未归类 Sheet 应只有 1 行数据
    ws = wb[UNCATEGORIZED_SHEET_NAME]
    assert ws.cell(row=2, column=2).value == "/api/one"
    assert ws.cell(row=3, column=2).value is None  # 只有一行


def _upload_xlsx_bytes(client, xlsx_bytes: bytes):
    return client.post(
        "/admin/api/apis/import",
        files={"file": ("test.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )


def test_import_endpoint_rejects_non_xlsx(client):
    r = client.post(
        "/admin/api/apis/import",
        files={"file": ("test.json", b'{"apis":[]}', "application/json")},
    )
    assert r.status_code == 400
    assert r.json()["detail"]["code"] == 40410


def test_import_endpoint_creates_new_api(client):
    def build(wb):
        ws = wb.create_sheet("新网管")
        for col_idx, h in enumerate(MAIN_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        ws.cell(row=2, column=1, value="GET")
        ws.cell(row=2, column=2, value="/api/imported")
        ws.cell(row=2, column=3, value="导入接口")
        ws.cell(row=2, column=7, value="static")
    data = _build_test_workbook_bytes(build)

    r = _upload_xlsx_bytes(client, data)
    assert r.status_code == 200, r.text
    result = r.json()["data"]
    assert result["created"] == 1
    assert result["updated"] == 0
    assert "新网管" in result["autoCreatedDomains"]

    # 校验 DB 里出现了
    r2 = client.get("/admin/api/apis")
    apis = r2.json()["data"]["items"]
    assert any(a["path"] == "/api/imported" for a in apis)


def test_import_endpoint_updates_existing_api(client):
    r = client.post("/admin/api/apis", json={
        "method": "GET", "path": "/api/exists", "name": "原名字",
        "dataSource": "static", "config": {},
    })
    original_id = r.json()["data"]["id"]

    def build(wb):
        ws = wb.create_sheet(UNCATEGORIZED_SHEET_NAME)
        for col_idx, h in enumerate(MAIN_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        ws.cell(row=2, column=1, value="GET")
        ws.cell(row=2, column=2, value="/api/exists")
        ws.cell(row=2, column=3, value="新名字")
        ws.cell(row=2, column=7, value="static")
    data = _build_test_workbook_bytes(build)

    r = _upload_xlsx_bytes(client, data)
    assert r.status_code == 200
    assert r.json()["data"]["updated"] == 1
    assert r.json()["data"]["created"] == 0

    r2 = client.get(f"/admin/api/apis/{original_id}")
    assert r2.json()["data"]["name"] == "新名字"


def test_import_endpoint_preserves_unknown_config_keys(client):
    # 预先造一个带未来字段的接口
    r = client.post("/admin/api/apis", json={
        "method": "POST", "path": "/api/keep-me", "name": "keep",
        "dataSource": "static",
        "config": {"customFuture": "x", "auth": {"type": "none"}},
    })

    # 用 Excel 更新它（只碰鉴权）
    def build(wb):
        ws = wb.create_sheet(UNCATEGORIZED_SHEET_NAME)
        for col_idx, h in enumerate(MAIN_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        ws.cell(row=2, column=1, value="POST")
        ws.cell(row=2, column=2, value="/api/keep-me")
        ws.cell(row=2, column=3, value="keep")
        ws.cell(row=2, column=7, value="static")
        # 鉴权类型 = 第 9 列
        ws.cell(row=2, column=9, value="xtoken")
        # 鉴权头名 = 第 10 列
        ws.cell(row=2, column=10, value="X-New-Token")
    data = _build_test_workbook_bytes(build)

    r = _upload_xlsx_bytes(client, data)
    assert r.status_code == 200

    r2 = client.get(f"/admin/api/apis?path=/api/keep-me")
    items = r2.json()["data"]["items"]
    assert len(items) == 1
    detail = client.get(f"/admin/api/apis/{items[0]['id']}").json()["data"]
    assert detail["config"]["customFuture"] == "x"
    assert detail["config"]["auth"] == {"type": "xtoken", "headerName": "X-New-Token"}


def test_import_endpoint_cross_sheet_move_changes_domain(client):
    # 先造两个域
    dom1 = client.post("/admin/api/domains", json={"name": "网管1"}).json()["data"]["id"]
    dom2 = client.post("/admin/api/domains", json={"name": "网管2"}).json()["data"]["id"]

    # 造一个绑到网管1 的接口
    client.post("/admin/api/apis", json={
        "method": "GET", "path": "/api/mover", "name": "mover",
        "domainId": dom1, "dataSource": "static", "config": {},
    })

    # 用 Excel 把它挪到网管2
    def build(wb):
        ws = wb.create_sheet("网管2")
        for col_idx, h in enumerate(MAIN_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        ws.cell(row=2, column=1, value="GET")
        ws.cell(row=2, column=2, value="/api/mover")
        ws.cell(row=2, column=3, value="mover")
        ws.cell(row=2, column=7, value="static")
    data = _build_test_workbook_bytes(build)

    r = _upload_xlsx_bytes(client, data)
    assert r.status_code == 200
    assert r.json()["data"]["updated"] == 1

    # 接口应该在 dom2 里，不在 dom1 里
    r_dom1 = client.get(f"/admin/api/apis?domainId={dom1}").json()["data"]["items"]
    r_dom2 = client.get(f"/admin/api/apis?domainId={dom2}").json()["data"]["items"]
    assert not any(a["path"] == "/api/mover" for a in r_dom1)
    assert any(a["path"] == "/api/mover" for a in r_dom2)


def test_import_endpoint_reports_row_errors_but_continues(client):
    def build(wb):
        ws = wb.create_sheet(UNCATEGORIZED_SHEET_NAME)
        for col_idx, h in enumerate(MAIN_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        # 坏行：非法 method
        ws.cell(row=2, column=1, value="FOO")
        ws.cell(row=2, column=2, value="/api/bad")
        ws.cell(row=2, column=3, value="bad")
        ws.cell(row=2, column=7, value="static")
        # 好行
        ws.cell(row=3, column=1, value="GET")
        ws.cell(row=3, column=2, value="/api/good")
        ws.cell(row=3, column=3, value="good")
        ws.cell(row=3, column=7, value="static")
    data = _build_test_workbook_bytes(build)

    r = _upload_xlsx_bytes(client, data)
    assert r.status_code == 200
    result = r.json()["data"]
    assert result["created"] == 1
    assert len(result["errors"]) == 1
    assert "FOO" in result["errors"][0]


def test_import_endpoint_fatal_error_on_empty_workbook(client):
    def build(wb):
        wb.create_sheet(INSTRUCTION_SHEET_NAME)
    data = _build_test_workbook_bytes(build)

    r = _upload_xlsx_bytes(client, data)
    assert r.status_code == 400
    assert "未找到" in r.json()["detail"]["message"]


def test_import_endpoint_composite_key_fully_replaced(client):
    """锁定 spec 的"Excel 权威视图"语义：Excel 里的 request/auth/fault 等 composite 键
    整体覆盖 DB 里对应键；空 cell 表示"该字段应被清除"，而不是"保留旧值"。
    只有真正未建模的顶层键（如 customFuture）在 UPDATE 时保留。"""
    # 预造：一个 request 里有 headers + query 两个数组的接口
    r = client.post("/admin/api/apis", json={
        "method": "GET",
        "path": "/api/replace-me",
        "name": "replace",
        "dataSource": "static",
        "config": {
            "request": {
                "headers": [{"name": "H1", "required": True}],
                "query": [{"name": "q1", "type": "string", "required": True}],
            },
            "customFuture": "keep_me",
        },
    })
    assert r.status_code == 200
    api_id = r.json()["data"]["id"]

    # Excel 只填 headers cell（覆盖为 H2），query cell 留空 → 期望 query 被清除
    def build(wb):
        ws = wb.create_sheet(UNCATEGORIZED_SHEET_NAME)
        for col_idx, h in enumerate(MAIN_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        ws.cell(row=2, column=1, value="GET")
        ws.cell(row=2, column=2, value="/api/replace-me")
        ws.cell(row=2, column=3, value="replace")
        ws.cell(row=2, column=7, value="static")
        # 请求头列 = 11：只写一个新 header
        ws.cell(row=2, column=11, value="H2|是|||")
        # Query 参数列 = 12：留空
    data = _build_test_workbook_bytes(build)

    r = _upload_xlsx_bytes(client, data)
    assert r.status_code == 200
    assert r.json()["data"]["updated"] == 1

    detail = client.get(f"/admin/api/apis/{api_id}").json()["data"]
    # request 整块被替换：headers 是新的 H2，query 因为 cell 空而不存在
    assert detail["config"]["request"] == {
        "headers": [{"name": "H2", "required": True, "expectValue": "", "example": "", "description": ""}]
    }
    # customFuture 是"未建模"的顶层键，保留
    assert detail["config"]["customFuture"] == "keep_me"


def test_build_workbook_static_body_dict_is_json_stringified():
    """历史数据 staticBody 可能是 dict，导出不能崩。"""
    config = {"staticBody": {"accessSession": "x", "expires": 1800}}
    apis = [_make_api_row(config=json.dumps(config))]
    wb = build_workbook(api_rows=apis, domains=[], topologies=[])
    reloaded = _load_bytes(wb)
    ws = reloaded[UNCATEGORIZED_SHEET_NAME]
    body_col = MAIN_HEADERS.index("静态响应体") + 1
    cell = ws.cell(row=2, column=body_col).value
    # 应是 JSON 字符串（含 accessSession 关键字）
    assert isinstance(cell, str)
    assert "accessSession" in cell


def test_build_workbook_response_template_nested_format():
    """老数据 config.response.template（嵌套）应被识别，兼容新格式 responseTemplate。"""
    config = {"response": {"template": "[\"{{items}}\"]"}}
    apis = [_make_api_row(config=json.dumps(config))]
    wb = build_workbook(api_rows=apis, domains=[], topologies=[])
    reloaded = _load_bytes(wb)
    ws = reloaded[UNCATEGORIZED_SHEET_NAME]
    tpl_col = MAIN_HEADERS.index("响应模板") + 1
    assert ws.cell(row=2, column=tpl_col).value == "[\"{{items}}\"]"


def test_build_workbook_response_template_top_level_still_works():
    """新数据 config.responseTemplate 顶层字段继续工作。"""
    config = {"responseTemplate": "{\"count\":\"{{total}}\"}"}
    apis = [_make_api_row(config=json.dumps(config))]
    wb = build_workbook(api_rows=apis, domains=[], topologies=[])
    reloaded = _load_bytes(wb)
    ws = reloaded[UNCATEGORIZED_SHEET_NAME]
    tpl_col = MAIN_HEADERS.index("响应模板") + 1
    assert ws.cell(row=2, column=tpl_col).value == "{\"count\":\"{{total}}\"}"


def test_parse_workbook_renamed_sheet_ignores_stale_a1_comment():
    """用户重命名 Sheet 后，A1 comment 里的原始域名应被视为过时，走新 Sheet 名。"""
    def build(wb):
        # 模拟：导出时 Sheet 名叫 "ManageOne网管"，A1 comment 里存 "__ORIGINAL_DOMAIN__=ManageOne网管"
        # 用户随后把 Sheet 重命名为 "新网管"
        ws = wb.create_sheet("新网管")
        for col_idx, h in enumerate(MAIN_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        from openpyxl.comments import Comment
        ws["A1"].comment = Comment("__ORIGINAL_DOMAIN__=ManageOne网管", "system")
        ws.cell(row=2, column=1, value="GET")
        ws.cell(row=2, column=2, value="/api/x")
        ws.cell(row=2, column=3, value="X")
        ws.cell(row=2, column=7, value="static")
    data = _build_test_workbook_bytes(build)

    # 已有域里有 ManageOne网管，但用户重命名了 Sheet 为 新网管
    result = parse_workbook(
        io.BytesIO(data),
        existing_domains=[{"id": "dom_manageone", "name": "ManageOne网管", "created_at": "2020-01-01"}],
        existing_topologies=[],
    )
    row = result.rows[0]
    # domain_id 应为 None（新域待创建），不应是 dom_manageone
    assert row["domain_id"] is None
    assert row["_new_domain_name"] == "新网管"
    assert "新网管" in result.auto_created_domains
    assert "ManageOne网管" not in result.auto_created_domains


def test_parse_workbook_a1_comment_still_wins_when_sheet_name_is_sanitized():
    """A1 comment 存 "网管/A"，Sheet 名被清洗成 "网管_A"，comment 仍应生效。"""
    def build(wb):
        ws = wb.create_sheet("网管_A")
        for col_idx, h in enumerate(MAIN_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        from openpyxl.comments import Comment
        ws["A1"].comment = Comment("__ORIGINAL_DOMAIN__=网管/A", "system")
        ws.cell(row=2, column=1, value="GET")
        ws.cell(row=2, column=2, value="/api/x")
        ws.cell(row=2, column=3, value="X")
        ws.cell(row=2, column=7, value="static")
    data = _build_test_workbook_bytes(build)

    result = parse_workbook(
        io.BytesIO(data),
        existing_domains=[{"id": "dom_a", "name": "网管/A", "created_at": "2020-01-01"}],
        existing_topologies=[],
    )
    assert result.rows[0]["domain_id"] == "dom_a"


def test_parse_workbook_renamed_sheet_matches_existing_other_domain():
    """用户重命名 Sheet 到另一个已存在的域名 → 迁移到那个已存在域。"""
    def build(wb):
        ws = wb.create_sheet("网管B")
        for col_idx, h in enumerate(MAIN_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        from openpyxl.comments import Comment
        ws["A1"].comment = Comment("__ORIGINAL_DOMAIN__=网管A", "system")
        ws.cell(row=2, column=1, value="GET")
        ws.cell(row=2, column=2, value="/api/x")
        ws.cell(row=2, column=3, value="X")
        ws.cell(row=2, column=7, value="static")
    data = _build_test_workbook_bytes(build)

    result = parse_workbook(
        io.BytesIO(data),
        existing_domains=[
            {"id": "dom_a", "name": "网管A", "created_at": "2020-01-01"},
            {"id": "dom_b", "name": "网管B", "created_at": "2020-01-02"},
        ],
        existing_topologies=[],
    )
    assert result.rows[0]["domain_id"] == "dom_b"


def test_parse_workbook_matches_sanitized_form_helper():
    """单元级：_matches_sanitized_form 的关键 case。"""
    from app.admin._api_excel import _matches_sanitized_form
    assert _matches_sanitized_form("网管/A", "网管_A") is True
    assert _matches_sanitized_form("网管A", "网管A") is True
    assert _matches_sanitized_form("网管/A", "网管_A~2") is True  # dedupe 情况
    assert _matches_sanitized_form("ManageOne网管", "新网管") is False  # 重命名场景
    assert _matches_sanitized_form("", "未命名") is True  # 空名 fallback
