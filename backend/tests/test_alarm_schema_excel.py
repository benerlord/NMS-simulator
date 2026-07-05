"""告警模板 Excel 导入导出 e2e 测试。"""
import io

import openpyxl


def _create_schema(client, code: str, name: str, description: str = "",
                    display_field_key: str = None, fields: list = None) -> str:
    payload = {
        "code": code,
        "name": name,
        "description": description,
        "displayFieldKey": display_field_key,
        "fields": fields or [
            {"fieldKey": "level", "fieldLabel": "级别", "fieldType": "text",
             "maxLength": 20, "required": True, "sortOrder": 0}
        ],
    }
    r = client.post("/admin/api/alarm-schemas", json=payload)
    assert r.status_code == 200, r.text
    return r.json()["data"]["id"]


# ============== 导出测试 ==============

def test_export_all_returns_xlsx_with_summary_sheet(client):
    """导出全部 → xlsx 含'模板汇总' Sheet + 每 code 独立字段 Sheet。"""
    _create_schema(client, "as_a", "模板A")
    _create_schema(client, "as_b", "模板B")

    r = client.post("/admin/api/alarm-schemas/export", json={})
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert "模板汇总" in wb.sheetnames
    assert "as_a" in wb.sheetnames
    assert "as_b" in wb.sheetnames


def test_export_ids_only_returns_selected(client):
    """按 ids 导出 → xlsx 只含指定模板。"""
    sid_a = _create_schema(client, "as_only_a", "只A")
    _create_schema(client, "as_only_b", "只B")

    r = client.post("/admin/api/alarm-schemas/export", json={"ids": [sid_a]})
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert "as_only_a" in wb.sheetnames
    assert "as_only_b" not in wb.sheetnames


def test_export_summary_contains_display_field_key_column(client):
    """汇总 Sheet 表头含'展示字段Key'列，值正确。"""
    _create_schema(
        client, "as_disp", "模板显示", display_field_key="level",
        fields=[
            {"fieldKey": "level", "fieldLabel": "级别", "fieldType": "text",
             "maxLength": 20, "required": True, "sortOrder": 0}
        ],
    )

    r = client.post("/admin/api/alarm-schemas/export", json={})
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb["模板汇总"]
    headers = [c.value for c in ws[1] if c.value]
    assert "展示字段Key" in headers

    # 找到 as_disp 行
    idx = {h: i for i, h in enumerate(headers)}
    code_col = idx["Code"] + 1
    disp_col = idx["展示字段Key"] + 1
    found = None
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[code_col - 1].value == "as_disp":
            found = row[disp_col - 1].value
            break
    assert found == "level"


def test_export_field_sheet_contains_mapping_target_column(client):
    """每模板字段 Sheet 表头含'映射节点属性'列，值正确。"""
    _create_schema(
        client, "as_map", "映射模板",
        fields=[
            {"fieldKey": "severity", "fieldLabel": "严重度",
             "fieldType": "text", "maxLength": 20,
             "mappingTarget": "node_severity", "sortOrder": 0}
        ],
    )

    r = client.post("/admin/api/alarm-schemas/export", json={})
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb["as_map"]
    headers = [c.value for c in ws[1] if c.value]
    assert "映射节点属性" in headers

    idx = {h: i for i, h in enumerate(headers)}
    map_col_1based = idx["映射节点属性"] + 1
    assert ws.cell(row=2, column=map_col_1based).value == "node_severity"


# ============== 导入 preview 测试 ==============

def _build_import_xlsx(rows: list) -> io.BytesIO:
    """构造一份最小导入 xlsx。rows: [{code, name, description, displayFieldKey, fields: [...]}]"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "模板汇总"
    ws.append(["Code", "名称", "描述", "展示字段Key"])
    for r in rows:
        ws.append([r.get("code"), r.get("name"), r.get("description"), r.get("displayFieldKey")])

    for r in rows:
        code = r.get("code")
        fields = r.get("fields") or [
            {"fieldKey": "level", "fieldLabel": "级别", "fieldType": "text",
             "maxLength": 20, "required": True, "sortOrder": 0}
        ]
        fs = wb.create_sheet(title=code)
        fs.append(["字段标识", "显示名称", "字段类型", "最大长度",
                    "默认值", "选项", "必填", "排序", "映射节点属性"])
        for f in fields:
            fs.append([
                f.get("fieldKey"), f.get("fieldLabel"), f.get("fieldType"),
                f.get("maxLength"), f.get("defaultValue"), f.get("options"),
                "是" if f.get("required") else "否",
                f.get("sortOrder", 0), f.get("mappingTarget"),
            ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_import_preview_categorizes_create_and_update(client):
    """已存在 code → toUpdate；不存在 → toCreate。"""
    _create_schema(client, "as_exists", "已存在")
    buf = _build_import_xlsx([
        {"code": "as_exists", "name": "已存在改名"},
        {"code": "as_new_1", "name": "新1"},
        {"code": "as_new_2", "name": "新2"},
    ])

    r = client.post(
        "/admin/api/alarm-schemas/import/preview",
        files={"file": ("test.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text
    data = r.json()["data"]

    to_create_codes = [item["code"] for item in data["toCreate"]]
    to_update_codes = [item["code"] for item in data["toUpdate"]]
    assert sorted(to_create_codes) == ["as_new_1", "as_new_2"]
    assert to_update_codes == ["as_exists"]


def test_import_preview_records_old_name_on_update(client):
    """覆盖项 oldName 与新 name 不同时正确记录。"""
    _create_schema(client, "as_rename", "旧名字")
    buf = _build_import_xlsx([
        {"code": "as_rename", "name": "新名字"},
    ])

    r = client.post(
        "/admin/api/alarm-schemas/import/preview",
        files={"file": ("test.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200
    item = r.json()["data"]["toUpdate"][0]
    assert item["code"] == "as_rename"
    assert item["name"] == "新名字"
    assert item["oldName"] == "旧名字"


def test_import_preview_missing_summary_sheet_returns_400(client):
    """缺'模板汇总' Sheet → 400。"""
    wb = openpyxl.Workbook()
    wb.active.title = "OtherSheet"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    r = client.post(
        "/admin/api/alarm-schemas/import/preview",
        files={"file": ("bad.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 400
    assert "模板汇总" in r.json()["detail"]["message"]


def test_import_preview_missing_required_records_error(client):
    """Code 或名称缺失的行 → errors 记录。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "模板汇总"
    ws.append(["Code", "名称", "描述"])
    ws.append([None, "缺 code 的行", "desc"])
    ws.append(["as_ok", None, "缺 name 的行"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    r = client.post(
        "/admin/api/alarm-schemas/import/preview",
        files={"file": ("bad.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200
    data = r.json()["data"]
    assert len(data["errors"]) == 2
    assert data["toCreate"] == []
    assert data["toUpdate"] == []


# ============== 正式导入测试 ==============

def test_import_creates_new_schema_with_fields(client):
    """新模板 + 字段一起导入。"""
    buf = _build_import_xlsx([
        {"code": "as_imp_new", "name": "导入新",
         "fields": [
             {"fieldKey": "severity", "fieldLabel": "严重度",
              "fieldType": "text", "maxLength": 30, "sortOrder": 0},
             {"fieldKey": "count", "fieldLabel": "计数",
              "fieldType": "number", "sortOrder": 1},
         ]},
    ])

    r = client.post(
        "/admin/api/alarm-schemas/import",
        files={"file": ("test.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text
    data = r.json()["data"]
    assert data["created"] == 1
    assert data["updated"] == 0
    assert data["totalFields"] == 2

    r = client.get("/admin/api/alarm-schemas").json()["data"]
    match = next(it for it in r if it["code"] == "as_imp_new")
    detail = client.get(f"/admin/api/alarm-schemas/{match['id']}").json()["data"]
    assert [f["fieldKey"] for f in detail["fields"]] == ["severity", "count"]


def test_import_overwrites_existing_schema(client):
    """已存在 code → name 覆盖，字段全部替换。"""
    _create_schema(
        client, "as_imp_ov", "旧名字",
        fields=[
            {"fieldKey": "old_field", "fieldLabel": "旧字段",
             "fieldType": "text", "maxLength": 10, "sortOrder": 0}
        ],
    )
    buf = _build_import_xlsx([
        {"code": "as_imp_ov", "name": "新名字",
         "fields": [
             {"fieldKey": "new_field", "fieldLabel": "新字段",
              "fieldType": "text", "maxLength": 30, "sortOrder": 0}
         ]},
    ])

    r = client.post(
        "/admin/api/alarm-schemas/import",
        files={"file": ("test.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200
    data = r.json()["data"]
    assert data["updated"] == 1
    assert data["created"] == 0

    lst = client.get("/admin/api/alarm-schemas").json()["data"]
    match = next(it for it in lst if it["code"] == "as_imp_ov")
    assert match["name"] == "新名字"
    detail = client.get(f"/admin/api/alarm-schemas/{match['id']}").json()["data"]
    field_keys = [f["fieldKey"] for f in detail["fields"]]
    assert field_keys == ["new_field"]
    assert "old_field" not in field_keys


def test_import_text_field_missing_maxlen_defaults_to_255(client):
    """text 字段最大长度为空 → 落库 255。"""
    buf = _build_import_xlsx([
        {"code": "as_imp_mx", "name": "空长度",
         "fields": [
             {"fieldKey": "note", "fieldLabel": "备注",
              "fieldType": "text", "maxLength": None, "sortOrder": 0}
         ]},
    ])

    r = client.post(
        "/admin/api/alarm-schemas/import",
        files={"file": ("test.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200
    lst = client.get("/admin/api/alarm-schemas").json()["data"]
    match = next(it for it in lst if it["code"] == "as_imp_mx")
    detail = client.get(f"/admin/api/alarm-schemas/{match['id']}").json()["data"]
    note = next(f for f in detail["fields"] if f["fieldKey"] == "note")
    assert note["maxLength"] == 255


def test_import_invalid_field_type_records_error(client):
    """field_type 非白名单 → errors 记录跳过。"""
    buf = _build_import_xlsx([
        {"code": "as_imp_bad", "name": "有非法类型",
         "fields": [
             {"fieldKey": "good_field", "fieldLabel": "合法",
              "fieldType": "text", "maxLength": 20, "sortOrder": 0},
             {"fieldKey": "bad_field", "fieldLabel": "非法",
              "fieldType": "invalid_type", "sortOrder": 1},
         ]},
    ])

    r = client.post(
        "/admin/api/alarm-schemas/import",
        files={"file": ("test.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200
    data = r.json()["data"]
    assert data["created"] == 1
    assert data["totalFields"] == 1  # 只有 good_field 入库
    assert any("invalid_type" in e for e in data["errors"])


def test_import_field_key_conflict_with_fixed_col_records_error(client):
    """field_key 与固定列（id/node_id/alarm_index/created_at/updated_at）冲突 → errors 记录。"""
    buf = _build_import_xlsx([
        {"code": "as_imp_fx", "name": "固定列冲突",
         "fields": [
             {"fieldKey": "id", "fieldLabel": "冲突 id",
              "fieldType": "text", "maxLength": 20, "sortOrder": 0},
             {"fieldKey": "safe_key", "fieldLabel": "安全",
              "fieldType": "text", "maxLength": 20, "sortOrder": 1},
         ]},
    ])

    r = client.post(
        "/admin/api/alarm-schemas/import",
        files={"file": ("test.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200
    data = r.json()["data"]
    assert data["totalFields"] == 1
    assert any("id" in e and "固定列" in e for e in data["errors"])


def test_import_invalid_mapping_target_records_error(client):
    """mapping_target 非合法标识符 → errors 记录跳过。"""
    buf = _build_import_xlsx([
        {"code": "as_imp_mp", "name": "非法映射",
         "fields": [
             {"fieldKey": "level", "fieldLabel": "级别",
              "fieldType": "text", "maxLength": 20,
              "mappingTarget": "123abc", "sortOrder": 0},
         ]},
    ])

    r = client.post(
        "/admin/api/alarm-schemas/import",
        files={"file": ("test.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200
    data = r.json()["data"]
    assert data["totalFields"] == 0
    assert any("123abc" in e for e in data["errors"])


def test_import_partial_failure_isolated_per_row(client):
    """某模板字段解析失败不影响其他模板。"""
    buf = _build_import_xlsx([
        {"code": "as_imp_ok", "name": "正常",
         "fields": [
             {"fieldKey": "level", "fieldLabel": "级别",
              "fieldType": "text", "maxLength": 20, "sortOrder": 0},
         ]},
        {"code": "as_imp_bad2", "name": "含错误字段",
         "fields": [
             {"fieldKey": "bad", "fieldLabel": "坏",
              "fieldType": "not_a_type", "sortOrder": 0},
         ]},
        {"code": "as_imp_ok2", "name": "又正常",
         "fields": [
             {"fieldKey": "count", "fieldLabel": "计数",
              "fieldType": "number", "sortOrder": 0},
         ]},
    ])

    r = client.post(
        "/admin/api/alarm-schemas/import",
        files={"file": ("test.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200
    data = r.json()["data"]
    assert data["created"] == 3  # 3 个模板都创建
    assert data["totalFields"] == 2  # 只有 2 个字段成功
    assert len(data["errors"]) >= 1
