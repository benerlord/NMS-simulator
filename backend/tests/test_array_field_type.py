"""字段类型新增 array 测试 — 覆盖 spec §3.1 / §3.2 + §3.6 Excel I/O。"""
import json

import pytest


# ============================================================
# 1. Pattern 接受 array
# ============================================================

def test_array_field_pattern_accepted(client):
    """fieldType='array' 通过 pattern 校验。"""
    r = client.post("/admin/api/node-types", json={
        "code": "arr_router", "name": "ArrRouter", "category": "physical",
        "fields": [
            {"fieldKey": "ports", "fieldLabel": "端口列表", "fieldType": "array"},
        ],
    })
    assert r.status_code == 200, r.text


def test_create_node_type_with_array_field(client):
    """POST 类型含 array 字段 → 落库 + GET 回来类型保留。"""
    r = client.post("/admin/api/node-types", json={
        "code": "arr_dev", "name": "ArrDev", "category": "physical",
        "fields": [
            {"fieldKey": "tags", "fieldLabel": "标签", "fieldType": "array"},
        ],
    })
    tid = r.json()["data"]["id"]
    r = client.get(f"/admin/api/node-types/{tid}")
    fields = r.json()["data"]["fields"]
    assert len(fields) == 1
    assert fields[0]["fieldType"] == "array"


# ============================================================
# 2. validate_array_default
# ============================================================

def test_create_with_array_default_value_valid(client):
    """defaultValue='[\"a\",\"b\"]' → 成功。"""
    r = client.post("/admin/api/node-types", json={
        "code": "arr_dv1", "name": "D1", "category": "physical",
        "fields": [
            {"fieldKey": "tags", "fieldLabel": "标签", "fieldType": "array",
             "defaultValue": '["a","b"]'},
        ],
    })
    assert r.status_code == 200, r.text
    tid = r.json()["data"]["id"]
    fields = client.get(f"/admin/api/node-types/{tid}").json()["data"]["fields"]
    assert fields[0]["defaultValue"] == '["a","b"]'


def test_create_with_array_default_value_invalid_not_list(client):
    """defaultValue='\"abc\"' → 422（不是 array）。"""
    r = client.post("/admin/api/node-types", json={
        "code": "arr_dv2", "name": "D2", "category": "physical",
        "fields": [
            {"fieldKey": "tags", "fieldLabel": "标签", "fieldType": "array",
             "defaultValue": '"abc"'},
        ],
    })
    assert r.status_code == 422
    assert "JSON array" in r.text


def test_create_with_array_default_value_invalid_json(client):
    """defaultValue='[1,2' → 422（语法错）。"""
    r = client.post("/admin/api/node-types", json={
        "code": "arr_dv3", "name": "D3", "category": "physical",
        "fields": [
            {"fieldKey": "tags", "fieldLabel": "标签", "fieldType": "array",
             "defaultValue": "[1,2"},
        ],
    })
    assert r.status_code == 422
    assert "合法 JSON" in r.text


def test_create_with_array_default_value_empty_array(client):
    """defaultValue='[]' → 成功。"""
    r = client.post("/admin/api/node-types", json={
        "code": "arr_dv4", "name": "D4", "category": "physical",
        "fields": [
            {"fieldKey": "tags", "fieldLabel": "标签", "fieldType": "array",
             "defaultValue": "[]"},
        ],
    })
    assert r.status_code == 200, r.text


def test_create_with_array_default_value_null(client):
    """defaultValue 不传 → 成功（不校验）。"""
    r = client.post("/admin/api/node-types", json={
        "code": "arr_dv5", "name": "D5", "category": "physical",
        "fields": [
            {"fieldKey": "tags", "fieldLabel": "标签", "fieldType": "array"},
        ],
    })
    assert r.status_code == 200, r.text


# ============================================================
# 3. 节点 attrs 接口接受 array JSON 字符串
# ============================================================

def test_set_attrs_with_json_array_string(client):
    """PUT node attrs value='[\"a\",\"b\"]' → GET 回来一致。"""
    r = client.post("/admin/api/node-types", json={
        "code": "arr_n", "name": "N", "category": "physical",
        "fields": [{"fieldKey": "tags", "fieldLabel": "标签", "fieldType": "array"}],
    })
    tid = r.json()["data"]["id"]
    r = client.post("/admin/api/topologies", json={"name": "T-arr"})
    topo = r.json()["data"]["id"]
    n1 = client.post(f"/admin/api/topologies/{topo}/nodes",
                     json={"nodeTypeId": tid, "name": "n1"}).json()["data"]["id"]

    r = client.put(f"/admin/api/nodes/{n1}/attrs", json=[
        {"fieldKey": "tags", "value": '["a","b","c"]'},
    ])
    assert r.status_code == 200, r.text

    attrs = client.get(f"/admin/api/nodes/{n1}").json()["data"].get("attrs", {})
    assert attrs.get("tags") == '["a","b","c"]'


# ============================================================
# 4. 改字段类型为 array
# ============================================================

def test_update_node_type_change_field_type_to_array(client):
    """已有 text 字段，PUT 改 fieldType 为 array → 成功。"""
    r = client.post("/admin/api/node-types", json={
        "code": "arr_chg", "name": "Chg", "category": "physical",
        "fields": [
            {"fieldKey": "tags", "fieldLabel": "标签", "fieldType": "text", "maxLength": 100},
        ],
    })
    tid = r.json()["data"]["id"]

    r = client.put(f"/admin/api/node-types/{tid}", json={
        "fields": [
            {"fieldKey": "tags", "fieldLabel": "标签", "fieldType": "array"},
        ],
    })
    assert r.status_code == 200, r.text

    fields = client.get(f"/admin/api/node-types/{tid}").json()["data"]["fields"]
    assert fields[0]["fieldType"] == "array"


# ============================================================
# 5. 边类型 + 告警模板对称
# ============================================================

def test_edge_type_array_field_symmetric(client):
    """边类型 fields[fieldType='array'] → 成功。"""
    r = client.post("/admin/api/edge-types", json={
        "code": "arr_edge", "name": "ArrEdge",
        "fields": [
            {"fieldKey": "subnets", "fieldLabel": "子网", "fieldType": "array",
             "defaultValue": '["10.0.0.0/24"]'},
        ],
    })
    assert r.status_code == 200, r.text
    tid = r.json()["data"]["id"]
    fields = client.get(f"/admin/api/edge-types/{tid}").json()["data"]["fields"]
    assert fields[0]["fieldType"] == "array"


def test_alarm_schema_array_field(client):
    """告警模板 fields[fieldType='array'] → 成功。"""
    r = client.post("/admin/api/alarm-schemas", json={
        "code": "arr_alarm", "name": "ArrAlarm",
        "fields": [
            {"fieldKey": "tags", "fieldLabel": "标签", "fieldType": "array",
             "defaultValue": "[]"},
        ],
    })
    assert r.status_code == 200, r.text


# ============================================================
# 6. Excel I/O
# ============================================================

def test_excel_export_array_field_default_preserved(client):
    """导出 → array 字段的默认值 JSON 字符串保留在 cell 里。"""
    from io import BytesIO
    from openpyxl import load_workbook

    r = client.post("/admin/api/node-types", json={
        "code": "arr_xl1", "name": "XL1", "category": "physical",
        "fields": [
            {"fieldKey": "tags", "fieldLabel": "标签", "fieldType": "array",
             "defaultValue": '["a","b"]'},
        ],
    })
    tid = r.json()["data"]["id"]

    r = client.post("/admin/api/node-types/export", json={"ids": [tid]})
    assert r.status_code == 200

    wb = load_workbook(BytesIO(r.content))
    sheet = wb["arr_xl1"]
    headers = {cell.value: idx for idx, cell in enumerate(sheet[1], start=1)}
    row2 = sheet[2]
    default_cell = row2[headers["默认值"] - 1]
    type_cell = row2[headers["字段类型"] - 1]
    assert type_cell.value == "array"
    assert default_cell.value == '["a","b"]'


def test_excel_export_numeric_array_field_default_preserved(client):
    """导出 → 纯数字数组 [1,2,3] 也以原 JSON 字符串保留（不被 openpyxl 识别为数字）。"""
    from io import BytesIO
    from openpyxl import load_workbook

    r = client.post("/admin/api/node-types", json={
        "code": "arr_xl_num", "name": "XLNum", "category": "physical",
        "fields": [
            {"fieldKey": "ports", "fieldLabel": "端口", "fieldType": "array",
             "defaultValue": "[1,2,3]"},
        ],
    })
    tid = r.json()["data"]["id"]

    r = client.post("/admin/api/node-types/export", json={"ids": [tid]})
    assert r.status_code == 200

    wb = load_workbook(BytesIO(r.content))
    sheet = wb["arr_xl_num"]
    headers = {cell.value: idx for idx, cell in enumerate(sheet[1], start=1)}
    default_cell = sheet[2][headers["默认值"] - 1]
    # cell 值应保持为字符串 "[1,2,3]"，不被识别为数字
    assert default_cell.value == "[1,2,3]"
    assert isinstance(default_cell.value, str), f"expected str, got {type(default_cell.value).__name__}"


def test_excel_import_array_field_default_parsed(client):
    """导入 cell '[\"a\",\"b\"]'（fieldType=array）→ 默认值落库。"""
    from io import BytesIO
    from openpyxl import Workbook

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "类型汇总"
    ws1.append(["编码", "名称", "分类", "图标", "颜色", "形状",
                "渲染模式", "DN模板", "描述", "创建时间", "更新时间"])
    ws1.append(["xl_imp1", "ImpType", "physical", "", "", "", "none", "", "", "", ""])

    ws2 = wb.create_sheet(title="xl_imp1")
    ws2.append(["字段标识", "显示名称", "字段类型", "最大长度",
                "默认值", "选项", "必填", "排序"])
    ws2.append(["tags", "标签", "array", "", '["a","b"]', "", "否", 0])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post("/admin/api/node-types/import",
                    files={"file": ("t.xlsx", buf.read(),
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200, r.text

    types = client.get("/admin/api/node-types").json()["data"]["items"]
    imp = next(t for t in types if t["code"] == "xl_imp1")
    fields = imp["fields"]
    assert len(fields) == 1
    assert fields[0]["fieldType"] == "array"
    assert fields[0]["defaultValue"] == '["a","b"]'


def test_excel_import_array_invalid_default_rejected(client):
    """导入非法 array default（fieldType=array, defaultValue='abc'）→ errors 收集，该字段被跳过。"""
    from io import BytesIO
    from openpyxl import Workbook

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "类型汇总"
    ws1.append(["编码", "名称", "分类", "图标", "颜色", "形状",
                "渲染模式", "DN模板", "描述", "创建时间", "更新时间"])
    ws1.append(["xl_imp2", "BadImp", "physical", "", "", "", "none", "", "", "", ""])

    ws2 = wb.create_sheet(title="xl_imp2")
    ws2.append(["字段标识", "显示名称", "字段类型", "最大长度",
                "默认值", "选项", "必填", "排序"])
    ws2.append(["tags", "标签", "array", "", "abc", "", "否", 0])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post("/admin/api/node-types/import",
                    files={"file": ("t.xlsx", buf.read(),
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200
    data = r.json()["data"]
    # 字段被跳过，errors 不为空
    assert len(data.get("errors", [])) > 0
